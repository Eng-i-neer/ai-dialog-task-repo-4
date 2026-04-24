"""Analyze pricing Excel structure: sheets, dimensions, sample rows, merged cells."""

from pathlib import Path

import openpyxl

PATH = Path(
    r"c:\Users\59571\Desktop\deutsch-app\舅妈网站\汇森国际-东欧COD报价20260331生效(5).xlsx"
)


def main() -> None:
    wb = openpyxl.load_workbook(PATH, read_only=False, data_only=True)

    print("=" * 72)
    print("ALL SHEET NAMES")
    print("=" * 72)
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("=" * 72)
        print(f"SHEET: {sheet_name}")
        print("=" * 72)

        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Column count (max_column): {max_col}")
        print(f"Row count (max_row): {max_row}")
        print()

        print("First 5 rows (values; empty cells shown as blank):")
        n = min(5, max_row) if max_row else 0
        for r in range(1, n + 1):
            cells = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    cells.append("")
                else:
                    s = str(v).replace("\n", " ")
                    if len(s) > 50:
                        s = s[:47] + "..."
                    cells.append(s)
            print(f"  Row {r}: {cells}")
        if n == 0:
            print("  (no rows)")
        print()

        merged = list(ws.merged_cells.ranges)
        print(f"Merged cell ranges: {len(merged)}")
        if merged:
            for m in sorted(merged, key=lambda x: (x.min_row, x.min_col, x.max_row, x.max_col)):
                print(f"  {m}")
        else:
            print("  (none)")
        print()

    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
