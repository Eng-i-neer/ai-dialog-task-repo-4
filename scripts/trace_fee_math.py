"""
Trace the mathematical relationship between input fee categories and template output values.
Key discovery needed: how 服务费/地派服务费/短信费 etc map to 头程运费/尾程运费 formula columns.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')

# Use 君悦 as primary example
input_path = None
for f in os.listdir(BASE / '中介提供'):
    if '-中文1-' in f and f.endswith('.xlsx'):
        input_path = BASE / '中介提供' / f
        break

template_path = BASE / '反馈客户' / '原始模板' / '汇森-君悦（东欧）对账单-20260330.xlsx'

# Parse input
wb_in = openpyxl.load_workbook(str(input_path), data_only=True)
waybill_fees = {}
for sname in wb_in.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = wb_in[sname]
    header_row = None
    for r in range(1, 15):
        v = ws.cell(r, 3).value
        if v and '运单' in str(v):
            header_row = r
            break
    if not header_row:
        continue
    for r in range(header_row + 1, ws.max_row + 1):
        wb_id = ws.cell(r, 3).value
        if not wb_id or not isinstance(wb_id, str):
            continue
        wb_id = wb_id.strip()
        amt = ws.cell(r, 13).value or 0
        charge_w = ws.cell(r, 8).value
        if wb_id not in waybill_fees:
            waybill_fees[wb_id] = {}
        waybill_fees[wb_id][sname] = {'amount': amt, 'charge_weight': charge_w}
wb_in.close()

# Parse template freight sheet - get actual values
wb_tmpl = openpyxl.load_workbook(str(template_path), data_only=True)
ws_freight = None
for sname in wb_tmpl.sheetnames:
    if '运费' in sname and '杂费' not in sname:
        ws_freight = wb_tmpl[sname]
        break

print("=" * 100)
print("运费 Sheet 逐运单对比: 模板值 vs 输入各科目金额")
print("=" * 100)
print(f"{'运单号':<30} {'模板头程':>8} {'模板尾程':>8} {'模板COD手续':>10} {'模板小计':>8} | {'输入服务费':>10} {'输入地派':>8} {'输入尾程':>8} {'输入头程':>8} {'输入短信':>8} {'输入转寄':>8} {'输入管理':>8}")
print("-" * 160)

for r in range(2, ws_freight.max_row + 1):
    wb_id = ws_freight.cell(r, 4).value
    if not wb_id:
        break
    wb_id = str(wb_id).strip()
    
    tmpl_head = ws_freight.cell(r, 13).value or 0  # 头程运费(EUR)
    tmpl_tail = ws_freight.cell(r, 14).value or 0  # 尾程运费(EUR)
    tmpl_cod_fee = ws_freight.cell(r, 15).value or 0  # COD手续费收取
    tmpl_subtotal = ws_freight.cell(r, 17).value or 0  # 小计
    
    inp = waybill_fees.get(wb_id, {})
    srv = inp.get('服务费', {}).get('amount', 0)
    
    # Sum all 地派 fees
    delivery = 0
    delivery_detail = ''
    for k, v in inp.items():
        if '地派' in k:
            delivery += v['amount']
            delivery_detail = k
    
    tail = inp.get('尾程运费', {}).get('amount', 0)
    head = inp.get('头程运费', {}).get('amount', 0)
    sms = inp.get('短信费', {}).get('amount', 0)
    forward = inp.get('转寄操作费', {}).get('amount', 0)
    mgmt = inp.get('账号管理费', {}).get('amount', 0)
    warehouse = inp.get('海外仓操作费', {}).get('amount', 0)
    agent_delivery = inp.get('代理送货费', {}).get('amount', 0)
    cod_fee = inp.get('代收COD手续费', {}).get('amount', 0)
    reject = inp.get('拒收返程费', {}).get('amount', 0)
    
    # Try to find what sums to the template values
    input_total = srv + delivery + tail + head + sms + forward + mgmt + warehouse + agent_delivery + cod_fee + reject
    
    print(f"{wb_id:<30} {tmpl_head:>8.2f} {tmpl_tail:>8.2f} {tmpl_cod_fee:>10.2f} {tmpl_subtotal:>8.2f} | {srv:>10.2f} {delivery:>8.2f} {tail:>8.2f} {head:>8.2f} {sms:>8.2f} {forward:>8.2f} {mgmt:>8.2f}")
    
    if r <= 15:
        # For first 10 rows, also show the detailed breakdown
        all_fees = {k: v['amount'] for k, v in inp.items() if v['amount'] != 0}
        print(f"  -> 输入全部: {all_fees}")
        print(f"  -> 模板: 头程={tmpl_head}, 尾程={tmpl_tail}, COD手续费={tmpl_cod_fee}, 小计={tmpl_subtotal}")
        
        # The formulas say:
        # 头程 = IFS(L="GS", I*7.7, "SC", I*9.5, "IC", I*10)
        # 尾程 = IF(K>2, 3.8+(ROUNDUP(K,0)-2)*0.8, 3.8)
        # These are FORMULA-computed, not input-derived!
        tail_weight = ws_freight.cell(r, 11).value or 0
        weight = ws_freight.cell(r, 9).value or 0
        category = ws_freight.cell(r, 12).value or ''
        print(f"  -> 模板 weight={weight}, tail_weight={tail_weight}, category={category}")

wb_tmpl.close()

# Now check F货附加费
print(f"\n\n{'='*100}")
print("F货附加费 Sheet 追踪 — 是否每条运费运单都出现在F货附加费中？")
print("=" * 100)

wb_tmpl = openpyxl.load_workbook(str(template_path), data_only=True)
freight_waybills = set()
f_waybills = set()

for sname in wb_tmpl.sheetnames:
    if '运费' in sname and '杂费' not in sname:
        ws = wb_tmpl[sname]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 4).value
            if v:
                freight_waybills.add(str(v).strip())
    if 'F' in sname and '附加' in sname:
        ws = wb_tmpl[sname]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 4).value
            if v:
                f_waybills.add(str(v).strip())

print(f"运费运单数: {len(freight_waybills)}")
print(f"F附加费运单数: {len(f_waybills)}")
print(f"运费中有但F附加费中没有: {len(freight_waybills - f_waybills)}")
print(f"F附加费中有但运费中没有: {len(f_waybills - freight_waybills)}")

if freight_waybills == f_waybills:
    print(">>> F附加费 = 运费运单（完全一致）!")
elif f_waybills.issubset(freight_waybills):
    print(">>> F附加费 ⊂ 运费运单")
    missing = freight_waybills - f_waybills
    print(f"  缺少: {list(missing)[:5]}...")

wb_tmpl.close()

# Also check 李志 template for similar pattern
print(f"\n\n{'='*100}")
print("李志 运费 Sheet 公式验证 — 哪些输入科目被公式隐含吸收？")
print("=" * 100)
print("""
李志 运费公式:
  C14 头程运费 = ROUNDUP(IFS(M="GS",J*7.7,"SC",J*9.5,"IC",J*10)*7.9342,2)
      → 只用了 J(头程计费重量) 和 M(普特敏货) ← 这是客户报价，不是代理成本!
  C15 尾程运费 = ROUNDUP(IF(L>2,8+(ROUNDUP(L,0)-2)*1.5,8)*7.9342,2)
      → 只用了 L(尾程计费重) ← 也是客户报价
  C16 代收手续费 = ROUND(7*7.9342,2)
      → 固定 7 EUR * 汇率 ← 客户向终端收取的手续费
  C17 附加费 = ROUNDUP(1.5*7.9342,2)
      → 固定 1.5 EUR * 汇率 ← 转寄附加费

关键发现: 运费Sheet的公式是按【客户报价】计算的，不是按代理输入的实际成本!
代理输入的 服务费/地派服务费/短信费 等是【代理成本】，不直接对应客户运费Sheet的任何列。
""")

# Verify: 李志模板
template_lz = BASE / '反馈客户' / '原始模板' / '20260330-汇森李志（东欧）对账单.xlsx'
wb_lz = openpyxl.load_workbook(str(template_lz), data_only=True)
ws_lz = wb_lz.worksheets[1]
input_lz = None
for f in os.listdir(BASE / '中介提供'):
    if '-中文-' in f and f.endswith('.xlsx'):
        input_lz = BASE / '中介提供' / f

wb_in_lz = openpyxl.load_workbook(str(input_lz), data_only=True)
fees_lz = {}
for sname in wb_in_lz.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = wb_in_lz[sname]
    hr = None
    for r in range(1, 15):
        v = ws.cell(r, 3).value
        if v and '运单' in str(v):
            hr = r
            break
    if not hr:
        continue
    for r in range(hr + 1, ws.max_row + 1):
        wb_id = ws.cell(r, 3).value
        if not wb_id or not isinstance(wb_id, str):
            continue
        wb_id = wb_id.strip()
        amt = ws.cell(r, 13).value or 0
        if wb_id not in fees_lz:
            fees_lz[wb_id] = {}
        fees_lz[wb_id][sname] = amt
wb_in_lz.close()

# Pick a sample from 李志
sample_wb = str(ws_lz.cell(2, 4).value or '')
if sample_wb in fees_lz:
    print(f"\n李志 样本运单: {sample_wb}")
    print(f"  输入科目: {fees_lz[sample_wb]}")
    print(f"  模板运费行: 头程={ws_lz.cell(2,14).value}, 尾程={ws_lz.cell(2,15).value}, "
          f"COD手续费={ws_lz.cell(2,16).value}, 附加费={ws_lz.cell(2,17).value}, 小计={ws_lz.cell(2,19).value}")

wb_lz.close()
