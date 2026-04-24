"""验证 HR12603041510002 的上架费"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

GEN = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\20260330-汇森李志（东欧）对账单-自动生成v4.xlsx'
wb = openpyxl.load_workbook(GEN, data_only=False)
ws = wb['20260330期尾程杂费']

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 2).value or '') == 'HR12603041510002':
        print(f"HR12603041510002 在 Row {r}")
        print(f"  C12 (上架费): {repr(ws.cell(r, 12).value)}")
        print(f"  C13 (退件费): {repr(ws.cell(r, 13).value)}")
        print(f"  C15 (增值税): {repr(ws.cell(r, 15).value)}")
        print(f"  C17 (小计):   {repr(ws.cell(r, 17).value)}")
        print(f"  C18 (备注):   {repr(ws.cell(r, 18).value)}")
        break
wb.close()
