"""
Deep verification of remaining categories:
- 退件费 detailed per-country check (including 70% rule verification)
- VAT sheet structure for 李志 and 小美 (they had no separate VAT sheet)
- F附加费 detailed value check
- Verify all sheets across all customers
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

print("=" * 120)
print("PART 1: 所有客户的 Sheet 结构一览")
print("=" * 120)

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=True)
    print(f"\n  {customer}: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        data_rows = 0
        for r in range(2, ws.max_row + 1):
            has_data = False
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if v and '合计' not in str(v):
                    has_data = True
                    break
            if has_data:
                data_rows += 1
            else:
                break
        print(f"    [{sname}] {data_rows}行数据")
    wb.close()


print(f"\n\n{'='*120}")
print("PART 2: 退件费详细 — 逐客户逐行公式 + 70%规则验证")
print("=" * 120)

OLD_TAIL = {
    '波兰':     {'first2': 3.8, 'extra1': 0.8, 'rule': '70%'},
    '罗马尼亚':  {'first2': 4.8, 'extra1': 0.6, 'rule': '70%'},
    '匈牙利':   {'first2': 4.1, 'extra1': 0.6, 'rule': '70%'},
    '捷克':     {'first2': 3.8, 'extra1': 0.6, 'rule': '70%'},
    '斯洛伐克':  {'first2': 4.0, 'extra1': 0.6, 'rule': '70%'},
    '保加利亚':  {'first2': 4.3, 'extra1': 0.8, 'rule': '70%'},
    '克罗地亚':  {'first2': 5.6, 'extra1': 0.9, 'rule': '70%'},
    '斯洛文尼亚': {'first2': 5.2, 'extra1': 0.8, 'rule': '70%'},
    '西班牙':   {'first2': 4.0, 'extra1': 1.0, 'rule': '100%'},
    '葡萄牙':   {'first2': 4.0, 'extra1': 1.0, 'rule': '100%'},
    '希腊':     {'first2': 5.7, 'extra1': 0.8, 'rule': '100%'},
    '意大利':   {'first2': 6.7, 'extra1': 1.0, 'rule': '100%'},
    '奥地利':   {'first2': 6.5, 'extra1': 1.0, 'rule': '100%'},
    '德国':     {'first2': 8.0, 'extra1': 1.5, 'rule': '100%'},
}

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    found = False
    for sname in wb_f.sheetnames:
        if '杂费' not in sname:
            continue

        ws_f = wb_f[sname]
        ws_v = wb_v[sname]

        headers = {}
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h:
                headers[str(h).strip()] = c

        return_col = shelf_col = vat_col = dest_col = weight_col = None
        for h, c in headers.items():
            if '退件' in h and ('入仓' in h or 'RMB' in h or 'EUR' in h):
                return_col = c
            if '上架' in h:
                shelf_col = c
            if '增值税' in h or 'VAT' in h:
                vat_col = c
            if '目的' in h:
                dest_col = c
            if '计费' in h and '重' in h:
                weight_col = c

        if return_col is None:
            continue

        found = True
        print(f"\n  {customer} [{sname}] 退件费=C{return_col}, 上架费=C{shelf_col}, VAT=C{vat_col}, 目的地=C{dest_col}")
        print(f"    Headers: {headers}")

        country_formulas = {}
        for r in range(2, ws_f.max_row + 1):
            dest = ws_v.cell(r, dest_col).value if dest_col else None
            formula = ws_f.cell(r, return_col).value
            if dest is None and formula is None:
                break
            d = str(dest or '?').strip()
            if '合计' in d:
                break
            if d not in country_formulas and formula and isinstance(formula, str) and formula.startswith('='):
                country_formulas[d] = (r, formula)

        for country, (row, formula) in sorted(country_formulas.items()):
            tail = OLD_TAIL.get(country)
            has_70 = '70%' in formula or '*0.7' in formula

            m = re.search(r'(\d+\.?\d*)\s*\+\s*\(.*?\)\s*\*\s*(\d+\.?\d*)', formula)
            f1 = float(m.group(1)) if m else None
            e1 = float(m.group(2)) if m else None

            checks = []
            if tail:
                if f1 == tail['first2'] and e1 == tail['extra1']:
                    checks.append('费率✓')
                else:
                    checks.append(f'费率✗(预期{tail["first2"]}/{tail["extra1"]} 实际{f1}/{e1})')

                expected_70 = tail['rule'] == '70%'
                if has_70 == expected_70:
                    checks.append(f'规则{tail["rule"]}✓')
                else:
                    checks.append(f'规则✗(预期{tail["rule"]} 实际{"70%" if has_70 else "100%"})')
            else:
                checks.append(f'费率={f1}/{e1}')
                checks.append(f'含70%={"是" if has_70 else "否"}')

            print(f"    {country}(R{row}): {' | '.join(checks)}")
            print(f"      公式: {formula[:90]}")

    if not found:
        print(f"\n  {customer}: 无杂费sheet/退件费公式")

    wb_f.close()
    wb_v.close()


print(f"\n\n{'='*120}")
print("PART 3: F附加费 — 详细金额提取")
print("=" * 120)

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    found = False
    for sname in wb_f.sheetnames:
        if 'F' not in sname:
            continue
        if '附加' not in sname and '加费' not in sname:
            continue

        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        found = True

        headers = {}
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h:
                headers[str(h).strip()] = c

        print(f"\n  {customer} [{sname}]")
        print(f"    Headers: {headers}")

        for r in range(2, min(ws_f.max_row + 1, 25)):
            row_data = []
            for c in range(1, ws_f.max_column + 1):
                fv = ws_f.cell(r, c).value
                vv = ws_v.cell(r, c).value
                h = ws_f.cell(1, c).value or f'C{c}'
                if fv is not None:
                    if isinstance(fv, str) and fv.startswith('='):
                        row_data.append(f"{h}=[{fv[:30]}]")
                    else:
                        row_data.append(f"{h}={fv}")
                elif vv is not None:
                    row_data.append(f"{h}=v:{vv}")
            if row_data:
                print(f"    R{r}: {', '.join(row_data)}")

    if not found:
        print(f"\n  {customer}: 无F附加费sheet")

    wb_f.close()
    wb_v.close()


print(f"\n\n{'='*120}")
print("PART 4: 李志/小美 VAT — 检查杂费sheet中的VAT列 或 独立VAT sheet")
print("=" * 120)

for customer in ['李志', '小美']:
    fname = CUSTOMERS[customer]
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    for sname in wb_f.sheetnames:
        ws_f = wb_f[sname]
        has_vat = False
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h and ('增值税' in str(h) or 'VAT' in str(h)):
                has_vat = True
                break
        if has_vat or '增值税' in sname or 'VAT' in sname:
            print(f"\n  {customer} [{sname}]:")
            for c in range(1, ws_f.max_column + 1):
                h = ws_f.cell(1, c).value
                if h:
                    print(f"    C{c}={h}", end='')
            print()
            for r in range(2, min(ws_f.max_row + 1, 6)):
                vals = []
                for c in range(1, ws_f.max_column + 1):
                    fv = ws_f.cell(r, c).value
                    if fv:
                        vals.append(f"C{c}={fv}")
                if vals:
                    print(f"    R{r}: {', '.join(vals)}")

    wb_f.close()
    wb_v.close()
