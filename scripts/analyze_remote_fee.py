"""分析偏远费的规则 - 邮编来源和费率"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'

tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
inp_wb = openpyxl.load_workbook(INPUT, data_only=True)

ws_sur = tmpl_wb['20260330期尾程杂费']

# 收集所有偏远费行
remote_rows = []
for r in range(2, ws_sur.max_row + 1):
    fee = ws_sur.cell(r, 16).value  # C16 偏远费
    if fee:
        remote_rows.append({
            'row': r,
            'waybill': ws_sur.cell(r, 2).value,
            'dest': ws_sur.cell(r, 10).value,
            'fee': fee,
            'zipcode': ws_sur.cell(r, 19).value,  # C19
            'region': ws_sur.cell(r, 20).value,    # C20
        })

print(f"偏远费行数: {len(remote_rows)}")

# 按地区分组
by_region = {}
for item in remote_rows:
    region = str(item['region'] or 'None')
    if region not in by_region:
        by_region[region] = {'fee': item['fee'], 'count': 0, 'zips': set()}
    by_region[region]['count'] += 1
    by_region[region]['zips'].add(str(item['zipcode']))

for region, info in sorted(by_region.items()):
    print(f"\n  地区: {region}")
    print(f"    费用: {info['fee']} RMB")
    print(f"    数量: {info['count']}")
    print(f"    邮编: {sorted(info['zips'])[:20]}")
    if info['fee']:
        eur = info['fee'] / 7.9342
        print(f"    反推EUR: {eur:.4f}")

# 查输入文件中这些运单的邮编来源
# 邮编可能在输入的 "尾程运费" 或 "IT地派服务费" sheet中
print("\n\n检查输入文件中的邮编信息:")
for sname in inp_wb.sheetnames:
    ws = inp_wb[sname]
    # 检查header中是否有邮编相关列
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and '邮编' in str(v):
                print(f"  Sheet '{sname}' R{r} C{c}: {v}")
            if v and '地址' in str(v):
                print(f"  Sheet '{sname}' R{r} C{c}: {v}")
            if v and '城市' in str(v):
                print(f"  Sheet '{sname}' R{r} C{c}: {v}")

# 查找IT地派服务费sheet的完整header
print("\n\nIT地派服务费 sheet结构:")
for sname in inp_wb.sheetnames:
    if 'IT' in sname:
        ws = inp_wb[sname]
        for r in range(1, 15):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v:
                    vals.append(f"C{c}={v}")
            if vals:
                print(f"  R{r}: {vals}")

# 查找一个具体的偏远费运单的所有数据
print("\n\n示例运单 IT12603181810028 在各sheet中的数据:")
target = 'IT12603181810028'
for sname in inp_wb.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = inp_wb[sname]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip() == target:
                vals = []
                for cc in range(1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if vv is not None:
                        vals.append(f"C{cc}={vv}")
                print(f"  Sheet '{sname}' R{r}: {vals}")
                break

tmpl_wb.close()
inp_wb.close()
