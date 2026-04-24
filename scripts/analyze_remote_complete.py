"""完整分析偏远费：报价表偏远邮编 + 输入文件邮编来源 + 费率计算"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

# 1. 从报价表读取偏远邮编
pricing_wb = openpyxl.load_workbook(PRICING, data_only=True)
ws_remote = pricing_wb['偏远邮编']

print("="*80)
print("报价表 - 偏远邮编表结构")
print("="*80)
# Header
for c in range(1, ws_remote.max_column + 1):
    v = ws_remote.cell(1, c).value
    print(f"  C{c}: {v}")

# 建立邮编->地区映射
remote_zips = {}  # zip_code -> {'country': ..., 'region': ...}
for r in range(2, ws_remote.max_row + 1):
    # Col1-3: 克罗地亚
    z1 = ws_remote.cell(r, 1).value
    c2 = ws_remote.cell(r, 2).value
    c3 = ws_remote.cell(r, 3).value
    if z1:
        country = str(c2 or '').strip()
        region = str(c3 or '').strip()
        remote_zips[str(z1).strip()] = {'country': country or '克罗地亚', 'region': region or '克罗地亚'}
    
    # Col4-5: 意大利偏远邮编(西西里岛)
    z4 = ws_remote.cell(r, 4).value
    c5 = ws_remote.cell(r, 5).value
    if z4:
        remote_zips[str(z4).strip()] = {'country': '意大利', 'region': str(c5 or '西西里岛').strip()}
    
    # Col6-7: 意大利偏远邮编(撒丁岛)
    z6 = ws_remote.cell(r, 6).value
    c7 = ws_remote.cell(r, 7).value
    if z6:
        remote_zips[str(z6).strip()] = {'country': '意大利', 'region': str(c7 or '撒丁岛').strip()}
    
    # Col8-9: 意大利偏远邮编(其他岛屿)
    z8 = ws_remote.cell(r, 8).value
    c9 = ws_remote.cell(r, 9).value
    if z8:
        remote_zips[str(z8).strip()] = {'country': '意大利', 'region': str(c9 or '其他岛屿').strip()}

print(f"\n总偏远邮编数: {len(remote_zips)}")

# 按地区统计
region_stats = {}
for z, info in remote_zips.items():
    r = info['region']
    if r not in region_stats:
        region_stats[r] = 0
    region_stats[r] += 1
for r, cnt in sorted(region_stats.items()):
    print(f"  {r}: {cnt} 个邮编")

# 2. 报价表的偏远费率
print("\n" + "="*80)
print("报价表 - 偏远费率")
print("="*80)
ws_price = pricing_wb['欧洲COD']
# Row 54 有偏远费率
for r in range(54, 56):
    for c in range(1, 12):
        v = ws_price.cell(r, c).value
        if v:
            val = str(v).replace('\n', '\\n')[:120]
            print(f"  R{r} C{c}: {val}")

# 3. 从输入文件查找邮编信息
print("\n" + "="*80)
print("输入文件 - 搜索邮编/地址信息")
print("="*80)
inp_wb = openpyxl.load_workbook(INPUT, data_only=True)

# 搜索所有sheet的header中是否有邮编/地址相关列
for sname in inp_wb.sheetnames:
    ws = inp_wb[sname]
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and any(kw in str(v) for kw in ['邮编', '邮政', 'zip', 'postal', '地址', 'address', '城市', 'city', '收件']):
                print(f"  Sheet '{sname}' R{r} C{c}: {v}")

# 搜索 "备注" 或 "指定路线" 列看看有没有邮编信息
# 检查尾程运费 sheet 的 C15 备注列
print("\n尾程运费 sheet 备注/其他列:")
ws_tail = inp_wb['尾程运费']
for r in range(9, 10):
    for c in range(1, ws_tail.max_column + 1):
        v = ws_tail.cell(r, c).value
        if v:
            print(f"  Header R{r} C{c}: {v}")

# 看一些IT运单的完整数据
target_wbs = ['IT12603181810028', 'IT12603161910046', 'IT12603181810007']
print(f"\n偏远费IT运单在所有sheet中的数据:")
for sname in inp_wb.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = inp_wb[sname]
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v and str(v).strip() in target_wbs:
                all_vals = []
                for cc in range(1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if vv is not None:
                        all_vals.append(f"C{cc}={vv}")
                print(f"  '{sname}' R{r}: {all_vals}")
                break

# 4. 对比模板中的偏远费运单
print("\n" + "="*80)
print("模板偏远费验证 - 邮编匹配报价表")
print("="*80)
tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
ws_sur = tmpl_wb['20260330期尾程杂费']

for r in range(2, ws_sur.max_row + 1):
    fee = ws_sur.cell(r, 16).value
    if not fee:
        continue
    zipcode = str(ws_sur.cell(r, 19).value or '')
    region = ws_sur.cell(r, 20).value
    waybill = ws_sur.cell(r, 2).value
    weight = ws_sur.cell(r, 7).value
    
    # 查报价表
    matched = remote_zips.get(zipcode)
    
    # 反推费率
    eur = fee / 7.9342 if fee else 0
    per_kg = eur / weight if weight and weight > 0 else 0
    
    if r <= 470 and r >= 440:  # 只打印部分
        print(f"  R{r}: {waybill} | 邮编={zipcode} | 地区={region} | "
              f"重量={weight}KG | 偏远费={fee}RMB | EUR={eur:.4f} | "
              f"EUR/KG={per_kg:.4f} | "
              f"报价匹配={matched}")

# 汇总
print(f"\n费率汇总:")
by_region2 = {}
for r in range(2, ws_sur.max_row + 1):
    fee = ws_sur.cell(r, 16).value
    if not fee:
        continue
    region = str(ws_sur.cell(r, 20).value or '')
    weight = ws_sur.cell(r, 7).value or 0
    eur = fee / 7.9342
    per_kg = eur / weight if weight > 0 else 0
    if region not in by_region2:
        by_region2[region] = []
    by_region2[region].append({'fee': fee, 'eur': eur, 'weight': weight, 'per_kg': per_kg})

for region, items in sorted(by_region2.items()):
    rates = set(round(i['per_kg'], 2) for i in items)
    print(f"  {region}: {len(items)}条, EUR/KG费率={rates}")
    print(f"    报价表: 卡拉布里亚/西西里岛/撒丁岛 +0.8EUR/KG, 其他岛屿 +21EUR")

pricing_wb.close()
inp_wb.close()
tmpl_wb.close()
