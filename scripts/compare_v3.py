"""对比基于模板生成的文件 vs 参考文件 — 只检查相同数据字段"""
import openpyxl
import glob

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

# Find the latest generated file
import os
desktop = os.path.expanduser('~/Desktop')
gen_files = sorted(
    [f for f in os.listdir(desktop) if f.startswith('20260330-result-')],
    reverse=True
)
GEN_FILE = os.path.join(desktop, gen_files[0])
print(f"Comparing: {GEN_FILE}")

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
gen = openpyxl.load_workbook(GEN_FILE, data_only=True)

# ── COD Sheet ──
ws_ref = ref[ref.sheetnames[0]]
ws_gen = gen[gen.sheetnames[0]]

ref_cod = {}
for r in range(2, ws_ref.max_row + 1):
    wb = ws_ref.cell(r, 4).value
    if wb and isinstance(wb, str) and len(wb) > 5:
        ref_cod[wb] = {c: ws_ref.cell(r, c).value for c in range(1, 15)}

gen_cod = {}
for r in range(2, ws_gen.max_row + 1):
    wb = ws_gen.cell(r, 4).value
    if wb and isinstance(wb, str) and len(wb) > 5:
        gen_cod[wb] = {c: ws_gen.cell(r, c).value for c in range(1, 15)}

common = set(ref_cod) & set(gen_cod)
print(f"\nCOD: ref={len(ref_cod)}, gen={len(gen_cod)}, common={len(common)}")

field_names = {
    1: 'customer_code', 2: 'ship_type', 3: 'date', 4: 'waybill',
    5: 'cust_ref', 6: 'transfer_no', 7: 'dest', 8: 'category',
    9: 'eur_amount',
}

for col, name in field_names.items():
    match = 0
    for wb in common:
        rv = ref_cod[wb][col]
        gv = gen_cod[wb][col]
        if rv == gv:
            match += 1
        elif isinstance(rv, (int, float)) and isinstance(gv, (int, float)) and abs(rv - gv) < 0.1:
            match += 1
    print(f"  {name:15s}: {match}/{len(common)} ({match/len(common)*100:.1f}%)")

# ── Freight Sheet ──
ws_ref_f = ref[ref.sheetnames[1]]
ws_gen_f = gen[gen.sheetnames[1]]

ref_fr = {}
for r in range(2, ws_ref_f.max_row + 1):
    wb = ws_ref_f.cell(r, 4).value
    if wb and isinstance(wb, str):
        ref_fr[wb] = {c: ws_ref_f.cell(r, c).value for c in range(1, 21)}

gen_fr = {}
for r in range(2, ws_gen_f.max_row + 1):
    wb = ws_gen_f.cell(r, 4).value
    if wb and isinstance(wb, str):
        gen_fr[wb] = {c: ws_gen_f.cell(r, c).value for c in range(1, 21)}

common_f = set(ref_fr) & set(gen_fr)
print(f"\nFreight: ref={len(ref_fr)}, gen={len(gen_fr)}, common={len(common_f)}")

freight_fields = {
    3: 'date', 4: 'waybill', 5: 'cust_ref', 6: 'transfer_no',
    7: 'dest', 8: 'category', 9: 'actual_weight',
    10: 'head_weight', 11: 'dimensions', 12: 'tail_qty',
}
for col, name in freight_fields.items():
    match = 0
    for wb in common_f:
        rv = ref_fr[wb][col]
        gv = gen_fr[wb][col]
        if rv == gv:
            match += 1
        elif isinstance(rv, (int, float)) and isinstance(gv, (int, float)) and abs(rv - gv) < 0.01:
            match += 1
    print(f"  {name:15s}: {match}/{len(common_f)} ({match/len(common_f)*100:.1f}%)")

print(f"\nSheet names preserved: {gen.sheetnames}")
