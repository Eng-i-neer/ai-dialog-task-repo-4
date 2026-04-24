"""
批量差异对比脚本 — 对比每个客户的自动生成文件 vs 原始模板
"""
import sys, io, os, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from datetime import datetime
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'
OUTPUT_BASE = BASE / '反馈客户' / '自动生成'

CUSTOMERS = ['李志', '君悦', '小美', 'J', '涵江', '阿甘', '威总', '峰总']


def find_file(directory, keyword):
    for f in os.listdir(directory):
        if f.endswith('.xlsx') and keyword in f:
            return directory / f
    return None


def normalize_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v.startswith('='):
            return '[FORMULA]'
        if v == '':
            return None
        return v
    if isinstance(v, float):
        return round(v, 4)
    return v


def compare_sheets(ws_template, ws_gen, sheet_name, max_compare_rows=600):
    """Compare two sheets cell by cell, returning structured diff."""
    result = {
        'sheet_name': sheet_name,
        'template_rows': 0,
        'gen_rows': 0,
        'header_match': True,
        'header_diffs': [],
        'row_diffs': [],
        'missing_in_gen': [],
        'extra_in_gen': [],
    }

    # Compare headers
    max_col = max(ws_template.max_column or 1, ws_gen.max_column or 1)
    for c in range(1, max_col + 1):
        th = ws_template.cell(1, c).value
        gh = ws_gen.cell(1, c).value
        if str(th or '').strip() != str(gh or '').strip():
            result['header_match'] = False
            result['header_diffs'].append({
                'col': c,
                'template': str(th or ''),
                'generated': str(gh or ''),
            })

    # Count data rows in template
    t_waybills = {}
    for r in range(2, min(ws_template.max_row + 1, max_compare_rows)):
        vals = [ws_template.cell(r, c).value for c in range(1, min(max_col + 1, 10))]
        if all(v is None for v in vals):
            break
        if any('合计' in str(v or '') for v in vals):
            break
        result['template_rows'] += 1
        wb_col = 4 if ws_template.cell(1, 4).value and '运单' in str(ws_template.cell(1, 4).value or '') else 2
        wb_val = ws_template.cell(r, wb_col).value
        if wb_val:
            t_waybills[str(wb_val).strip()] = r

    # Count data rows in generated
    g_waybills = {}
    for r in range(2, min(ws_gen.max_row + 1, max_compare_rows)):
        vals = [ws_gen.cell(r, c).value for c in range(1, min(max_col + 1, 10))]
        if all(v is None for v in vals):
            break
        if any('合计' in str(v or '') for v in vals):
            break
        result['gen_rows'] += 1
        wb_col = 4 if ws_gen.cell(1, 4).value and '运单' in str(ws_gen.cell(1, 4).value or '') else 2
        wb_val = ws_gen.cell(r, wb_col).value
        if wb_val:
            g_waybills[str(wb_val).strip()] = r

    # Find missing/extra waybills
    t_set = set(t_waybills.keys())
    g_set = set(g_waybills.keys())
    result['missing_in_gen'] = sorted(t_set - g_set)
    result['extra_in_gen'] = sorted(g_set - t_set)

    # Compare matched waybills cell by cell
    common = t_set & g_set
    diff_count = 0
    for wb_id in sorted(common):
        tr = t_waybills[wb_id]
        gr = g_waybills[wb_id]
        for c in range(1, max_col + 1):
            tv = normalize_value(ws_template.cell(tr, c).value)
            gv = normalize_value(ws_gen.cell(gr, c).value)
            if tv != gv:
                if tv == '[FORMULA]' or gv == '[FORMULA]':
                    continue
                diff_count += 1
                if diff_count <= 20:
                    result['row_diffs'].append({
                        'waybill': wb_id,
                        'col': c,
                        'col_header': str(ws_template.cell(1, c).value or f'C{c}'),
                        'template_val': str(tv),
                        'gen_val': str(gv),
                    })

    result['total_cell_diffs'] = diff_count
    return result


def diff_customer(customer):
    """Diff generated vs template for one customer."""
    template_path = find_file(TEMPLATE_DIR, customer)
    gen_path = find_file(OUTPUT_BASE / customer, customer)

    if not template_path or not gen_path:
        return {'customer': customer, 'error': f'Files not found: template={template_path}, gen={gen_path}'}

    wb_t = openpyxl.load_workbook(str(template_path), data_only=True)
    wb_g = openpyxl.load_workbook(str(gen_path), data_only=True)

    result = {
        'customer': customer,
        'template_file': template_path.name,
        'gen_file': gen_path.name,
        'template_sheets': wb_t.sheetnames,
        'gen_sheets': wb_g.sheetnames,
        'sheet_count_match': len(wb_t.sheetnames) == len(wb_g.sheetnames),
        'sheet_diffs': [],
    }

    for sname in wb_t.sheetnames:
        if sname not in wb_g.sheetnames:
            result['sheet_diffs'].append({
                'sheet_name': sname, 'error': 'Sheet missing in generated file',
            })
            continue

        ws_t = wb_t[sname]
        ws_g = wb_g[sname]
        diff = compare_sheets(ws_t, ws_g, sname)
        result['sheet_diffs'].append(diff)

    wb_t.close()
    wb_g.close()
    return result


def main():
    all_results = {}
    summary_lines = []

    for customer in CUSTOMERS:
        print(f"\n{'='*60}")
        print(f"Diffing: {customer}")

        result = diff_customer(customer)
        all_results[customer] = result

        if 'error' in result:
            print(f"  ERROR: {result['error']}")
            summary_lines.append(f"  {customer}: ERROR - {result['error']}")
            continue

        print(f"  Template: {result['template_file']} ({len(result['template_sheets'])} sheets)")
        print(f"  Generated: {result['gen_file']} ({len(result['gen_sheets'])} sheets)")

        customer_issues = []
        for sd in result['sheet_diffs']:
            if 'error' in sd:
                print(f"  [{sd['sheet_name']}] {sd['error']}")
                customer_issues.append(f"{sd['sheet_name']}: {sd['error']}")
                continue

            t_rows = sd['template_rows']
            g_rows = sd['gen_rows']
            match_str = 'OK' if t_rows == g_rows else f'MISMATCH ({t_rows} vs {g_rows})'

            missing = len(sd.get('missing_in_gen', []))
            extra = len(sd.get('extra_in_gen', []))
            cell_diffs = sd.get('total_cell_diffs', 0)

            status = 'OK'
            issues = []
            if t_rows != g_rows:
                issues.append(f'rows: {t_rows}→{g_rows}')
            if missing > 0:
                issues.append(f'{missing} missing')
            if extra > 0:
                issues.append(f'{extra} extra')
            if cell_diffs > 0:
                issues.append(f'{cell_diffs} cell diffs')
            if not sd['header_match']:
                issues.append('header mismatch')

            status = ', '.join(issues) if issues else 'PERFECT'
            print(f"  [{sd['sheet_name']}] rows: {t_rows}/{g_rows}, {status}")

            if sd.get('header_diffs'):
                for hd in sd['header_diffs'][:3]:
                    print(f"    Header C{hd['col']}: '{hd['template']}' vs '{hd['generated']}'")

            if sd.get('missing_in_gen') and missing <= 5:
                print(f"    Missing: {sd['missing_in_gen']}")
            if sd.get('extra_in_gen') and extra <= 5:
                print(f"    Extra: {sd['extra_in_gen']}")
            if sd.get('row_diffs'):
                for rd in sd['row_diffs'][:5]:
                    print(f"    Diff [{rd['waybill']}] C{rd['col']}({rd['col_header']}): "
                          f"'{rd['template_val']}' vs '{rd['gen_val']}'")

            if issues:
                customer_issues.append(f"{sd['sheet_name']}: {status}")

        summary_lines.append(f"  {customer}: {len(customer_issues)} issues" if customer_issues else f"  {customer}: PERFECT")

    # Save detailed JSON
    report_path = BASE / '反馈客户' / '自动生成' / 'diff_report.json'
    with open(str(report_path), 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n\n{'='*60}")
    print("SUMMARY OF ALL CUSTOMERS")
    print(f"{'='*60}")
    for line in summary_lines:
        print(line)
    print(f"\nDetailed report saved to: {report_path}")


if __name__ == '__main__':
    main()
