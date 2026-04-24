import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
OUTPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
out = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)

ws_in = inp['COD']
ws_out = out[out.sheetnames[0]]

in_data = {}
for r in range(10, 198):
    wb = ws_in.cell(r, 3).value
    if wb:
        in_data[wb] = {
            'date': ws_in.cell(r, 2).value,
            'waybill': wb,
            'transfer_no': ws_in.cell(r, 4).value,
            'route': ws_in.cell(r, 5).value,
            'dest': ws_in.cell(r, 11).value,
            'amount_eur': ws_in.cell(r, 13).value,
            'cust_ref': ws_in.cell(r, 16).value,
            'category': ws_in.cell(r, 17).value,
        }

print("=== FIELD MAPPING: Input COD -> Output COD ===\n")
count = 0
for r in range(2, ws_out.max_row + 1):
    out_wb = ws_out.cell(r, 4).value
    if out_wb and out_wb in in_data:
        d = in_data[out_wb]
        o = {
            'custno': ws_out.cell(r, 1).value,
            'type': ws_out.cell(r, 2).value,
            'date': ws_out.cell(r, 3).value,
            'waybill': ws_out.cell(r, 4).value,
            'cust_ref': ws_out.cell(r, 5).value,
            'track': ws_out.cell(r, 6).value,
            'dest': ws_out.cell(r, 7).value,
            'cat': ws_out.cell(r, 8).value,
            'amt': ws_out.cell(r, 9).value,
            'rate': ws_out.cell(r, 10).value,
            'cny': ws_out.cell(r, 11).value,
            'diff': ws_out.cell(r, 12).value,
            'total': ws_out.cell(r, 13).value,
        }
        print(f"--- Row {r} (waybill: {out_wb}) ---")
        print(f"  OUT A(客号)={o['custno']}  <- FIXED constant")
        print(f"  OUT B(type)={o['type']}  <- IN route={d['route']}")
        print(f"  OUT C(date)={o['date']}  <- IN date={d['date']}  MATCH={o['date']==d['date']}")
        print(f"  OUT D(waybill)={o['waybill']}  <- IN waybill  MATCH=True")
        print(f"  OUT E(cust_ref)={o['cust_ref']}  <- IN cust_ref={d['cust_ref']}  MATCH={o['cust_ref']==d['cust_ref']}")
        print(f"  OUT F(track)={o['track']}  <- IN transfer={d['transfer_no']}  MATCH={o['track']==d['transfer_no']}")
        print(f"  OUT G(dest)={o['dest']}  <- IN dest={d['dest']}  MATCH={o['dest']==d['dest']}")
        print(f"  OUT H(cat)={o['cat']}  <- IN cat={d['category']}  MATCH={o['cat']==d['category']}")
        print(f"  OUT I(EUR)={o['amt']}  <- IN EUR={d['amount_eur']}  MATCH={o['amt']==d['amount_eur']}")
        print(f"  OUT J(rate)={o['rate']}  <- EXTERNAL (exchange rate)")
        if o['amt'] and o['rate']:
            calc = round(o['amt'] * o['rate'], 2)
            print(f"  OUT K(CNY)={o['cny']}  <- CALC: {o['amt']}*{o['rate']}={calc}")
        print(f"  OUT L(diff)={o['diff']}  <- CALCULATED")
        if o['cny'] and o['diff']:
            print(f"  OUT M(total)={o['total']}  <- CALC: {o['cny']}-{o['diff']}={round(o['cny']-o['diff'],2)}")
        print()
        count += 1
        if count >= 5:
            break

# Now check route -> ship_type mapping
print("\n=== ROUTE -> SHIP TYPE MAPPING ===")
route_type = {}
for r in range(2, ws_out.max_row + 1):
    out_wb = ws_out.cell(r, 4).value
    out_type = ws_out.cell(r, 2).value
    if out_wb and out_wb in in_data and out_type:
        route = in_data[out_wb]['route']
        key = (route, out_type)
        route_type[key] = route_type.get(key, 0) + 1

for k, v in sorted(route_type.items()):
    print(f"  route={k[0]} -> type={k[1]}: {v} records")

# Check output rows NOT in input COD (historical)
print("\n=== OUTPUT ROWS NOT IN INPUT COD ===")
for r in range(2, ws_out.max_row + 1):
    out_wb = ws_out.cell(r, 4).value
    if out_wb and out_wb not in in_data:
        print(f"  Row {r}: waybill={out_wb}, type={ws_out.cell(r,2).value}, date={ws_out.cell(r,3).value}, note={ws_out.cell(r,14).value}")
