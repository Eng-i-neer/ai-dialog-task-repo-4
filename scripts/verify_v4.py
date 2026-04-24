"""验证v4生成文件：表头不变、公式保留、日期格式不变"""
import openpyxl
import os

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
desktop = os.path.expanduser('~/Desktop')
gen_files = sorted([f for f in os.listdir(desktop) if f.startswith('20260330-v4-')], reverse=True)
GEN_FILE = os.path.join(desktop, gen_files[0])
print(f"Checking: {gen_files[0]}")

tmpl = openpyxl.load_workbook(TEMPLATE)
gen = openpyxl.load_workbook(GEN_FILE)

for si in range(len(tmpl.sheetnames)):
    ts = tmpl[tmpl.sheetnames[si]]
    gs = gen[gen.sheetnames[si]]
    print(f"\n{'='*60}")
    print(f"  Sheet: {tmpl.sheetnames[si]}")

    # 1. Headers identical?
    header_ok = True
    for c in range(1, max(ts.max_column, gs.max_column) + 1):
        if ts.cell(1, c).value != gs.cell(1, c).value:
            header_ok = False
            print(f"  !! HEADER Col {c}: template={ts.cell(1,c).value} vs gen={gs.cell(1,c).value}")
    if header_ok:
        print(f"  Headers: OK (identical)")

    # 2. Date format preserved?
    for c in range(1, ts.max_column + 1):
        tnf = ts.cell(2, c).number_format
        gnf = gs.cell(2, c).number_format
        if tnf != gnf:
            print(f"  !! FORMAT Col {c} Row 2: template='{tnf}' vs gen='{gnf}'")

    # 3. Formulas preserved? (check row 4 which had formulas in template)
    formula_cols = []
    for c in range(1, ts.max_column + 1):
        for r in [4, 5, 6]:
            tv = ts.cell(r, c).value
            gv = gs.cell(r, c).value
            if tv and isinstance(tv, str) and tv.startswith('='):
                formula_cols.append(c)
                if gv and isinstance(gv, str) and gv.startswith('='):
                    # Compare formula structure (ignore row numbers)
                    import re
                    t_norm = re.sub(r'(?<=[A-Z])\d+', 'N', tv)
                    g_norm = re.sub(r'(?<=[A-Z])\d+', 'N', gv)
                    if t_norm != g_norm:
                        print(f"  !! FORMULA Col {c} Row {r}: template={tv[:50]} vs gen={gv[:50]}")
                    break
                else:
                    print(f"  !! FORMULA LOST Col {c} Row {r}: template={tv[:50]} vs gen={repr(gv)[:50]}")
                break

    if formula_cols:
        print(f"  Formula columns: {sorted(set(formula_cols))}")

    # 4. Column widths
    width_diffs = 0
    for c in range(1, ts.max_column + 1):
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(c)
        tw = ts.column_dimensions[letter].width
        gw = gs.column_dimensions[letter].width
        if tw != gw:
            width_diffs += 1
    if width_diffs:
        print(f"  !! Column widths: {width_diffs} differences")
    else:
        print(f"  Column widths: OK (identical)")
