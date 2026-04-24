import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# 1. Get all waybills from template 尾程杂费 sheet
template_wb = openpyxl.load_workbook(
    r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\反馈客户\原始模板\20260330-汇森李志（东欧）对账单.xlsx',
    read_only=True, data_only=True)
ws = template_wb['20260330期尾程杂费']
rows = list(ws.rows)

surcharge_waybills = set()
surcharge_transfers = set()
for row in rows[1:]:
    vals = [c.value for c in row]
    wb = str(vals[1]).strip() if vals[1] else None  # col B = 运单号码
    tr = str(vals[3]).strip() if vals[3] else None  # col D = 转单号
    if wb:
        surcharge_waybills.add(wb)
    if tr:
        surcharge_transfers.add(tr)

# Also get the 62 remote-fee ones specifically
remote_transfers = set()
for row in rows[1:]:
    vals = [c.value for c in row]
    tr = str(vals[3]).strip() if vals[3] else None
    remote_fee = vals[15] if len(vals) > 15 else None
    if remote_fee and remote_fee != 0 and str(remote_fee) not in ('0', '0.0', ''):
        if tr:
            remote_transfers.add(tr)

# And the ones WITHOUT remote fee
non_remote_transfers = surcharge_transfers - remote_transfers

print(f'尾程杂费 total: {len(surcharge_transfers)}')
print(f'  with 偏远费: {len(remote_transfers)}')
print(f'  without 偏远费: {len(non_remote_transfers)}')
template_wb.close()

# 2. Now scan the 代理账单 to find which sheets these orders appear in
agent_wb = openpyxl.load_workbook(
    r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\鑫腾跃 -中文-对账单20260330.xlsx',
    read_only=True, data_only=True)

print(f'\n代理账单 sheets: {agent_wb.sheetnames}')

# For each sheet, find waybill column and check overlap
for sn in agent_wb.sheetnames:
    ws = agent_wb[sn]
    sheet_rows = list(ws.rows)
    if not sheet_rows:
        continue
    
    headers = [str(c.value).strip() if c.value else '' for c in sheet_rows[0]]
    
    # Find waybill column
    wb_idx = None
    for i, h in enumerate(headers):
        if '运单' in h and '号' in h:
            wb_idx = i
            break
    if wb_idx is None:
        for i, h in enumerate(headers):
            if '单号' in h and '客户' not in h:
                wb_idx = i
                break
    
    if wb_idx is None:
        print(f'\n  [{sn}] - no waybill column found. Headers: {headers[:8]}')
        continue
    
    sheet_waybills = set()
    for row in sheet_rows[1:]:
        vals = [c.value for c in row]
        if wb_idx < len(vals) and vals[wb_idx]:
            sheet_waybills.add(str(vals[wb_idx]).strip())
    
    # Match against surcharge waybills
    matched_surcharge = sheet_waybills & surcharge_waybills
    matched_remote = sheet_waybills & remote_transfers  # won't match (different format)
    
    # Also try matching surcharge_transfers against sheet
    matched_transfers = sheet_waybills & surcharge_transfers
    
    total_matches = len(matched_surcharge) + len(matched_transfers)
    if total_matches > 0 or len(sheet_waybills) > 0:
        # Count how many of the remote 62 are here
        remote_here = len(sheet_waybills & remote_transfers)
        non_remote_here = len(sheet_waybills & non_remote_transfers)
        
        print(f'\n  [{sn}] total_rows={len(sheet_waybills)}, col={headers[wb_idx]}')
        print(f'    matched 尾程杂费 by waybill: {len(matched_surcharge)}')
        print(f'    matched 尾程杂费 by transfer: {len(matched_transfers)}')

# 3. More useful: check by waybill format (IT/DE/HR)
# The surcharge sheet uses IT运单号 as waybill, the agent bill uses IT运单号 too
# Let's check which agent sheets contain the 501 surcharge waybills directly
print('\n\n=== Detailed: which agent sheets contain 尾程杂费 orders? ===')

# Build a map: waybill -> list of agent sheets it appears in
waybill_to_sheets = {}
for sn in agent_wb.sheetnames:
    ws = agent_wb[sn]
    sheet_rows = list(ws.rows)
    if not sheet_rows:
        continue
    headers = [str(c.value).strip() if c.value else '' for c in sheet_rows[0]]
    wb_idx = None
    for i, h in enumerate(headers):
        if '运单' in h and '号' in h:
            wb_idx = i
            break
    if wb_idx is None:
        for i, h in enumerate(headers):
            if '单号' in h and '客户' not in h:
                wb_idx = i
                break
    if wb_idx is None:
        continue
    
    for row in sheet_rows[1:]:
        vals = [c.value for c in row]
        if wb_idx < len(vals) and vals[wb_idx]:
            w = str(vals[wb_idx]).strip()
            waybill_to_sheets.setdefault(w, []).append(sn)

# Now check: for the 501 surcharge orders, which agent sheets do they appear in?
sheet_count = {}
remote_sheet_count = {}
for wb in surcharge_waybills:
    sheets = waybill_to_sheets.get(wb, ['NOT_FOUND'])
    for s in sheets:
        sheet_count[s] = sheet_count.get(s, 0) + 1

# Same for the 62 remote ones (they use transfer format F265xxxx)
# We need to match them differently - let's get the waybill format from template
template_wb2 = openpyxl.load_workbook(
    r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\反馈客户\原始模板\20260330-汇森李志（东欧）对账单.xlsx',
    read_only=True, data_only=True)
ws2 = template_wb2['20260330期尾程杂费']
rows2 = list(ws2.rows)

remote_waybills = set()
non_remote_waybills = set()
for row in rows2[1:]:
    vals = [c.value for c in row]
    wb = str(vals[1]).strip() if vals[1] else None
    remote_fee = vals[15] if len(vals) > 15 else None
    has_remote = remote_fee and remote_fee != 0 and str(remote_fee) not in ('0', '0.0', '')
    if wb:
        if has_remote:
            remote_waybills.add(wb)
        else:
            non_remote_waybills.add(wb)

template_wb2.close()

print('\nAll 501 尾程杂费 orders appear in these agent sheets:')
for s, cnt in sorted(sheet_count.items(), key=lambda x: -x[1]):
    print(f'  {s}: {cnt}')

# Now specifically for remote 62
print(f'\nThe 62 偏远费 orders (by waybill {list(remote_waybills)[:3]}...) appear in:')
for wb in remote_waybills:
    sheets = waybill_to_sheets.get(wb, ['NOT_FOUND'])
    for s in sheets:
        remote_sheet_count[s] = remote_sheet_count.get(s, 0) + 1

for s, cnt in sorted(remote_sheet_count.items(), key=lambda x: -x[1]):
    print(f'  {s}: {cnt}')

# Check the 31 extra DB remote (IT format) in agent sheets
print('\n=== The 31 DB-only remote orders in agent sheets ===')
extra_31 = [
    'IT12603091810073', 'IT12603091810036', 'IT12603091810061', 'IT12603091810062',
    'IT12603091810039', 'IT12603091810019', 'IT12603091810072', 'IT12603091810034',
    'IT12603091810037', 'IT12603091810074', 'IT12603091810035', 'IT12603091810065',
    'IT12603091810076', 'IT12603091810075', 'IT12603091810064', 'IT12603091810063',
    'IT12603091810067', 'IT12603091810025', 'IT12603021810027', 'IT12603021810025',
    'IT12603021810029', 'IT12603061810004', 'IT12603021810023', 'IT12603021810026',
    'IT12603021810024', 'IT12603031910007', 'IT12603021810022', 'IT12603041810006',
    'IT12603021810020', 'IT12603021810021', 'IT12603031910006',
]
extra_sheets = {}
for wb in extra_31:
    sheets = waybill_to_sheets.get(wb, ['NOT_FOUND'])
    for s in sheets:
        extra_sheets[s] = extra_sheets.get(s, 0) + 1

for s, cnt in sorted(extra_sheets.items(), key=lambda x: -x[1]):
    print(f'  {s}: {cnt}')

agent_wb.close()
