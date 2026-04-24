"""
逐客户逐行检查尾程运费公式是否按国家变化，
以及头程运费的实际计算方式。
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

PRICING = {
    '波兰':     {'first2': 4.0, 'extra1': 0.9},
    '罗马尼亚':  {'first2': 4.8, 'extra1': 0.6},
    '匈牙利':   {'first2': 4.4, 'extra1': 0.7},
    '捷克':     {'first2': 4.1, 'extra1': 0.8},
    '斯洛伐克':  {'first2': 4.3, 'extra1': 0.7},
    '保加利亚':  {'first2': 4.3, 'extra1': 0.8},
    '克罗地亚':  {'first2': 6.4, 'extra1': 1.0},
    '斯洛文尼亚': {'first2': 5.9, 'extra1': 0.9},
    '西班牙':   {'first2': 4.0, 'extra1': 1.0},
    '葡萄牙':   {'first2': 4.0, 'extra1': 1.0},
    '希腊':     {'first2': 5.7, 'extra1': 0.8},
    '意大利':   {'first2': 7.3, 'extra1': 1.1},
    '奥地利':   {'first2': 6.5, 'extra1': 1.0},
    '德国':     {'first2': 8.3, 'extra1': 1.5},
}

HEAD_RATES = {'GS': 7.7, 'SC': 9.5, 'IC': 10}


def extract_tail_params(formula):
    """Extract (first2, extra1) from IF(K>2, first2+(K-2)*extra1, first2) pattern."""
    if not formula or not isinstance(formula, str):
        return None, None, formula
    m = re.search(r'IF\(\w+>2\s*,\s*([\d.]+)\s*\+.*\*([\d.]+)\s*,\s*([\d.]+)\)', formula)
    if m:
        return float(m.group(1)), float(m.group(2)), formula
    return None, None, formula


print("=" * 100)
print("PART 1: 每个客户的尾程运费公式 — 是否逐行按国家变化")
print("=" * 100)

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    ws_f = ws_v = None
    tail_col = None
    dest_col = None
    for sname in wb_f.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws_f = wb_f[sname]
            ws_v = wb_v[sname]
            for c in range(1, ws_f.max_column + 1):
                h = str(ws_f.cell(1, c).value or '')
                if '尾程' in h and '运费' in h:
                    tail_col = c
                if '目的' in h:
                    dest_col = c
            break

    if not ws_f or not tail_col:
        wb_f.close()
        wb_v.close()
        continue

    print(f"\n--- {customer} [{ws_f.title}] 尾程=C{tail_col}, 目的地=C{dest_col} ---")

    # Collect per-country formulas
    country_formulas = {}
    for r in range(2, ws_f.max_row + 1):
        dest = ws_v.cell(r, dest_col).value if dest_col else None
        formula = ws_f.cell(r, tail_col).value
        val = ws_v.cell(r, tail_col).value
        if dest is None and formula is None:
            break
        if '合计' in str(dest or ''):
            break
        dest_str = str(dest or '?').strip()

        if dest_str not in country_formulas:
            country_formulas[dest_str] = {'formulas': set(), 'values': [], 'count': 0}
        country_formulas[dest_str]['count'] += 1
        if isinstance(formula, str) and formula.startswith('='):
            f1, f2, _ = extract_tail_params(formula)
            country_formulas[dest_str]['formulas'].add((f1, f2))
        if val is not None:
            country_formulas[dest_str]['values'].append(val)

    for dest, info in sorted(country_formulas.items()):
        formulas = info['formulas']
        vals = info['values']
        pricing = PRICING.get(dest)
        pricing_str = f" 报价={pricing['first2']}/{pricing['extra1']}" if pricing else " 报价=未知"

        for f1, f2 in formulas:
            match = ''
            if pricing and f1 is not None:
                if f1 == pricing['first2'] and f2 == pricing['extra1']:
                    match = ' ✓'
                else:
                    match = f" ✗ (报价={pricing['first2']}/{pricing['extra1']})"
            print(f"  {dest}({info['count']}): 首2={f1} 续1={f2}{match}{pricing_str}")

    wb_f.close()
    wb_v.close()


print(f"\n\n{'='*100}")
print("PART 2: 头程运费 — 实际计算追踪（取3个客户各取几条样本）")
print("=" * 100)

samples = [
    ('李志', '20260330-汇森李志（东欧）对账单.xlsx'),
    ('君悦', '汇森-君悦（东欧）对账单-20260330.xlsx'),
    ('峰总', '汇森-峰总（东欧）对账单-20260330.xlsx'),
]

for customer, fname in samples:
    fpath = TEMPLATE_DIR / fname
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)

    ws_v = ws_f = None
    for sname in wb_v.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws_v = wb_v[sname]
            ws_f = wb_f[sname]
            break
    if not ws_v:
        wb_v.close()
        wb_f.close()
        continue

    # Find key columns
    headers = {}
    for c in range(1, ws_v.max_column + 1):
        h = ws_v.cell(1, c).value
        if h:
            headers[str(h).strip()] = c

    head_col = None
    head_formula_col = None
    weight_col = None
    cat_col = None
    dest_col = None
    tail_w_col = None

    for h, c in headers.items():
        if '头程运费' in h:
            head_col = c
        if '头程计费' in h:
            weight_col = c
        if '普特敏' in h:
            cat_col = c
        if '目的' in h:
            dest_col = c
        if '尾程计费' in h:
            tail_w_col = c
        if '重量' in h and '头程' not in h and '尾程' not in h:
            if not weight_col:
                weight_col = c

    print(f"\n--- {customer} 头程运费分析 ---")
    print(f"  头程运费列=C{head_col}, 重量列=C{weight_col}, 货类列=C{cat_col}, 目的地=C{dest_col}")

    # Get the formula
    head_formula = ws_f.cell(2, head_col).value if head_col else None
    print(f"  公式: {head_formula}")

    # Also find the input-side data
    input_path = None
    code_map = {'李志': '中文', '君悦': '中文1', '峰总': '中文12'}
    code = code_map.get(customer)
    if code:
        for f in os.listdir(BASE / '中介提供'):
            if f'-{code}-' in f and f.endswith('.xlsx'):
                input_path = BASE / '中介提供' / f

    input_head_fees = {}
    input_service_fees = {}
    if input_path:
        wb_in = openpyxl.load_workbook(str(input_path), data_only=True)
        for sname in wb_in.sheetnames:
            ws_in = wb_in[sname]
            hr = None
            for r in range(1, 15):
                v = ws_in.cell(r, 3).value
                if v and '运单' in str(v):
                    hr = r
                    break
            if not hr:
                continue
            for r in range(hr + 1, ws_in.max_row + 1):
                wb_id = ws_in.cell(r, 3).value
                if not wb_id:
                    continue
                wb_id = str(wb_id).strip()
                amt = ws_in.cell(r, 13).value
                cw = ws_in.cell(r, 8).value
                if '头程运费' in sname:
                    input_head_fees[wb_id] = {'amt': amt, 'cw': cw}
                if '服务费' in sname:
                    input_service_fees[wb_id] = {'amt': amt, 'cw': cw}
        wb_in.close()

    print(f"\n  {'运单号':<25} {'目的地':<8} {'货类':>4} {'重量':>6} {'尾程重':>6} {'模板头程':>10} {'公式计算':>10} {'输入头程EUR':>12} {'输入服务费':>10}")
    print("  " + "-" * 110)

    count = 0
    for r in range(2, min(ws_v.max_row + 1, 30)):
        wb_id = ws_v.cell(r, 4).value
        if not wb_id:
            break
        wb_id = str(wb_id).strip()
        dest = ws_v.cell(r, dest_col).value if dest_col else ''
        cat = ws_v.cell(r, cat_col).value if cat_col else ''
        weight = ws_v.cell(r, weight_col).value if weight_col else 0
        tail_w = ws_v.cell(r, tail_w_col).value if tail_w_col else 0
        head_val = ws_v.cell(r, head_col).value if head_col else 0

        # Calculate what formula should give
        w = weight or 0
        rate = HEAD_RATES.get(str(cat or '').strip(), 10)
        if customer == '李志':
            # RMB: =ROUNDUP(IFS(...)*7.9342,2)
            import math
            head_eur = w * rate
            head_calc = math.ceil(head_eur * 7.9342 * 100) / 100
        else:
            head_calc = w * rate

        in_head = input_head_fees.get(wb_id, {}).get('amt', '-')
        in_head_cw = input_head_fees.get(wb_id, {}).get('cw', '-')
        in_svc = input_service_fees.get(wb_id, {}).get('amt', '-')

        print(f"  {wb_id:<25} {str(dest):<8} {str(cat):>4} {w:>6} {tail_w or '-':>6} {head_val or 0:>10} {head_calc:>10.2f} {str(in_head):>12} {str(in_svc):>10}")
        count += 1
        if count >= 15:
            break

    wb_v.close()
    wb_f.close()
