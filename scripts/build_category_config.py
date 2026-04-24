"""
Build precise category-driven configuration:
For each input category, determine exactly which output sheet it targets,
and the exact column mapping in that sheet.
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
INPUT_DIR = BASE / '中介提供'
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMER_MAP = {
    '中文':   '李志',
    '中文1':  '君悦',
    '中文3':  '小美',
    '中文5':  'J',
    '中文6':  '涵江',
    '中文7':  '阿甘',
    '中文9':  '威总',
    '中文12': '峰总',
}

# Which input categories each customer has
input_cats_by_customer = {}
for code, customer in CUSTOMER_MAP.items():
    for f in os.listdir(INPUT_DIR):
        if f'-{code}-' in f and f.endswith('.xlsx'):
            wb = openpyxl.load_workbook(str(INPUT_DIR / f), data_only=True)
            cats = set()
            for sname in wb.sheetnames:
                if sname in ('汇总', '总表'):
                    continue
                ws = wb[sname]
                hr = None
                for r in range(1, 15):
                    v = ws.cell(r, 3).value
                    if v and '运单' in str(v):
                        hr = r
                        break
                if not hr:
                    continue
                has_data = False
                for r in range(hr + 1, min(hr + 3, ws.max_row + 1)):
                    if ws.cell(r, 3).value:
                        has_data = True
                        break
                if has_data:
                    cats.add(sname)
            wb.close()
            input_cats_by_customer[customer] = cats

# For each output sheet type, which input categories determine membership
# and the exact column mapping per template
output_sheet_types = {
    'COD回款': {
        'input_trigger': 'COD',
        'description': '有COD科目的运单进入此Sheet',
    },
    '运费': {
        'input_trigger': '尾程运费',
        'description': '有尾程运费科目的运单进入此Sheet (= 有服务费的运单)',
    },
    '增值税': {
        'input_trigger': '目的地增值税',
        'description': '有目的地增值税科目的运单进入此Sheet (= 有头程运费的运单)',
    },
    '尾程杂费': {
        'input_trigger': ['上架费', '尾程退件操作费', '尾程退件操作费(补退)', '拒收返程费'],
        'description': '有上架费/退件费等杂费科目的运单进入此Sheet',
    },
    'F附加费': {
        'input_trigger': '尾程运费',
        'description': '与运费Sheet完全相同的运单集',
    },
}

print("=" * 80)
print("按科目维度的客户分布")
print("=" * 80)

for otype, info in output_sheet_types.items():
    trigger = info['input_trigger']
    if isinstance(trigger, str):
        trigger = [trigger]

    customers_with = []
    for customer, cats in input_cats_by_customer.items():
        if any(t in cats for t in trigger):
            customers_with.append(customer)

    print(f"\n  {otype} ({info['description']})")
    print(f"    触发科目: {trigger}")
    print(f"    有此科目的客户({len(customers_with)}): {customers_with}")

# Now scan each template for precise column configs
print(f"\n\n{'='*80}")
print("按输出Sheet类型，各客户的精确列结构")
print("=" * 80)

for customer in CUSTOMER_MAP.values():
    tmpl = None
    for f in os.listdir(TEMPLATE_DIR):
        if customer in f and f.endswith('.xlsx'):
            tmpl = TEMPLATE_DIR / f
    if not tmpl:
        continue

    wb = openpyxl.load_workbook(str(tmpl), data_only=False)
    print(f"\n{'='*60}")
    print(f"客户: {customer}")

    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h:
                headers[c] = str(h).strip()

        # Detect sheet type
        stype = None
        if 'COD' in sname:
            stype = 'COD回款'
        elif '运费' in sname and '杂费' not in sname:
            stype = '运费'
        elif '增值税' in sname:
            stype = '增值税'
        elif '杂费' in sname:
            stype = '尾程杂费'
        elif 'F' in sname and '附加' in sname:
            stype = 'F附加费'
        elif '汇总' in sname:
            stype = '汇总'
        elif '仓储' in sname:
            stype = '仓储'
        elif '理赔' in sname:
            stype = '理赔'
        else:
            stype = '?'

        if stype in ('汇总', '仓储', '理赔', '?'):
            continue

        # Key columns to identify
        waybill_col = None
        c5_field = None
        c6_field = None
        weight_col = None
        tail_weight_col = None
        category_col = None

        for c, h in headers.items():
            if '运单' in h and '编号' not in h:
                waybill_col = c
            if c == 5:
                c5_field = h
            if c == 6:
                c6_field = h

        # Get formula samples
        formulas = {}
        for c in range(1, ws.max_column + 1):
            for r in range(2, min(ws.max_row + 1, 6)):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.startswith('='):
                    formulas[c] = v
                    break

        print(f"\n  [{stype}] {sname}")
        print(f"    运单列: C{waybill_col}")
        print(f"    C5={c5_field}, C6={c6_field}")
        print(f"    Headers: {headers}")
        if formulas:
            print(f"    Formulas: {dict((c, f[:60]) for c, f in formulas.items())}")

    wb.close()
