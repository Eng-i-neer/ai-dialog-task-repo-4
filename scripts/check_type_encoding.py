"""检查 type 字段不匹配是否是编码问题"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
GEN_FILE = r'C:\Users\59571\Desktop\deutsch-app\舅妈网站\20260330-汇森李志（东欧）对账单-自动生成.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
gen = openpyxl.load_workbook(GEN_FILE, data_only=True)

ws_ref = ref[ref.sheetnames[0]]
ws_gen = gen[gen.sheetnames[0]]

# Check type encoding
mismatches = 0
real_mismatches = 0
for r in range(2, ws_ref.max_row + 1):
    ref_wb = ws_ref.cell(r, 4).value
    ref_type = ws_ref.cell(r, 2).value
    if not ref_wb:
        continue

    # Find in gen
    for gr in range(2, ws_gen.max_row + 1):
        gen_wb = ws_gen.cell(gr, 4).value
        if gen_wb == ref_wb:
            gen_type = ws_gen.cell(gr, 2).value
            if ref_type != gen_type:
                mismatches += 1
                # Check bytes
                ref_bytes = ref_type.encode('utf-8') if ref_type else b''
                gen_bytes = gen_type.encode('utf-8') if gen_type else b''
                if ref_bytes == gen_bytes:
                    print(f"  {ref_wb}: ENCODING MATCH (bytes same, str different?)")
                else:
                    real_mismatches += 1
                    if real_mismatches <= 10:
                        print(f"  {ref_wb}: ref='{ref_type}'({ref_bytes.hex()}), gen='{gen_type}'({gen_bytes.hex()})")
            break

print(f"\nTotal str mismatches: {mismatches}")
print(f"Real content mismatches: {real_mismatches}")

# Check the actual distribution of type values
ref_types = {}
gen_types = {}
for r in range(2, 200):
    rv = ws_ref.cell(r, 2).value
    if rv:
        ref_types[rv] = ref_types.get(rv, 0) + 1

for r in range(2, 200):
    gv = ws_gen.cell(r, 2).value
    if gv:
        gen_types[gv] = gen_types.get(gv, 0) + 1

print(f"\nRef type values: {ref_types}")
print(f"Gen type values: {gen_types}")

# Check COD CNY mismatch - specific examples
print("\n=== CNY Mismatches ===")
for r in range(2, ws_ref.max_row + 1):
    ref_wb = ws_ref.cell(r, 4).value
    ref_cny = ws_ref.cell(r, 11).value
    if not ref_wb:
        continue
    for gr in range(2, ws_gen.max_row + 1):
        gen_wb = ws_gen.cell(gr, 4).value
        if gen_wb == ref_wb:
            gen_cny = ws_gen.cell(gr, 11).value
            if ref_cny and gen_cny and abs(ref_cny - gen_cny) >= 0.02:
                ref_eur = ws_ref.cell(r, 9).value
                ref_rate = ws_ref.cell(r, 10).value
                print(f"  {ref_wb}: ref_cny={ref_cny}, gen_cny={gen_cny}, ref_eur={ref_eur}, ref_rate={ref_rate}")
            break
