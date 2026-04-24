"""读取报价文件中关于退件费/尾程派送费的规则"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'

wb = openpyxl.load_workbook(PRICING, data_only=True)
print("所有 Sheet 名称:")
for i, name in enumerate(wb.sheetnames):
    print(f"  [{i}] {name}")

print("\n" + "="*80)

for sname in wb.sheetnames:
    ws = wb[sname]
    # 检查是否包含"退件"、"派送"、"拒收"等关键词
    has_keyword = False
    for r in range(1, min(ws.max_row + 1, 5)):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(r, c).value or '')
            if any(kw in v for kw in ['退件', '派送', '拒收', '返程', '尾程']):
                has_keyword = True
                break
    
    if has_keyword or '派送' in sname or '退' in sname or '尾程' in sname or '报价' in sname:
        print(f"\n{'='*80}")
        print(f"Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols)")
        print(f"{'='*80}")
        for r in range(1, min(ws.max_row + 1, 30)):
            row_vals = []
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_vals.append(f"C{c}={v}")
            if row_vals:
                print(f"  R{r}: {' | '.join(row_vals)}")

wb.close()
