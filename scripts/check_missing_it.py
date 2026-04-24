"""检查模板中有但生成中缺失的IT运单的费用来源"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
inp_wb = openpyxl.load_workbook(INPUT, data_only=True)

ws_sur = tmpl_wb['20260330期尾程杂费']

# 缺失的运单列表
missing = ['IT12603181810028', 'IT12603181810007', 'IT12603161910046',
           'IT12603191810010', 'IT12603181810010', 'IT12603161910054',
           'IT12603181810016', 'IT12603161910048', 'IT12603161910052',
           'IT12603161910062']

# 模板中这些运单的数据
for wb in missing:
    for r in range(2, ws_sur.max_row + 1):
        if str(ws_sur.cell(r, 2).value) == wb:
            vals = {}
            for c in range(1, 21):
                v = ws_sur.cell(r, c).value
                if v is not None:
                    hdr = ws_sur.cell(1, c).value or f'C{c}'
                    vals[hdr] = v
            print(f"模板 R{r} {wb}: {vals}")
            break

# 输入文件中这些运单的费用
print("\n输入文件中的费用:")
for sname in inp_wb.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = inp_wb[sname]
    header_row = None
    for r in range(1, 15):
        v = ws.cell(r, 3).value
        if v and '运单' in str(v):
            header_row = r
            break
    if not header_row:
        continue
    for r in range(header_row + 1, ws.max_row + 1):
        wb_val = str(ws.cell(r, 3).value or '').strip()
        if wb_val in missing:
            amt = ws.cell(r, 13).value
            print(f"  {sname}: {wb_val} amount={amt}")

tmpl_wb.close()
inp_wb.close()
