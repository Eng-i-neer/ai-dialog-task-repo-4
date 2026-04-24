"""分析输入文件各费用sheet如何映射到尾程杂费表的列"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
inp_wb = openpyxl.load_workbook(INPUT, data_only=True)

ws_sur = tmpl_wb['20260330期尾程杂费']

# 建立模板中运单 -> 各列值的索引
tmpl_data = {}
for r in range(2, ws_sur.max_row + 1):
    wb = ws_sur.cell(r, 2).value
    if not wb:
        continue
    tmpl_data[str(wb)] = {
        'row': r,
        'shelf_fee': ws_sur.cell(r, 12).value,      # C12 上架费
        'return_fee': ws_sur.cell(r, 13).value,      # C13 尾程退件入仓费
        'second_delivery': ws_sur.cell(r, 14).value, # C14 二派费
        'vat': ws_sur.cell(r, 15).value,             # C15 增值税
        'remote_fee': ws_sur.cell(r, 16).value,      # C16 偏远费
        'subtotal': ws_sur.cell(r, 17).value,        # C17 小计
        'remark': ws_sur.cell(r, 18).value,          # C18 备注
        'zipcode': ws_sur.cell(r, 19).value,         # C19 邮编
        'c20': ws_sur.cell(r, 20).value,             # C20
    }

# 分析输入文件中的费用sheet，看哪些运单出现在尾程杂费表
surcharge_sheets = ['上架费', '尾程退件操作费', '目的地增值税', '尾程退件操作费(补退)']

# 建立输入文件中的运单费用索引
inp_fees = {}
for sname in inp_wb.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws_in = inp_wb[sname]
    header_row = None
    for r in range(1, 15):
        v = ws_in.cell(r, 3).value
        if v and '运单' in str(v):
            header_row = r
            break
    if not header_row:
        continue
    
    for r in range(header_row + 1, ws_in.max_row + 1):
        wb = ws_in.cell(r, 3).value
        if not wb or not isinstance(wb, str):
            continue
        wb = wb.strip()
        amount = ws_in.cell(r, 13).value
        if amount is None:
            continue
        if wb not in inp_fees:
            inp_fees[wb] = {}
        inp_fees[wb][sname] = amount

# 对模板中的每条运单，检查它在输入文件中有哪些费用
print("="*80)
print("模板运单 -> 输入文件费用映射验证")
print("="*80)

# 按备注类型分组分析
remark_groups = {}
for wb, td in tmpl_data.items():
    remark = td['remark'] or 'None'
    if remark not in remark_groups:
        remark_groups[remark] = []
    
    inp = inp_fees.get(wb, {})
    remark_groups[remark].append({
        'waybill': wb,
        'template': td,
        'input_fees': inp,
    })

for remark, items in sorted(remark_groups.items(), key=lambda x: -len(x[1])):
    print(f"\n备注类型: '{remark}' ({len(items)} 条)")
    for item in items[:3]:
        td = item['template']
        inp = item['input_fees']
        print(f"  {item['waybill']}:")
        print(f"    模板: 上架费={td['shelf_fee']}, 退件费={td['return_fee']}, "
              f"增值税={td['vat']}, 偏远费={td['remote_fee']}, "
              f"邮编={td['zipcode']}, C20={td['c20']}")
        print(f"    输入: {inp}")
        
        # 验证值是否匹配
        if td['shelf_fee'] and '上架费' in inp:
            shelf_in = inp['上架费']
            shelf_tmpl = td['shelf_fee']
            # 输入是EUR，模板是RMB
            print(f"    上架费验证: 输入EUR={shelf_in}, 模板RMB={shelf_tmpl}, "
                  f"输入*7.9342={shelf_in*7.9342 if isinstance(shelf_in,(int,float)) else 'N/A'}")
        
        if td['return_fee'] and '尾程退件操作费' in inp:
            ret_in = inp['尾程退件操作费']
            ret_tmpl = td['return_fee']
            print(f"    退件费验证: 输入EUR={ret_in}, 模板RMB={ret_tmpl}, "
                  f"输入*7.9342={ret_in*7.9342 if isinstance(ret_in,(int,float)) else 'N/A'}")
        
        if td['vat'] and '目的地增值税' in inp:
            vat_in = inp['目的地增值税']
            vat_tmpl = td['vat']
            print(f"    增值税验证: 输入EUR={vat_in}, 模板RMB={vat_tmpl}, "
                  f"输入*7.9342={vat_in*7.9342 if isinstance(vat_in,(int,float)) else 'N/A'}")

# 检查模板中不存在于输入文件的运单
print("\n" + "="*80)
print("模板中有但输入文件中没有费用的运单")
print("="*80)
missing = [wb for wb in tmpl_data if wb not in inp_fees]
print(f"  共 {len(missing)} 条")
for wb in missing[:10]:
    td = tmpl_data[wb]
    print(f"  {wb}: 上架={td['shelf_fee']}, 退件={td['return_fee']}, "
          f"增值税={td['vat']}, 偏远={td['remote_fee']}, 备注={td['remark']}")

# 检查输入文件中有费用但模板中没出现的运单
print("\n" + "="*80)
print("输入文件中有杂费但模板中不存在的运单")
print("="*80)
surcharge_fee_names = {'上架费', '尾程退件操作费', '目的地增值税', '尾程退件操作费(补退)'}
extra = []
for wb, fees in inp_fees.items():
    matching_fees = {k: v for k, v in fees.items() if k in surcharge_fee_names}
    if matching_fees and wb not in tmpl_data:
        extra.append((wb, matching_fees))
print(f"  共 {len(extra)} 条")
for wb, fees in extra[:10]:
    print(f"  {wb}: {fees}")

tmpl_wb.close()
inp_wb.close()
