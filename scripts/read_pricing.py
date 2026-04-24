# -*- coding: utf-8 -*-
"""Read pricing table to understand COD rules per country."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'web'))
sys.stdout.reconfigure(encoding='utf-8')
os.environ['FLASK_ENV'] = 'testing'

import openpyxl

pricing_files = [
    r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\汇森国际-东欧COD报价20250121生效.xlsx',
]

for fp in pricing_files:
    if not os.path.exists(fp):
        continue
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    print(f"\n=== {os.path.basename(fp)} ===")
    print(f"Sheets: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n--- Sheet: {sname} ---")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 50), values_only=True), 1):
            vals = [str(c) if c else '' for c in row[:15]]
            line = ' | '.join(v[:30] for v in vals)
            if any(v.strip() for v in vals):
                print(f"  R{row_idx}: {line}")
    wb.close()
