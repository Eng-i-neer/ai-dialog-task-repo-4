"""精确验证报价规则 - v3"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
OUTPUT  = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

output_wb = openpyxl.load_workbook(OUTPUT, data_only=True)

# ========== COD 分析 ==========
ws = output_wb['20260330期COD回款']
print("="*80)
print("COD回款表 - 列结构")
print("="*80)
for c in range(1, 15):
    print(f"  Col {c}: {ws.cell(1,c).value}")

# 样本分析
print("\n前10条数据详情:")
for r in range(2, 12):
    if not ws.cell(r,4).value:
        continue
    print(f"\n  Row {r}:")
    for c in range(1, 15):
        hdr = ws.cell(1,c).value or f'Col{c}'
        val = ws.cell(r,c).value
        print(f"    {hdr}: {val}")

# 统计COD手续费规律
print("\n\nCOD手续费分析:")
dest_cod_fees = {}
for r in range(2, ws.max_row + 1):
    if not ws.cell(r,4).value:
        continue
    dest = str(ws.cell(r,7).value or '')
    cod_amount = ws.cell(r,9).value  # 代收金额
    rate = ws.cell(r,10).value
    amount_cny = ws.cell(r,11).value  # 金额(CNY)
    cod_fee_cny = ws.cell(r,12).value  # COD手续费(CNY)
    total = ws.cell(r,13).value  # 小计
    
    if cod_amount and rate:
        # 金额CNY = 代收金额 * 汇率
        calc_amount_cny = cod_amount * rate
        # 小计 = 金额CNY - COD手续费CNY
        if isinstance(amount_cny, (int, float)) and isinstance(cod_fee_cny, (int, float)):
            calc_total = amount_cny - cod_fee_cny
            
            # COD手续费CNY => 反推 EUR = cod_fee_cny / rate
            cod_fee_eur = cod_fee_cny / rate if rate else 0
            pct3 = cod_amount * 0.03
            
            if dest not in dest_cod_fees:
                dest_cod_fees[dest] = []
            dest_cod_fees[dest].append({
                'r': r, 'cod_amount': cod_amount, 'rate': rate,
                'amount_cny': amount_cny, 'cod_fee_cny': cod_fee_cny,
                'total': total,
                'calc_amount_cny': calc_amount_cny,
                'cod_fee_eur': cod_fee_eur, 'pct3': pct3,
                'calc_total': calc_total,
            })

for dest, items in sorted(dest_cod_fees.items()):
    print(f"\n目的地: {dest} ({len(items)}条)")
    for item in items[:3]:
        cod_fee_eur = item['cod_fee_eur']
        pct3 = item['pct3']
        cod_amount = item['cod_amount']
        # 验证规则
        # 德国: max(3%, 7.0)
        rule7 = max(pct3, 7.0)
        # 意大利: max(3%, 2.0)
        rule2 = max(pct3, 2.0)
        
        print(f"  R{item['r']}: 代收EUR={cod_amount:.2f} | 3%={pct3:.4f} | "
              f"COD手续费反推EUR={cod_fee_eur:.4f} | "
              f"max(3%,7)={rule7:.4f} | max(3%,2)={rule2:.4f} | "
              f"手续费CNY={item['cod_fee_cny']}")

# ========== 运费分析 ==========
ws = output_wb['20260330期运费']
print("\n\n" + "="*80)
print("运费表 - 列结构")
print("="*80)
for c in range(1, 21):
    print(f"  Col {c}: {ws.cell(1,c).value}")

# 样本分析
print("\n前10条数据详情:")
for r in range(2, 12):
    if not ws.cell(r,4).value:
        continue
    print(f"\n  Row {r}:")
    for c in range(1, 21):
        hdr = ws.cell(1,c).value or f'Col{c}'
        val = ws.cell(r,c).value
        print(f"    {hdr}: {val}")

# 运费验证: 头程运费 + 尾程运费 计算
print("\n\n头程运费验证:")
print("报价: 德国普货 头程=9.7/0.1kg, 首2KG尾程=8.3, 续1KG=1.5")
print("报价: 意大利普货 头程=9.7/0.1kg, 首2KG尾程=7.3, 续1KG=1.1")

freight_analysis = {}
for r in range(2, ws.max_row + 1):
    waybill = ws.cell(r,4).value
    if not waybill:
        continue
    dest = str(ws.cell(r,7).value or '')
    category = str(ws.cell(r,8).value or '')
    actual_wt = ws.cell(r,9).value   # 收件实重
    head_wt = ws.cell(r,10).value    # 头程计费重量
    size = ws.cell(r,11).value       # 尺寸
    tail_wt = ws.cell(r,12).value    # 尾程计费重
    cargo_type = str(ws.cell(r,13).value or '')  # 普特敏货
    head_fee = ws.cell(r,14).value   # 头程运费
    tail_fee = ws.cell(r,15).value   # 尾程运费
    cod_fee = ws.cell(r,16).value    # 代收手续费
    addon = ws.cell(r,17).value      # 附加费
    peak = ws.cell(r,18).value       # 德国旺季附加费
    subtotal = ws.cell(r,19).value   # 小计
    
    if dest not in freight_analysis:
        freight_analysis[dest] = []
    freight_analysis[dest].append({
        'r': r, 'waybill': waybill, 'category': category,
        'actual_wt': actual_wt, 'head_wt': head_wt,
        'size': size, 'tail_wt': tail_wt,
        'cargo_type': cargo_type,
        'head_fee': head_fee, 'tail_fee': tail_fee,
        'cod_fee': cod_fee, 'addon': addon, 'peak': peak,
        'subtotal': subtotal,
    })

for dest, items in sorted(freight_analysis.items()):
    print(f"\n目的地: {dest} ({len(items)}条)")
    
    # 品名统计
    cargo_types = set(i['cargo_type'] for i in items)
    print(f"  普特敏货类型: {cargo_types}")
    
    for item in items[:5]:
        # 验证头程运费 = 头程计费重量(KG) * 头程单价/0.1KG
        # 头程单价: 普货=9.7/0.1KG = 97/KG
        head_wt = item['head_wt']
        head_fee = item['head_fee']
        
        if head_wt and head_fee and isinstance(head_wt, (int, float)) and isinstance(head_fee, (int, float)):
            # 如果头程计费重量单位是KG:
            calc_head_97 = head_wt * 97  # 9.7 EUR per 0.1KG = 97 per KG
            
            # 验证尾程运费
            tail_wt = item['tail_wt']
            tail_fee = item['tail_fee']
            
            if tail_wt and tail_fee and isinstance(tail_wt, (int, float)) and isinstance(tail_fee, (int, float)):
                # 尾程: 首2KG + (超出部分/1KG * 续费)
                if '德' in dest:
                    tail_first2 = 8.3
                    tail_per_kg = 1.5
                elif '意' in dest:
                    tail_first2 = 7.3
                    tail_per_kg = 1.1
                else:
                    tail_first2 = 0
                    tail_per_kg = 0
                
                if tail_wt <= 2:
                    calc_tail = tail_first2
                else:
                    import math
                    extra_kg = math.ceil(tail_wt - 2)
                    calc_tail = tail_first2 + extra_kg * tail_per_kg
                
                print(f"  R{item['r']}: {item['waybill']} | {item['cargo_type']} | "
                      f"头程重={head_wt} 头程费={head_fee} calc={calc_head_97:.2f} {'OK' if abs(head_fee-calc_head_97)<0.1 else 'X'} | "
                      f"尾程重={tail_wt} 尾程费={tail_fee} calc={calc_tail:.2f} {'OK' if abs(tail_fee-calc_tail)<0.1 else 'X'} | "
                      f"COD手续费={item['cod_fee']} | 附加费={item['addon']} | 小计={item['subtotal']}")
            else:
                print(f"  R{item['r']}: {item['waybill']} | {item['cargo_type']} | "
                      f"头程重={head_wt} 头程费={head_fee} calc={calc_head_97:.2f} | "
                      f"尾程重={tail_wt} 尾程费={tail_fee}")

# ========== 尾程杂费 ==========
ws = output_wb['20260330期尾程杂费']
print("\n\n" + "="*80)
print("尾程杂费表 - 列结构")
print("="*80)
for c in range(1, ws.max_column + 1):
    print(f"  Col {c}: {ws.cell(1,c).value}")

print("\n全部数据:")
for r in range(2, ws.max_row + 1):
    if not ws.cell(r,4).value:
        continue
    print(f"\n  Row {r}:")
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(1,c).value or f'Col{c}'
        val = ws.cell(r,c).value
        if val is not None:
            print(f"    {hdr}: {val}")

output_wb.close()
