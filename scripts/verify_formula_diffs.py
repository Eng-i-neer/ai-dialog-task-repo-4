"""Quick verification: are the 'None' diffs due to formula columns?"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')

pairs = [
    ('君悦', 'COD回款', 11),
    ('君悦', '头程&尾程运费', 13),
    ('峰总', '头程&尾程运费', 13),
    ('涵江', '目的地增值税', 11),
]

for customer, sheet_keyword, col in pairs:
    gen_path = BASE / '反馈客户' / '自动生成' / customer
    for f in gen_path.iterdir():
        if f.suffix == '.xlsx':
            wb = openpyxl.load_workbook(str(f), data_only=False)
            for sname in wb.sheetnames:
                if sheet_keyword in sname:
                    ws = wb[sname]
                    v = ws.cell(2, col).value
                    print(f"{customer}/{sname} C{col} R2: {repr(v)} {'(FORMULA)' if isinstance(v, str) and v.startswith('=') else ''}")
            wb.close()
