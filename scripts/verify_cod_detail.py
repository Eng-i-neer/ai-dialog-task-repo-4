"""
Detailed COD formula analysis - extract 3% rate and minimum EUR charge correctly.
Formula patterns:
  李志(RMB): =ROUNDUP(IF(K4=0,0,MAX(K4*0.03*1, 7*7.9342)),2) - 55.7
  Others(EUR): =IF(K2=0,0,MAX(K2*0.03*1, 1.5*7.9342))
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

OLD_COD = {
    'default': {'rate': 0.03, 'min_eur': 1.5},
    '意大利':  {'rate': 0.03, 'min_eur': 2.0},
    '奥地利':  {'rate': 0.03, 'min_eur': 5.0},
    '德国':   {'rate': 0.03, 'min_eur': 7.0},
}

NEW_COD = {
    'default': {'rate': 0.03, 'min_eur': 1.5},
    '意大利':  {'rate': 0.03, 'min_eur': 2.0},
    '奥地利':  {'rate': 0.03, 'min_eur': 5.0},
    '德国':   {'rate': 0.03, 'min_eur': 7.0},
}

print("=" * 120)
print("COD手续费公式详细分析")
print("=" * 120)
print("旧/新报价都一样: 3%费率, 默认最低1.5EUR, 意大利2.0EUR, 奥地利5EUR, 德国7.0EUR")
print("不签收时: 意大利按2.0EUR/票收, 奥地利按5EUR/票收, 德国按7.0EUR/票收")
print()

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    cod_ws_f = cod_ws_v = None
    for sname in wb_f.sheetnames:
        if 'COD' in sname or 'cod' in sname:
            cod_ws_f = wb_f[sname]
            cod_ws_v = wb_v[sname]
            break

    if not cod_ws_f:
        print(f"\n{customer}: 无COD sheet")
        wb_f.close()
        wb_v.close()
        continue

    headers = {}
    for c in range(1, cod_ws_f.max_column + 1):
        h = cod_ws_f.cell(1, c).value
        if h:
            headers[str(h).strip()] = c

    cod_col = dest_col = None
    for h, c in headers.items():
        if 'COD' in h and ('手续费' in h or '费' in h) and '回款' not in h and '代收' not in h:
            cod_col = c
        if '目的' in h:
            dest_col = c

    if not cod_col:
        for h, c in headers.items():
            if 'COD' in h and c > 5:
                cod_col = c
                break

    print(f"\n{customer} [{cod_ws_f.title}] COD手续费=C{cod_col}, 目的地=C{dest_col}")
    print(f"  Headers: {headers}")

    country_formulas = {}
    for r in range(2, cod_ws_f.max_row + 1):
        dest = cod_ws_v.cell(r, dest_col).value if dest_col else None
        formula = cod_ws_f.cell(r, cod_col).value if cod_col else None
        if dest is None and formula is None:
            break
        d = str(dest or '?').strip()
        if '合计' in d:
            break
        if d not in country_formulas and formula and isinstance(formula, str):
            country_formulas[d] = (r, formula)

    for country, (row, formula) in sorted(country_formulas.items()):
        # Parse: MAX(K*0.03*1, X*7.9342) or MAX(K*0.03, X)
        # The minimum is the second arg of MAX: X*7.9342 → X is min EUR
        m = re.search(r'MAX\([^,]+,\s*(\d+\.?\d*)\s*\*\s*7\.9342', formula)
        if m:
            min_eur = float(m.group(1))
        else:
            m2 = re.search(r'MAX\([^,]+,\s*(\d+\.?\d*)\s*\)', formula)
            min_eur = float(m2.group(1)) if m2 else None

        # Parse rate
        m_rate = re.search(r'\*\s*(0\.\d+)', formula)
        rate = float(m_rate.group(1)) if m_rate else None

        # Parse offset (the -XX.XX at the end for 不签收扣款)
        m_offset = re.search(r'\)\s*-\s*([\d.]+)', formula)
        offset = float(m_offset.group(1)) if m_offset else None

        expected = OLD_COD.get(country, OLD_COD['default'])
        rate_ok = rate == expected['rate'] if rate else '?'
        min_ok = min_eur == expected['min_eur'] if min_eur is not None else '?'

        marks = []
        if rate_ok == True: marks.append('费率3%✓')
        elif rate: marks.append(f'费率{rate}✗')
        if min_ok == True: marks.append(f'最低{min_eur}EUR✓')
        elif min_eur is not None: marks.append(f'最低{min_eur}EUR 预期{expected["min_eur"]}✗')
        if offset: marks.append(f'不签收扣{offset}')

        print(f"  {country}(R{row}): {' | '.join(marks)}")
        print(f"    公式: {formula}")

    wb_f.close()
    wb_v.close()
