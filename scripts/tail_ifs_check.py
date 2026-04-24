"""
Check if any template uses IFS/SWITCH to vary tail-freight formula by country,
or if each row manually has a different IF formula per country.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
}

PRICING = {
    '波兰':     {'first2': 4.0, 'extra1': 0.9},
    '罗马尼亚':  {'first2': 4.8, 'extra1': 0.6},
    '匈牙利':   {'first2': 4.4, 'extra1': 0.7},
    '捷克':     {'first2': 4.1, 'extra1': 0.8},
    '斯洛伐克':  {'first2': 4.3, 'extra1': 0.7},
    '保加利亚':  {'first2': 4.3, 'extra1': 0.8},
    '克罗地亚':  {'first2': 6.4, 'extra1': 1.0},
    '斯洛文尼亚': {'first2': 5.9, 'extra1': 0.9},
    '西班牙':   {'first2': 4.0, 'extra1': 1.0},
    '葡萄牙':   {'first2': 4.0, 'extra1': 1.0},
    '希腊':     {'first2': 5.7, 'extra1': 0.8},
    '意大利':   {'first2': 7.3, 'extra1': 1.1},
    '奥地利':   {'first2': 6.5, 'extra1': 1.0},
    '德国':     {'first2': 8.3, 'extra1': 1.5},
}

def extract_tail_params_from_if(formula):
    """Extract (first2, extra1) from IF(K>2, first2+(K-2)*extra1, first2)."""
    if not formula or not isinstance(formula, str):
        return None, None, None
    m = re.search(r'(\d+\.?\d*)\s*\+\s*\(.*?\)\s*\*\s*(\d+\.?\d*)\s*,\s*(\d+\.?\d*)', formula)
    if m:
        return float(m.group(1)), float(m.group(2)), float(m.group(3))
    return None, None, None


for customer in ['峰总']:
    fname = CUSTOMERS[customer]
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    ws_f = ws_v = None
    for sname in wb_f.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws_f = wb_f[sname]
            ws_v = wb_v[sname]
            break

    tail_col = dest_col = None
    for c in range(1, ws_f.max_column + 1):
        h = str(ws_f.cell(1, c).value or '')
        if '尾程' in h and '运费' in h:
            tail_col = c
        if '目的' in h:
            dest_col = c

    print(f"\n{'='*120}")
    print(f"{customer} — 逐行尾程公式检查 (列{tail_col})")
    print(f"{'='*120}")

    prev_dest = None
    for r in range(2, ws_f.max_row + 1):
        dest = ws_v.cell(r, dest_col).value
        formula = ws_f.cell(r, tail_col).value
        val = ws_v.cell(r, tail_col).value
        weight = ws_v.cell(r, tail_col - 3).value if tail_col else None

        if dest is None and formula is None:
            break
        if '合计' in str(dest or ''):
            break

        d = str(dest or '?').strip()
        if d != prev_dest:
            print(f"\n  --- {d} (从行{r}开始) ---")
            prev_dest = d

        f1, e1, f1b = extract_tail_params_from_if(str(formula))
        pricing = PRICING.get(d)
        match = ''
        has_extra = ''
        if pricing and f1 is not None:
            if f1 == pricing['first2'] and e1 == pricing['extra1']:
                match = ' ✓报价匹配'
            else:
                match = f' ✗不匹配(报价={pricing["first2"]}/{pricing["extra1"]})'
        if isinstance(formula, str) and '+' in formula and formula.count('+') > 1:
            has_extra = ' [含额外加费]'

        print(f"    R{r}: 首2={f1} 续1={e1} 值={val} {match}{has_extra}")
        print(f"      公式: {formula}")

    wb_f.close()
    wb_v.close()
