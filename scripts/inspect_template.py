"""检查模板文件的完整样式结构，为基于模板填充做准备"""
import openpyxl
from openpyxl.utils import get_column_letter

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
wb = openpyxl.load_workbook(TEMPLATE)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*60}")
    print(f"  Sheet: {sname}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"  Merged cells: {len(ws.merged_cells.ranges)}")
    for mc in ws.merged_cells.ranges:
        print(f"    {mc}")
    
    # Column widths
    print(f"  Column widths:")
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        w = ws.column_dimensions[letter].width
        if w:
            print(f"    {letter}: {w}")
    
    # Row heights for header
    for r in range(1, min(3, ws.max_row + 1)):
        h = ws.row_dimensions[r].height
        print(f"  Row {r} height: {h}")
    
    # Header cell styles (row 1)
    print(f"  Header styles (Row 1):")
    for c in range(1, min(ws.max_column + 1, 21)):
        cell = ws.cell(1, c)
        v = cell.value
        if v:
            f = cell.font
            fill = cell.fill
            align = cell.alignment
            border = cell.border
            nf = cell.number_format
            print(f"    {get_column_letter(c)}: val={v}")
            print(f"      font: name={f.name}, size={f.size}, bold={f.bold}, color={f.color}")
            print(f"      fill: {fill.start_color.rgb if fill.start_color else None}")
            print(f"      align: h={align.horizontal}, v={align.vertical}, wrap={align.wrap_text}")

    # Check last rows for summary section
    print(f"  Last rows (summary area):")
    for r in range(max(1, ws.max_row - 10), ws.max_row + 1):
        vals = []
        for c in range(1, min(ws.max_column + 1, 21)):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(f"{get_column_letter(c)}={repr(v)}")
        if vals:
            print(f"    Row {r}: {vals}")
