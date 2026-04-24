"""
按科目维度的差异对比 — 对每个科目(Sheet类型)，集中对比所有有该科目的客户
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'
OUTPUT_BASE = BASE / '反馈客户' / '自动生成'

CUSTOMERS = ['李志', '君悦', '小美', 'J', '涵江', '阿甘', '威总', '峰总']

SHEET_TYPES = {
    'COD回款': lambda n: 'COD' in n,
    '运费': lambda n: '运费' in n and '杂费' not in n,
    '增值税': lambda n: '增值税' in n,
    '尾程杂费': lambda n: '杂费' in n,
    'F附加费': lambda n: ('F' in n.upper() or 'f' in n) and '附加' in n,
}


def find_file(directory, keyword):
    for f in os.listdir(directory):
        if f.endswith('.xlsx') and keyword in f:
            return directory / f
    return None


def normalize(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v.startswith('='):
            return '[F]'
        return v if v else None
    if isinstance(v, float):
        return round(v, 4)
    return v


def diff_sheet_pair(ws_t, ws_g, sname):
    """Compare two sheets, return structured diff focused on data quality."""
    max_col = max(ws_t.max_column or 1, ws_g.max_column or 1)

    # Headers
    header_diffs = []
    for c in range(1, max_col + 1):
        th = str(ws_t.cell(1, c).value or '').strip()
        gh = str(ws_g.cell(1, c).value or '').strip()
        if th != gh:
            header_diffs.append((c, th, gh))

    # Waybill column
    wb_col = 4
    for c in range(1, min(max_col + 1, 6)):
        h = ws_t.cell(1, c).value
        if h and '运单' in str(h):
            wb_col = c
            break

    # Data rows
    def get_waybills(ws):
        wbs = {}
        for r in range(2, min(ws.max_row + 1, 800)):
            vals = [ws.cell(r, c).value for c in range(1, min(max_col + 1, 8))]
            if all(v is None for v in vals):
                break
            if any('合计' in str(v or '') for v in vals):
                break
            wb_id = ws.cell(r, wb_col).value
            if wb_id:
                wbs[str(wb_id).strip()] = r
        return wbs

    t_wbs = get_waybills(ws_t)
    g_wbs = get_waybills(ws_g)

    t_set, g_set = set(t_wbs), set(g_wbs)
    common = t_set & g_set

    # Cell diffs for matched waybills — categorize by column
    col_diffs = {}
    formula_only_diffs = 0
    real_diffs = 0

    for wb_id in sorted(common):
        tr, gr = t_wbs[wb_id], g_wbs[wb_id]
        for c in range(1, max_col + 1):
            tv = normalize(ws_t.cell(tr, c).value)
            gv = normalize(ws_g.cell(gr, c).value)
            if tv == gv:
                continue
            if tv == '[F]' or gv == '[F]':
                formula_only_diffs += 1
                continue
            real_diffs += 1
            h = str(ws_t.cell(1, c).value or f'C{c}')
            if h not in col_diffs:
                col_diffs[h] = {'count': 0, 'samples': []}
            col_diffs[h]['count'] += 1
            if len(col_diffs[h]['samples']) < 2:
                col_diffs[h]['samples'].append({
                    'wb': wb_id, 'template': str(tv), 'generated': str(gv)
                })

    return {
        'sheet': sname,
        'template_rows': len(t_wbs),
        'gen_rows': len(g_wbs),
        'row_match': len(t_wbs) == len(g_wbs),
        'missing_in_gen': sorted(t_set - g_set),
        'extra_in_gen': sorted(g_set - t_set)[:10],
        'extra_count': len(g_set - t_set),
        'formula_diffs': formula_only_diffs,
        'real_diffs': real_diffs,
        'col_diffs': col_diffs,
        'header_diffs': header_diffs,
    }


def main():
    all_results = {}

    for cat_name, match_fn in SHEET_TYPES.items():
        print(f"\n{'='*70}")
        print(f"科目: {cat_name}")
        print(f"{'='*70}")

        cat_results = {}
        for customer in CUSTOMERS:
            tmpl = find_file(TEMPLATE_DIR, customer)
            gen = find_file(OUTPUT_BASE / customer, customer)
            if not tmpl or not gen:
                continue

            wb_t = openpyxl.load_workbook(str(tmpl), data_only=True)
            wb_g = openpyxl.load_workbook(str(gen), data_only=True)

            t_sheet = None
            g_sheet = None
            for sn in wb_t.sheetnames:
                if match_fn(sn):
                    t_sheet = sn
                    break
            for sn in wb_g.sheetnames:
                if match_fn(sn):
                    g_sheet = sn
                    break

            if not t_sheet or not g_sheet:
                wb_t.close()
                wb_g.close()
                continue

            ws_t = wb_t[t_sheet]
            ws_g = wb_g[g_sheet]

            diff = diff_sheet_pair(ws_t, ws_g, t_sheet)
            cat_results[customer] = diff

            # Print
            status_parts = []
            if not diff['row_match']:
                status_parts.append(f"rows {diff['template_rows']}→{diff['gen_rows']}")
            if diff['missing_in_gen']:
                status_parts.append(f"{len(diff['missing_in_gen'])} missing")
            if diff['extra_count'] > 0:
                status_parts.append(f"{diff['extra_count']} extra")
            if diff['formula_diffs'] > 0:
                status_parts.append(f"{diff['formula_diffs']} formula-only")
            if diff['real_diffs'] > 0:
                status_parts.append(f"{diff['real_diffs']} real diffs")

            status = ', '.join(status_parts) if status_parts else 'PERFECT MATCH'
            print(f"\n  {customer} [{diff['sheet']}]: {status}")

            if diff['real_diffs'] > 0:
                for col_name, info in sorted(diff['col_diffs'].items(), key=lambda x: -x[1]['count']):
                    print(f"    {col_name}: {info['count']} diffs")
                    for s in info['samples']:
                        print(f"      [{s['wb'][:20]}] T='{s['template']}' G='{s['generated']}'")

            if diff['missing_in_gen'] and len(diff['missing_in_gen']) <= 5:
                print(f"    Missing: {diff['missing_in_gen']}")

            wb_t.close()
            wb_g.close()

        all_results[cat_name] = cat_results

    # Final cross-category summary
    print(f"\n\n{'='*70}")
    print("CROSS-CATEGORY SUMMARY")
    print(f"{'='*70}")
    print(f"\n{'科目':<12} {'客户':<8} {'模板行':>6} {'生成行':>6} {'行匹配':>6} {'公式差':>6} {'实际差':>6} {'差异列'}") 
    print("-" * 80)
    for cat_name, cat_results in all_results.items():
        for customer, diff in cat_results.items():
            diff_cols = ', '.join(f"{k}({v['count']})" for k, v in diff['col_diffs'].items()) if diff['col_diffs'] else '-'
            row_ok = 'OK' if diff['row_match'] else 'MISS'
            print(f"{cat_name:<12} {customer:<8} {diff['template_rows']:>6} {diff['gen_rows']:>6} {row_ok:>6} {diff['formula_diffs']:>6} {diff['real_diffs']:>6} {diff_cols}")

    # Save
    report_path = OUTPUT_BASE / 'category_diff_report.json'
    with open(str(report_path), 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nReport saved: {report_path}")


if __name__ == '__main__':
    main()
