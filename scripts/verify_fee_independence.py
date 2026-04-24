"""
Verify: 输入科目（代理成本）与输出科目（客户报价）完全独立计算

核心假设验证:
1. 运费Sheet的数据行 = 输入中有 尾程运费 or 头程运费 的运单集
2. 运费Sheet的头程/尾程 = 公式按客户报价计算，与代理输入金额无关
3. F附加费 = 运费Sheet的运单集（每单固定2EUR）
4. 增值税Sheet的运单集 = 输入中有 头程运费 的运单集
5. COD Sheet的运单集 = 输入中有 COD 的运单集
6. 尾程杂费Sheet的运单集 = 输入中有 上架费/退件费 等的运单集
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')

test_cases = [
    ('中文1', '君悦'),
    ('中文6', '涵江'),
    ('中文12', '峰总'),
]

for code, customer in test_cases:
    input_path = None
    for f in os.listdir(BASE / '中介提供'):
        if f'-{code}-' in f and f.endswith('.xlsx'):
            input_path = BASE / '中介提供' / f
    template_path = None
    for f in os.listdir(BASE / '反馈客户' / '原始模板'):
        if customer in f and f.endswith('.xlsx'):
            template_path = BASE / '反馈客户' / '原始模板' / f

    if not input_path or not template_path:
        continue

    print(f"\n{'='*80}")
    print(f"客户: {customer} ({code})")
    print(f"{'='*80}")

    # Parse input waybills by category
    wb_in = openpyxl.load_workbook(str(input_path), data_only=True)
    cat_waybills = {}
    for sname in wb_in.sheetnames:
        if sname in ('汇总', '总表'):
            continue
        ws = wb_in[sname]
        hr = None
        for r in range(1, 15):
            v = ws.cell(r, 3).value
            if v and '运单' in str(v):
                hr = r
                break
        if not hr:
            continue
        wbs = set()
        for r in range(hr + 1, ws.max_row + 1):
            wb_id = ws.cell(r, 3).value
            if wb_id and isinstance(wb_id, str) and wb_id.strip():
                wbs.add(wb_id.strip())
        cat_waybills[sname] = wbs
    wb_in.close()

    # Parse template waybills by sheet
    wb_tmpl = openpyxl.load_workbook(str(template_path), data_only=True)
    tmpl_waybills = {}
    for sname in wb_tmpl.sheetnames:
        ws = wb_tmpl[sname]
        wb_col = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and '运单' in str(h):
                wb_col = c
                break
        if not wb_col:
            continue
        wbs = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, wb_col).value
            if v:
                wbs.add(str(v).strip())
        if wbs:
            tmpl_waybills[sname] = wbs
    wb_tmpl.close()

    # Hypothesis testing
    input_tail = cat_waybills.get('尾程运费', set())
    input_head = cat_waybills.get('头程运费', set())
    input_cod = cat_waybills.get('COD', set())
    input_vat = cat_waybills.get('目的地增值税', set())
    input_service = cat_waybills.get('服务费', set())
    input_shelf = cat_waybills.get('上架费', set())
    input_return = cat_waybills.get('尾程退件操作费', set())

    for tsname, twbs in tmpl_waybills.items():
        print(f"\n  模板 [{tsname}]: {len(twbs)} 运单")

        if 'COD' in tsname:
            print(f"    vs 输入[COD]: {len(input_cod)} -> 匹配={len(twbs & input_cod)}, "
                  f"模板多={len(twbs - input_cod)}, 输入多={len(input_cod - twbs)}")

        elif '运费' in tsname and '杂费' not in tsname:
            print(f"    vs 输入[尾程运费]: {len(input_tail)} -> 匹配={len(twbs & input_tail)}, "
                  f"模板多={len(twbs - input_tail)}, 输入多={len(input_tail - twbs)}")
            print(f"    vs 输入[服务费]:   {len(input_service)} -> 匹配={len(twbs & input_service)}, "
                  f"模板多={len(twbs - input_service)}, 输入多={len(input_service - twbs)}")

        elif '增值税' in tsname:
            print(f"    vs 输入[目的地增值税]: {len(input_vat)} -> 匹配={len(twbs & input_vat)}, "
                  f"模板多={len(twbs - input_vat)}, 输入多={len(input_vat - twbs)}")
            print(f"    vs 输入[头程运费]:     {len(input_head)} -> 匹配={len(twbs & input_head)}, "
                  f"模板多={len(twbs - input_head)}, 输入多={len(input_head - twbs)}")

        elif '杂费' in tsname:
            combined = input_shelf | input_return
            print(f"    vs 输入[上架+退件]: {len(combined)} -> 匹配={len(twbs & combined)}, "
                  f"模板多={len(twbs - combined)}, 输入多={len(combined - twbs)}")

        elif 'F' in tsname and '附加' in tsname:
            print(f"    vs 输入[尾程运费]: {len(input_tail)} -> 匹配={len(twbs & input_tail)}, "
                  f"模板多={len(twbs - input_tail)}, 输入多={len(input_tail - twbs)}")

    # Show which input categories have NO match
    print(f"\n  输入中不直接对应模板Sheet的科目:")
    matched_cats = {'COD', '头程运费', '尾程运费', '目的地增值税', '上架费', '尾程退件操作费', '尾程退件操作费(补退)'}
    for cat, wbs in cat_waybills.items():
        if cat not in matched_cats:
            print(f"    [{cat}]: {len(wbs)} 运单 — 不单独出现在模板中")
