"""分析直发/转发判断逻辑 和 加费=30的条件"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# Build input lookup
ws_in = inp['COD']
in_data = {}
for r in range(10, ws_in.max_row + 1):
    wb = ws_in.cell(r, 3).value
    if wb:
        in_data[wb] = {
            'route': ws_in.cell(r, 5).value,
            'cust_ref': ws_in.cell(r, 16).value,
        }

# Build lookup from 保险费 sheet (separate input sheet)
for sname in inp.sheetnames:
    ws_tmp = inp[sname]
    for r in range(10, ws_tmp.max_row + 1):
        wb = ws_tmp.cell(r, 3).value
        if wb and wb not in in_data:
            in_data[wb] = {
                'route': ws_tmp.cell(r, 5).value,
                'cust_ref': ws_tmp.cell(r, 16).value,
            }

# Check type from COD sheet
ws_cod = ref[ref.sheetnames[0]]
print("=== COD Sheet 直发/转发 分析 ===")
type_route = {}
for r in range(2, ws_cod.max_row + 1):
    wb = ws_cod.cell(r, 4).value
    ship_type = ws_cod.cell(r, 2).value
    if wb and ship_type and wb in in_data:
        route = in_data[wb].get('route', 'UNKNOWN')
        cust_ref = in_data[wb].get('cust_ref', '')
        key = (route, ship_type)
        if key not in type_route:
            type_route[key] = []
        type_route[key].append((wb, cust_ref))

for k, v in sorted(type_route.items()):
    print(f"\n  route={k[0]}, type={k[1]}: {len(v)} records")
    for wb, cr in v[:3]:
        print(f"    {wb} -> cust_ref={cr}")

# Check freight sheet for addon=30 vs addon=11.91
ws_freight = ref[ref.sheetnames[1]]
print("\n\n=== 加费=30 vs 11.91 分析 ===")
addon_30 = []
addon_12 = []
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    tail_qty = ws_freight.cell(r, 12).value
    dest = ws_freight.cell(r, 7).value
    actual_w = ws_freight.cell(r, 9).value
    head_w = ws_freight.cell(r, 10).value
    if not wb:
        continue
    item = {
        'wb': wb,
        'addon': addon,
        'tail_qty': tail_qty,
        'dest': dest,
        'actual_w': actual_w,
        'head_w': head_w,
        'cust_ref': ws_freight.cell(r, 5).value,
    }
    if addon and abs(addon - 30) < 0.1:
        addon_30.append(item)
    elif addon and abs(addon - 11.91) < 0.1:
        addon_12.append(item)

print(f"\n加费=30: {len(addon_30)} 条")
print(f"  tail_qty分布: { {it['tail_qty']:0 for it in addon_30} }")
# Check actual weight for addon=30
weights_30 = [it['actual_w'] for it in addon_30 if it['actual_w']]
print(f"  actual_weight范围: {min(weights_30):.3f} - {max(weights_30):.3f}")
print(f"  示例:")
for it in addon_30[:5]:
    print(f"    {it['wb']}: actual_w={it['actual_w']}, head_w={it['head_w']}, tail_qty={it['tail_qty']}, dest={it['dest']}")

print(f"\n加费=11.91: {len(addon_12)} 条")
weights_12 = [it['actual_w'] for it in addon_12 if it['actual_w']]
print(f"  actual_weight范围: {min(weights_12):.3f} - {max(weights_12):.3f}")
print(f"  示例:")
for it in addon_12[:5]:
    print(f"    {it['wb']}: actual_w={it['actual_w']}, head_w={it['head_w']}, tail_qty={it['tail_qty']}, dest={it['dest']}")

# Check if addon=30 correlates with actual_weight vs head_weight difference
print("\n=== addon=30 时 actual_w vs charge_w 关系 ===")
for it in addon_30[:10]:
    aw = it['actual_w'] or 0
    hw = it['head_w'] or 0
    diff = aw - hw if hw else 'N/A'
    print(f"  {it['wb']}: actual={aw}, head_charge={hw}, diff={diff}")

print("\n=== addon=11.91 时 actual_w vs charge_w 关系 ===")
for it in addon_12[:10]:
    aw = it['actual_w'] or 0
    hw = it['head_w'] or 0
    diff = aw - hw if hw else 'N/A'
    print(f"  {it['wb']}: actual={aw}, head_charge={hw}, diff={diff}")
