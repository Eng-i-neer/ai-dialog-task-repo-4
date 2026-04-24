"""检查模板 尾程杂费 C13 的公式情况"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws = tmpl['20260330期尾程杂费']

# 检查 C13 列中有公式的行
print("C13(尾程退件入仓费) 公式行:")
formula_rows_c13 = []
value_rows_c13 = []
for r in range(2, 510):
    v = ws.cell(r, 13).value
    if isinstance(v, str) and v.startswith('='):
        formula_rows_c13.append(r)
        if len(formula_rows_c13) <= 5:
            print(f"  R{r}: {v}")
    elif v is not None:
        value_rows_c13.append((r, v))

if formula_rows_c13:
    print(f"  公式行范围: R{min(formula_rows_c13)}-R{max(formula_rows_c13)} (共{len(formula_rows_c13)}行)")
else:
    print("  无公式行!")
    
if value_rows_c13:
    print(f"  有值但非公式行: {value_rows_c13[:5]}")

# 同时检查 C12 的公式行范围
print("\nC12(上架费) 公式行:")
formula_rows_c12 = []
for r in range(2, 510):
    v = ws.cell(r, 12).value
    if isinstance(v, str) and v.startswith('='):
        formula_rows_c12.append(r)
        if len(formula_rows_c12) <= 3:
            print(f"  R{r}: {v}")
if formula_rows_c12:
    print(f"  公式行范围: R{min(formula_rows_c12)}-R{max(formula_rows_c12)} (共{len(formula_rows_c12)}行)")

# C15 增值税
print("\nC15(增值税) 公式行:")
formula_rows_c15 = []
for r in range(2, 510):
    v = ws.cell(r, 15).value
    if isinstance(v, str) and v.startswith('='):
        formula_rows_c15.append(r)
        if len(formula_rows_c15) <= 3:
            print(f"  R{r}: {v}")
if formula_rows_c15:
    print(f"  公式行范围: R{min(formula_rows_c15)}-R{max(formula_rows_c15)} (共{len(formula_rows_c15)}行)")

tmpl.close()
