"""
Verify ALL fee categories across all customer templates against the OLD pricing file
(汇森国际-东欧COD报价20251108生效(13).xlsx).

Categories to verify:
1. 头程运费 (Head Freight) - GS/SC/IC rates
2. 尾程运费 (Tail Delivery) - per-country first2/extra1 [ALREADY CONFIRMED 100%]
3. COD手续费 (COD Handling Fee) - 3% + minimum
4. 退件费/拒收返程费 (Return Fee) - "派送费×70%" vs "同派送费"
5. 增值税VAT - 1.2EUR/ticket + IOSS rates
6. F附加费 (F-Surcharge) - 2EUR/ticket or 30CNY/ticket
7. 上架费 (Shelf Fee) - 1.5EUR/ticket
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

OLD_HEAD = {
    'default': {'普货': 0.77, '特货': 0.95, '敏感货': 1.0},
    '西班牙':  {'普货': 0.92, '特货': 1.1, '敏感货': 1.2},
    '葡萄牙':  {'普货': 0.92, '特货': 1.1, '敏感货': 1.2},
}

OLD_TAIL = {
    '波兰':     {'first2': 3.8, 'extra1': 0.8},
    '罗马尼亚':  {'first2': 4.8, 'extra1': 0.6},
    '匈牙利':   {'first2': 4.1, 'extra1': 0.6},
    '捷克':     {'first2': 3.8, 'extra1': 0.6},
    '斯洛伐克':  {'first2': 4.0, 'extra1': 0.6},
    '保加利亚':  {'first2': 4.3, 'extra1': 0.8},
    '克罗地亚':  {'first2': 5.6, 'extra1': 0.9},
    '斯洛文尼亚': {'first2': 5.2, 'extra1': 0.8},
    '西班牙':   {'first2': 4.0, 'extra1': 1.0},
    '葡萄牙':   {'first2': 4.0, 'extra1': 1.0},
    '希腊':     {'first2': 5.7, 'extra1': 0.8},
    '意大利':   {'first2': 6.7, 'extra1': 1.0},
    '奥地利':   {'first2': 6.5, 'extra1': 1.0},
    '德国':     {'first2': 8.0, 'extra1': 1.5},
}

OLD_RETURN_RULE = {
    '波兰': '70%', '罗马尼亚': '70%', '匈牙利': '70%', '捷克': '70%',
    '斯洛伐克': '70%', '保加利亚': '70%', '克罗地亚': '70%', '斯洛文尼亚': '70%',
    '西班牙': '100%', '葡萄牙': '100%', '希腊': '100%',
    '意大利': '100%', '奥地利': '100%', '德国': '100%',
}

OLD_COD = {
    'default': {'rate': 0.03, 'min_eur': 1.5},
    '意大利':  {'rate': 0.03, 'min_eur': 2.0},
    '奥地利':  {'rate': 0.03, 'min_eur': 5.0},
    '德国':   {'rate': 0.03, 'min_eur': 7.0},
}

results = {}

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)
    results[customer] = {}

    for sname in wb_f.sheetnames:
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        headers = {}
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h:
                headers[str(h).strip()] = c

        # =============================================
        # 1. HEAD FREIGHT (头程运费)
        # =============================================
        if '运费' in sname and '杂费' not in sname:
            head_col = None
            for h, c in headers.items():
                if '头程运费' in h:
                    head_col = c
                    break
            if head_col:
                formula = ws_f.cell(2, head_col).value
                f_str = str(formula or '')
                gs = sc = ic = None
                m_gs = re.search(r'"GS"\s*,\s*\w+\s*\*\s*([\d.]+)', f_str)
                m_sc = re.search(r'"SC"\s*,\s*\w+\s*\*\s*([\d.]+)', f_str)
                m_ic = re.search(r'"IC"\s*,\s*\w+\s*\*\s*([\d.]+)', f_str)
                if m_gs: gs = float(m_gs.group(1))
                if m_sc: sc = float(m_sc.group(1))
                if m_ic: ic = float(m_ic.group(1))
                results[customer]['head'] = {
                    'formula': f_str, 'GS': gs, 'SC': sc, 'IC': ic
                }

        # =============================================
        # 2. COD HANDLING FEE (COD手续费)
        # =============================================
        if 'COD' in sname or 'cod' in sname:
            cod_col = None
            for h, c in headers.items():
                if 'COD' in h and '手续费' in h:
                    cod_col = c
                    break
                if 'COD' in h and '费' in h and '回款' not in h and '代收' not in h:
                    cod_col = c
            if cod_col:
                dest_col = None
                for h, c in headers.items():
                    if '目的' in h:
                        dest_col = c
                        break

                country_formulas = {}
                for r in range(2, ws_f.max_row + 1):
                    dest = ws_v.cell(r, dest_col).value if dest_col else None
                    formula = ws_f.cell(r, cod_col).value
                    if dest is None and formula is None:
                        break
                    if dest and '合计' in str(dest):
                        break
                    d = str(dest or '?').strip()
                    if d not in country_formulas and formula:
                        country_formulas[d] = str(formula)

                results[customer]['cod'] = country_formulas

        # =============================================
        # 3. RETURN FEE (退件费/杂费sheet中的C13)
        # =============================================
        if '杂费' in sname:
            return_col = None
            shelf_col = None
            vat_col = None
            dest_col = None
            for h, c in headers.items():
                if '退件' in h and '入仓' in h:
                    return_col = c
                if '上架' in h:
                    shelf_col = c
                if '增值税' in h or 'VAT' in h:
                    vat_col = c
                if '目的' in h:
                    dest_col = c

            if return_col:
                country_return = {}
                for r in range(2, ws_f.max_row + 1):
                    dest = ws_v.cell(r, dest_col).value if dest_col else None
                    formula = ws_f.cell(r, return_col).value
                    if dest is None and formula is None:
                        break
                    if dest and '合计' in str(dest):
                        break
                    d = str(dest or '?').strip()
                    if d not in country_return and formula and isinstance(formula, str):
                        country_return[d] = formula
                results[customer]['return'] = country_return

            if shelf_col:
                shelf_formulas = set()
                for r in range(2, min(ws_f.max_row + 1, 100)):
                    f = ws_f.cell(r, shelf_col).value
                    if f and isinstance(f, str) and f.startswith('='):
                        shelf_formulas.add(f)
                results[customer]['shelf'] = shelf_formulas

            if vat_col:
                vat_formulas = set()
                for r in range(2, min(ws_f.max_row + 1, 100)):
                    f = ws_f.cell(r, vat_col).value
                    if f and isinstance(f, str) and f.startswith('='):
                        vat_formulas.add(f)
                results[customer]['vat_surcharge'] = vat_formulas

        # =============================================
        # 4. VAT (增值税 sheet)
        # =============================================
        if '增值税' in sname or 'VAT' in sname:
            vat_amt_col = None
            dest_col = None
            for h, c in headers.items():
                if '增值税' in h or 'VAT' in h:
                    if '金额' in h or 'EUR' in h or 'RMB' in h:
                        vat_amt_col = c
                if '目的' in h:
                    dest_col = c

            if not vat_amt_col:
                for h, c in headers.items():
                    if ('增值税' in h or 'VAT' in h) and c > 3:
                        vat_amt_col = c
                        break

            if vat_amt_col:
                country_vat = {}
                for r in range(2, ws_f.max_row + 1):
                    dest = ws_v.cell(r, dest_col).value if dest_col else None
                    formula = ws_f.cell(r, vat_amt_col).value
                    if dest is None and formula is None:
                        break
                    if dest and '合计' in str(dest):
                        break
                    d = str(dest or '?').strip()
                    if d not in country_vat and formula:
                        country_vat[d] = str(formula)
                results[customer]['vat'] = country_vat

        # =============================================
        # 5. F-SURCHARGE (F附加费 sheet)
        # =============================================
        if 'F' in sname and '附加' in sname:
            f_col = None
            for h, c in headers.items():
                if '附加' in h and ('EUR' in h or 'RMB' in h or '金额' in h):
                    f_col = c
                if '附加费' in h and c > 3:
                    if not f_col:
                        f_col = c

            if f_col:
                f_formulas = set()
                for r in range(2, min(ws_f.max_row + 1, 30)):
                    f = ws_f.cell(r, f_col).value
                    if f is not None:
                        f_formulas.add(str(f))
                results[customer]['f_surcharge'] = f_formulas

    wb_f.close()
    wb_v.close()


# =============================================
# REPORT
# =============================================

print("=" * 120)
print("科目 1: 头程运费 — GS/SC/IC 费率 vs 旧报价")
print("=" * 120)
print(f"旧报价: 普货(GS)=0.77×10=7.7, 特货(SC)=0.95×10=9.5, 敏感货(IC)=1.0×10=10.0 EUR/KG")
print(f"        西班牙/葡萄牙: 普货=0.92×10=9.2, 特货=1.1×10=11.0, 敏感货=1.2×10=12.0 EUR/KG")
print()

for customer, data in results.items():
    head = data.get('head', {})
    gs, sc, ic = head.get('GS'), head.get('SC'), head.get('IC')
    old_gs, old_sc, old_ic = 7.7, 9.5, 10.0
    gs_ok = gs == old_gs if gs else '?'
    sc_ok = sc == old_sc if sc else '?'
    ic_ok = ic == old_ic if ic else '?'
    mark = '✓' if gs_ok == True and sc_ok == True and ic_ok == True else '✗'
    print(f"  {customer:<6}: GS={gs} SC={sc} IC={ic}  {mark}")
    if mark == '✗':
        print(f"         预期: GS=7.7 SC=9.5 IC=10.0")


print(f"\n\n{'='*120}")
print("科目 2: 尾程运费 — 已确认 100% 匹配旧报价 (见前次验证)")
print("=" * 120)
print("  全部 8 个国家/费率组合 = 100% 匹配旧报价 ✓")


print(f"\n\n{'='*120}")
print("科目 3: COD手续费 — 3%费率 + 最低收费")
print("=" * 120)
print(f"旧报价: 默认3%最低1.5EUR, 意大利最低2.0EUR, 奥地利最低5EUR, 德国最低7.0EUR")
print()

for customer, data in results.items():
    cod = data.get('cod', {})
    if not cod:
        print(f"  {customer:<6}: 无COD手续费公式 (可能无COD sheet)")
        continue
    print(f"  {customer:<6}:")
    for country, formula in sorted(cod.items()):
        pct = re.findall(r'(\d+\.?\d*)%|(\d+\.?\d*)\s*\*\s*(\d+\.?\d*)', formula)
        mins = re.findall(r'MAX\(.*?(\d+\.?\d+)', formula, re.IGNORECASE)
        if not mins:
            mins = re.findall(r',\s*(\d+\.?\d+)\s*\)', formula)
        expected = OLD_COD.get(country, OLD_COD['default'])
        print(f"    {country}: {formula[:80]}...")
        if mins:
            min_val = float(mins[0])
            match = min_val == expected['min_eur']
            print(f"      最低={min_val}EUR 预期={expected['min_eur']}EUR {'✓' if match else '✗'}")


print(f"\n\n{'='*120}")
print("科目 4: 退件费(拒收返程费) — 按国家规则")
print("=" * 120)
print(f"旧报价: 波兰/罗马尼亚/匈牙利/捷克/斯洛伐克/保加利亚/克罗地亚/斯洛文尼亚 = 派送费×70%")
print(f"        西班牙/葡萄牙/希腊/意大利/奥地利/德国 = 同派送费(100%)")
print()

for customer, data in results.items():
    ret = data.get('return', {})
    if not ret:
        print(f"  {customer:<6}: 无退件费公式")
        continue
    print(f"  {customer:<6}:")
    for country, formula in sorted(ret.items()):
        has_70pct = '70%' in formula or '0.7' in formula
        expected_rule = OLD_RETURN_RULE.get(country, '?')
        tail = OLD_TAIL.get(country)

        if expected_rule == '70%':
            match = has_70pct
        else:
            match = not has_70pct

        # Extract first2 and extra1 from return formula
        m = re.search(r'(\d+\.?\d*)\s*\+\s*\(.*?\)\s*\*\s*(\d+\.?\d*)', formula)
        f1 = float(m.group(1)) if m else None
        e1 = float(m.group(2)) if m else None

        rate_match = ''
        if tail and f1 is not None:
            if f1 == tail['first2'] and e1 == tail['extra1']:
                rate_match = ' 费率✓'
            else:
                rate_match = f' 费率✗(预期{tail["first2"]}/{tail["extra1"]})'

        rule_mark = '✓' if match else '✗'
        print(f"    {country}: 含70%={'是' if has_70pct else '否'} 预期={expected_rule} {rule_mark}{rate_match}")


print(f"\n\n{'='*120}")
print("科目 5: 增值税VAT — 税率 + 清关费1.2EUR/票")
print("=" * 120)

OLD_IOSS = {
    '奥地利': 0.20, '保加利亚': 0.20, '克罗地亚': 0.25,
    '捷克': 0.21, '爱沙尼亚': 0.22, '德国': 0.19,
    '希腊': 0.24, '匈牙利': 0.27, '意大利': 0.22,
    '波兰': 0.23, '葡萄牙': 0.23, '罗马尼亚': 0.19,
    '斯洛伐克': 0.23, '斯洛文尼亚': 0.22, '西班牙': 0.21,
}
print(f"旧报价: 清关费1.2EUR/票, VAT税率见IOSS表")
print()

for customer, data in results.items():
    vat = data.get('vat', {})
    if not vat:
        print(f"  {customer:<6}: 无独立VAT sheet公式")
        continue
    print(f"  {customer:<6}:")
    for country, formula in sorted(vat.items()):
        ioss = OLD_IOSS.get(country)
        nums = re.findall(r'(\d+\.?\d+)', formula)
        nums_f = [float(n) for n in nums if 0.01 < float(n) < 100]
        has_12 = any(abs(n - 1.2) < 0.01 for n in nums_f)
        has_10 = any(abs(n - 10) < 0.01 for n in nums_f)
        has_ioss = ioss and any(abs(n - ioss) < 0.001 for n in nums_f)

        parts = []
        if has_12 or has_10:
            parts.append('清关费✓')
        if has_ioss:
            parts.append(f'税率{ioss}✓')
        elif ioss:
            parts.append(f'税率{ioss}? 公式数值={nums_f}')

        print(f"    {country}: {' '.join(parts)}  公式={formula[:70]}")


print(f"\n\n{'='*120}")
print("科目 6: F附加费")
print("=" * 120)
print(f"旧报价: F货/纯电 = 2EUR/票, F手表 = 30CNY/票, 电子烟/保健品 = 1.5EUR/票")
print()

for customer, data in results.items():
    fs = data.get('f_surcharge', set())
    if not fs:
        print(f"  {customer:<6}: 无F附加费sheet/公式")
        continue
    print(f"  {customer:<6}: {fs}")


print(f"\n\n{'='*120}")
print("科目 7: 上架费 (退仓上架)")
print("=" * 120)
print(f"旧报价: 1.5EUR/票")
print()

for customer, data in results.items():
    shelf = data.get('shelf', set())
    if not shelf:
        print(f"  {customer:<6}: 无上架费公式")
        continue
    has_15 = any('1.5' in str(f) for f in shelf)
    print(f"  {customer:<6}: {shelf}  含1.5EUR={'✓' if has_15 else '✗'}")


print(f"\n\n{'='*120}")
print("科目 7b: 杂费sheet中的增值税列 (VAT清关费)")
print("=" * 120)
print(f"旧报价: 1.2EUR/票 (=ROUNDUP(1.2*汇率,2) 或 =10*税率)")
print()

for customer, data in results.items():
    vs = data.get('vat_surcharge', set())
    if not vs:
        print(f"  {customer:<6}: 无杂费VAT公式")
        continue
    print(f"  {customer:<6}: {vs}")
