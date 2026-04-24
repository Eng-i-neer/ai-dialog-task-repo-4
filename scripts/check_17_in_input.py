"""检查这17条运单在输入文件上架费Sheet中的状态"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

targets = [
    'HR12603071410001', 'DE12602091410035', 'DE12603111510046',
    'DE12602281410146', 'DE12603071410023', 'DE12603021410263',
    'DE12602281410149', 'DE12602261610008', 'DE12603071410035',
    'DE12603021410268', 'DE12602061410029', 'DE12603131410009',
    'DE12603091510034', 'DE12603051510007', 'DE12603051510015',
    'DE12603061510015', 'DE12602111410013',
]

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# 在上架费Sheet中查找
print("=== 在输入文件'上架费'Sheet中查找 ===")
ws = inp['上架费']
# 先确定表头行
for r in range(1, 10):
    vals = [ws.cell(r, c).value for c in range(1, 16)]
    non_none = [(c, v) for c, v in enumerate(vals, 1) if v is not None]
    if len(non_none) > 5:
        print(f"  表头行 R{r}: {[v for _, v in non_none]}")
        break

found_in_shelf = set()
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(r, c).value or '')
        if v in targets:
            amount = None
            for ac in [13, 14]:
                a = ws.cell(r, ac).value
                if isinstance(a, (int, float)):
                    amount = a
                    break
            found_in_shelf.add(v)
            print(f"  R{r}: 找到 {v}, 金额={amount}")

not_found = set(targets) - found_in_shelf
print(f"\n在上架费Sheet中找到: {len(found_in_shelf)} / {len(targets)}")
print(f"未找到: {len(not_found)}")
for wb in sorted(not_found):
    print(f"  {wb}")

# 对未找到的，检查是否在其他Sheet中
if not_found:
    print(f"\n=== 检查未找到的运单在其他Sheet中的情况 ===")
    for sname in inp.sheetnames:
        ws2 = inp[sname]
        for wb_id in not_found:
            for r in range(2, ws2.max_row + 1):
                for c in range(1, min(ws2.max_column + 1, 15)):
                    if str(ws2.cell(r, c).value or '') == wb_id:
                        fee_name = ws2.cell(r, 12).value
                        amount = ws2.cell(r, 13).value
                        print(f"  {wb_id} -> Sheet '{sname}' R{r}: 费用={fee_name} 金额={amount}")
                        break

inp.close()
