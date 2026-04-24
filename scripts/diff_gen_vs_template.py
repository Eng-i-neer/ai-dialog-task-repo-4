"""对比生成文件 vs 模板文件的差异"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
GENERATED = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\20260330-汇森李志（东欧）对账单-自动生成v5.xlsx'

tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
gen_wb = openpyxl.load_workbook(GENERATED, data_only=False)
tmpl_data_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
gen_data_wb = openpyxl.load_workbook(GENERATED, data_only=True)

print(f"模板 sheets: {tmpl_wb.sheetnames}")
print(f"生成 sheets: {gen_wb.sheetnames}")

for si, sname in enumerate(tmpl_wb.sheetnames):
    ws_t = tmpl_wb[sname]
    ws_g = gen_wb[sname]
    ws_td = tmpl_data_wb[sname]
    ws_gd = gen_data_wb[sname]
    
    print(f"\n{'='*80}")
    print(f"Sheet[{si}]: {sname}")
    print(f"  模板: {ws_t.max_row} rows x {ws_t.max_column} cols")
    print(f"  生成: {ws_g.max_row} rows x {ws_g.max_column} cols")
    print(f"{'='*80}")
    
    # 1. Header对比 (Row 1)
    header_match = True
    for c in range(1, max(ws_t.max_column, ws_g.max_column) + 1):
        vt = ws_t.cell(1, c).value
        vg = ws_g.cell(1, c).value
        if vt != vg:
            print(f"  HEADER DIFF Col{c}: 模板={vt} | 生成={vg}")
            header_match = False
    if header_match:
        print(f"  Header: MATCH")
    
    # 2. 公式列对比 - 检查R4行的公式是否保留
    print(f"\n  公式保留检查 (R2-R6):")
    for r in range(2, min(7, ws_t.max_row + 1)):
        tmpl_formulas = {}
        gen_formulas = {}
        for c in range(1, ws_t.max_column + 1):
            vt = ws_t.cell(r, c).value
            vg = ws_g.cell(r, c).value
            if isinstance(vt, str) and vt.startswith('='):
                tmpl_formulas[c] = vt[:50]
            if isinstance(vg, str) and vg.startswith('='):
                gen_formulas[c] = vg[:50]
        
        if tmpl_formulas or gen_formulas:
            all_cols = set(list(tmpl_formulas.keys()) + list(gen_formulas.keys()))
            for c in sorted(all_cols):
                tf = tmpl_formulas.get(c, '(none)')
                gf = gen_formulas.get(c, '(none)')
                status = "OK" if tf == gf else "DIFF"
                if status == "DIFF":
                    print(f"    R{r} C{c}: {status}")
                    print(f"      模板: {tf}")
                    print(f"      生成: {gf}")
    
    # 3. 数据对比 - 用 data_only=True 比较实际值
    max_check = min(ws_t.max_row, ws_g.max_row, 200)
    data_diffs = 0
    
    # 分别检查模板有数据而生成没有、以及生成有数据而模板没有
    print(f"\n  数据对比 (前{max_check}行):")
    
    # 找到模板中有运单号的行
    tmpl_waybills = {}
    gen_waybills = {}
    
    waybill_col = None
    for c in range(1, ws_t.max_column + 1):
        h = str(ws_t.cell(1, c).value or '')
        if '运单' in h:
            waybill_col = c
            break
    
    if waybill_col:
        for r in range(2, ws_t.max_row + 1):
            wb_val = ws_td.cell(r, waybill_col).value
            if wb_val:
                tmpl_waybills[str(wb_val)] = r
        for r in range(2, ws_g.max_row + 1):
            wb_val = ws_gd.cell(r, waybill_col).value
            if wb_val:
                gen_waybills[str(wb_val)] = r
        
        common = set(tmpl_waybills.keys()) & set(gen_waybills.keys())
        only_tmpl = set(tmpl_waybills.keys()) - set(gen_waybills.keys())
        only_gen = set(gen_waybills.keys()) - set(tmpl_waybills.keys())
        
        print(f"    模板运单数: {len(tmpl_waybills)}")
        print(f"    生成运单数: {len(gen_waybills)}")
        print(f"    共有运单数: {len(common)}")
        print(f"    仅模板有: {len(only_tmpl)}")
        if only_tmpl:
            for wb in list(only_tmpl)[:10]:
                r = tmpl_waybills[wb]
                print(f"      {wb} (模板R{r})")
        print(f"    仅生成有: {len(only_gen)}")
        if only_gen:
            for wb in list(only_gen)[:10]:
                r = gen_waybills[wb]
                print(f"      {wb} (生成R{r})")
        
        # 对共有运单逐列比较
        diff_count = 0
        for wb in list(common)[:20]:
            tr = tmpl_waybills[wb]
            gr = gen_waybills[wb]
            for c in range(1, ws_t.max_column + 1):
                tv = ws_td.cell(tr, c).value
                gv = ws_gd.cell(gr, c).value
                if tv != gv:
                    # 忽略公式列的计算值差异(因为生成文件未被Excel重新计算)
                    tv_f = ws_t.cell(tr, c).value
                    gv_f = ws_g.cell(gr, c).value
                    if isinstance(tv_f, str) and tv_f.startswith('=') and isinstance(gv_f, str) and gv_f.startswith('='):
                        continue
                    diff_count += 1
                    if diff_count <= 30:
                        hdr = ws_t.cell(1, c).value or f'Col{c}'
                        print(f"    DIFF {wb}: {hdr}(C{c}) 模板R{tr}={tv} | 生成R{gr}={gv}")
        
        if diff_count == 0:
            print(f"    共有运单数据列: ALL MATCH")
        else:
            print(f"    共有运单数据差异总数: {diff_count}")
    else:
        print(f"    未找到运单号列")

tmpl_wb.close()
gen_wb.close()
tmpl_data_wb.close()
gen_data_wb.close()
