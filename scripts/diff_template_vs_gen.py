"""精确对比模板文件 vs 生成文件的每一列：表头、格式、列宽、样式"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

# Find latest generated file
desktop = os.path.expanduser('~/Desktop')
gen_files = sorted([f for f in os.listdir(desktop) if f.startswith('20260330-result-')], reverse=True)
GEN_FILE = os.path.join(desktop, gen_files[0])

tmpl = openpyxl.load_workbook(TEMPLATE)
gen = openpyxl.load_workbook(GEN_FILE)

for si in range(len(tmpl.sheetnames)):
    ts = tmpl[tmpl.sheetnames[si]]
    gs = gen[gen.sheetnames[si]]
    
    print(f"\n{'='*70}")
    print(f"  Sheet {si}: template='{tmpl.sheetnames[si]}' vs gen='{gen.sheetnames[si]}'")
    print(f"  Template: {ts.max_row} rows x {ts.max_column} cols")
    print(f"  Generated: {gs.max_row} rows x {gs.max_column} cols")
    print(f"{'='*70}")
    
    # Compare headers (row 1)
    print("\n  --- HEADER (Row 1) ---")
    max_col = max(ts.max_column, gs.max_column)
    for c in range(1, max_col + 1):
        tv = ts.cell(1, c).value
        gv = gs.cell(1, c).value
        if tv != gv:
            print(f"  !! Col {c} HEADER MISMATCH: template={repr(tv)} vs gen={repr(gv)}")
        elif tv:
            print(f"  OK Col {c}: {repr(tv)}")
    
    # Compare column widths
    print("\n  --- COLUMN WIDTHS ---")
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        tw = ts.column_dimensions[letter].width
        gw = gs.column_dimensions[letter].width
        if tw != gw:
            print(f"  !! Col {letter} WIDTH: template={tw} vs gen={gw}")
    
    # Compare cell formats (row 2)
    print("\n  --- ROW 2 CELL FORMATS ---")
    for c in range(1, max_col + 1):
        tc = ts.cell(2, c)
        gc = gs.cell(2, c)
        diffs = []
        if tc.number_format != gc.number_format:
            diffs.append(f"numfmt: t='{tc.number_format}' g='{gc.number_format}'")
        if str(tc.font) != str(gc.font):
            diffs.append(f"font differs")
        if str(tc.alignment) != str(gc.alignment):
            diffs.append(f"align differs")
        if diffs:
            print(f"  !! Col {c}: {'; '.join(diffs)}")
    
    # Compare first 3 data rows values
    print("\n  --- DATA ROWS 2-4 ---")
    for r in range(2, min(5, min(ts.max_row, gs.max_row) + 1)):
        for c in range(1, max_col + 1):
            tv = ts.cell(r, c).value
            gv = gs.cell(r, c).value
            if tv != gv:
                print(f"  Row {r} Col {c}: template={repr(tv)[:60]} vs gen={repr(gv)[:60]}")
