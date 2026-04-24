"""
Read the pricing file to find all tail delivery pricing tiers
and check if there's a 'customer' or 'old' pricing column.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
pricing_path = BASE / '汇森国际-东欧COD报价20260331生效(5).xlsx'

wb = openpyxl.load_workbook(str(pricing_path), data_only=True)
ws = wb['欧洲COD']

print("报价文件 [欧洲COD] 完整内容:")
print("=" * 120)
for r in range(1, ws.max_row + 1):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"  R{r}: {', '.join(vals)}")

wb.close()

# Also check the formula version
wb_f = openpyxl.load_workbook(str(pricing_path), data_only=False)
ws_f = wb_f['欧洲COD']
print(f"\n\n报价文件 [欧洲COD] 公式版:")
print("=" * 120)
for r in range(1, ws_f.max_row + 1):
    vals = []
    for c in range(1, ws_f.max_column + 1):
        v = ws_f.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"  R{r}: {', '.join(vals)}")
wb_f.close()
