"""
全面提取报价文件中的所有定价规则
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

PRICING_FILE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\报价规则\汇森国际-东欧COD报价20260331生效(5).xlsx')

wb = openpyxl.load_workbook(str(PRICING_FILE), data_only=True)
print("Sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*80}")
    print(f"Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")
    for r in range(1, min(ws.max_row + 1, 40)):
        row_vals = []
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(f"C{c}={v}")
        if row_vals:
            print(f"  R{r}: {' | '.join(row_vals)}")

wb.close()
