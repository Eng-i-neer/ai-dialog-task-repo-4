"""Check why J/阿甘/威总 show no COD formulas, and check 峰总's ? country."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

checks = {
    'J':   '汇森-J（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
}

for customer, fname in checks.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    for sname in wb_f.sheetnames:
        if 'COD' not in sname and 'cod' not in sname:
            continue
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        print(f"\n{'='*80}")
        print(f"{customer} [{sname}] rows={ws_f.max_row}, cols={ws_f.max_column}")

        # Print first row headers
        print("  Headers:", end=' ')
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h:
                print(f"C{c}={h}", end='  ')
        print()

        # Print first 5 data rows (formula + value)
        for r in range(2, min(ws_f.max_row + 1, 8)):
            print(f"  R{r}:", end=' ')
            for c in [4, 7, 9, 12]:
                fv = ws_f.cell(r, c).value
                vv = ws_v.cell(r, c).value
                h = ws_f.cell(1, c).value or f'C{c}'
                if fv is not None:
                    if isinstance(fv, str) and fv.startswith('='):
                        print(f"{h}=[公式] ", end='')
                    else:
                        print(f"{h}={fv} ", end='')
                elif vv is not None:
                    print(f"{h}=v:{vv} ", end='')
            print()

        # Check for 峰总 R2 dest
        if customer == '峰总':
            print(f"\n  峰总 R2 目的地值: formula='{ws_f.cell(2, 7).value}' value='{ws_v.cell(2, 7).value}'")

        # Check for 李志 克罗地亚 row
        if customer == '李志':
            print(f"\n  李志 检查克罗地亚和希腊行的COD最低收费:")
            for r in range(2, ws_f.max_row + 1):
                dest = ws_v.cell(r, 7).value
                if dest and str(dest).strip() in ['克罗地亚', '希腊']:
                    formula = ws_f.cell(r, 12).value
                    print(f"    R{r} {dest}: {formula}")
                    if r > 115:
                        break

    wb_f.close()
    wb_v.close()
