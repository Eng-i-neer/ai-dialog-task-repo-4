# -*- coding: utf-8 -*-
"""Scan ALL 李志 agent bills to catalog every header layout variation."""
import openpyxl, os, sys, json
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单'

files = []
for dp, dn, fns in os.walk(BASE):
    for f in fns:
        if '李志' in f and f.endswith('.xlsx') and not f.startswith('~$'):
            files.append(os.path.join(dp, f))

files.sort()

HEADER_KEYWORDS = ['序号', '运单号码', '运单号', '寄件日期', '目的地', '原币金额']

def find_real_header(ws, max_row=15):
    """Find the row that contains actual column headers (not metadata like 帐单号码)."""
    for r in range(1, min(max_row + 1, (ws.max_row or 0) + 1)):
        row_vals = []
        for c in range(1, min(25, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(str(v).strip())
        row_text = ' '.join(row_vals)
        matches = sum(1 for kw in HEADER_KEYWORDS if kw in row_text)
        if matches >= 3:
            return r
    return None

layout_variations = defaultdict(list)
all_column_names = set()
sheet_name_set = set()
header_row_positions = defaultdict(int)
files_with_issues = []

for fpath in files:
    fname = os.path.basename(fpath)
    period = fname.split('对账单')[-1].replace('.xlsx', '').replace('-补', '')
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR opening {fname}: {e}")
        continue

    for sname in wb.sheetnames:
        sheet_name_set.add(sname)
        ws = wb[sname]
        hr = find_real_header(ws)
        if not hr:
            continue

        header_row_positions[hr] += 1
        cols = []
        for c in range(1, min(25, (ws.max_column or 0) + 1)):
            v = ws.cell(hr, c).value
            if v is not None:
                col_name = str(v).strip()
                cols.append(col_name)
                all_column_names.add(col_name)
            else:
                cols.append('')

        cols_clean = [c for c in cols if c]
        layout_key = '|'.join(cols_clean)
        layout_variations[layout_key].append(f"{period}:{sname}")

        # Check first data row for sanity
        first_row = hr + 1
        c3 = ws.cell(first_row, 3).value
        if c3 and ('编号' in str(c3) or '名称' in str(c3)):
            files_with_issues.append(f"{fname}:{sname} - C3={c3}")

    wb.close()

print("=" * 90)
print(f"扫描完成: {len(files)} 个文件")
print(f"\n所有出现过的 Sheet 名称 ({len(sheet_name_set)}):")
for s in sorted(sheet_name_set):
    print(f"  - {s}")

print(f"\n表头行位置分布:")
for row, count in sorted(header_row_positions.items()):
    print(f"  Row {row}: {count} 次")

print(f"\n所有出现过的列名 ({len(all_column_names)}):")
for c in sorted(all_column_names):
    print(f"  - '{c}'")

print(f"\n共发现 {len(layout_variations)} 种不同的列布局:")
for i, (layout, occurrences) in enumerate(sorted(layout_variations.items(), key=lambda x: -len(x[1]))):
    cols = layout.split('|')
    print(f"\n--- 布局 {i+1} (出现 {len(occurrences)} 次) ---")
    for j, c in enumerate(cols, 1):
        print(f"  Col {j:2d}: {c}")
    if len(occurrences) <= 8:
        for occ in occurrences:
            print(f"    -> {occ}")
    else:
        for occ in occurrences[:4]:
            print(f"    -> {occ}")
        print(f"    ... 和另外 {len(occurrences)-4} 个")

if files_with_issues:
    print(f"\n⚠ 可能有问题的 Sheet:")
    for issue in files_with_issues:
        print(f"  {issue}")
