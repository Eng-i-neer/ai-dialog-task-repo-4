"""遍历输入文件所有Sheet，找出 DE12510231810002 和 DE12511241510023 出现在哪里"""
import openpyxl

INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
wb = openpyxl.load_workbook(INPUT, data_only=True)

targets = ['DE12510231810002', 'DE12511241510023']

for t in targets:
    print(f"\n{'='*60}")
    print(f"  Searching: {t}")
    print(f"{'='*60}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(r, c).value
                if v == t:
                    print(f"\n  Found in sheet '{sname}' row {r} col {c}")
                    print(f"  Full row data:")
                    for cc in range(1, min(ws.max_column + 1, 19)):
                        val = ws.cell(r, cc).value
                        if val is not None:
                            print(f"    Col {cc}: {repr(val)}")
