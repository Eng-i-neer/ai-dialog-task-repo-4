"""对比自动生成的输出与参考输出"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
GEN_FILE = r'C:\Users\59571\Desktop\deutsch-app\舅妈网站\20260330-汇森李志（东欧）对账单-自动生成.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
gen = openpyxl.load_workbook(GEN_FILE, data_only=True)

# ── Compare COD Sheet ──
print("=" * 60)
print("SHEET 1: COD回款 对比")
print("=" * 60)

ws_ref = ref[ref.sheetnames[0]]
ws_gen = gen[gen.sheetnames[0]]

# Build lookup by waybill for ref (col D) and gen (col D)
ref_cod = {}
for r in range(2, ws_ref.max_row + 1):
    wb = ws_ref.cell(r, 4).value
    if wb and isinstance(wb, str) and wb.startswith(('DE', 'IT', 'HR', 'GR')):
        ref_cod[wb] = {
            'type': ws_ref.cell(r, 2).value,
            'date': ws_ref.cell(r, 3).value,
            'cust_ref': ws_ref.cell(r, 5).value,
            'dest': ws_ref.cell(r, 7).value,
            'category': ws_ref.cell(r, 8).value,
            'eur': ws_ref.cell(r, 9).value,
            'cny': ws_ref.cell(r, 11).value,
            'diff': ws_ref.cell(r, 12).value,
            'total': ws_ref.cell(r, 13).value,
        }

gen_cod = {}
for r in range(2, ws_gen.max_row + 1):
    wb = ws_gen.cell(r, 4).value
    if wb and isinstance(wb, str) and wb.startswith(('DE', 'IT', 'HR', 'GR')):
        gen_cod[wb] = {
            'type': ws_gen.cell(r, 2).value,
            'date': ws_gen.cell(r, 3).value,
            'cust_ref': ws_gen.cell(r, 5).value,
            'dest': ws_gen.cell(r, 7).value,
            'category': ws_gen.cell(r, 8).value,
            'eur': ws_gen.cell(r, 9).value,
            'cny': ws_gen.cell(r, 11).value,
            'diff': ws_gen.cell(r, 12).value,
            'total': ws_gen.cell(r, 13).value,
        }

print(f"参考文件 COD 记录: {len(ref_cod)}")
print(f"生成文件 COD 记录: {len(gen_cod)}")

# Find common waybills
common = set(ref_cod.keys()) & set(gen_cod.keys())
only_ref = set(ref_cod.keys()) - set(gen_cod.keys())
only_gen = set(gen_cod.keys()) - set(ref_cod.keys())

print(f"共同运单: {len(common)}")
print(f"仅在参考中: {len(only_ref)} -> {list(only_ref)[:5]}")
print(f"仅在生成中: {len(only_gen)} -> {list(only_gen)[:5]}")

# Compare field by field for common waybills
match_counts = {'type': 0, 'date': 0, 'cust_ref': 0, 'dest': 0, 'category': 0, 'eur': 0, 'cny': 0}
mismatch_examples = {}
for wb in common:
    r = ref_cod[wb]
    g = gen_cod[wb]
    for field in match_counts:
        rv = r[field]
        gv = g[field]
        if rv == gv or (isinstance(rv, (int, float)) and isinstance(gv, (int, float)) and abs(rv - gv) < 0.02):
            match_counts[field] += 1
        else:
            if field not in mismatch_examples:
                mismatch_examples[field] = (wb, rv, gv)

print(f"\n字段匹配率 (共 {len(common)} 条):")
for field, count in match_counts.items():
    pct = count / len(common) * 100 if common else 0
    print(f"  {field:12s}: {count}/{len(common)} ({pct:.1f}%)")
    if field in mismatch_examples:
        wb, rv, gv = mismatch_examples[field]
        print(f"    Example mismatch @ {wb}: ref={rv}, gen={gv}")

# ── Compare Freight Sheet ──
print("\n" + "=" * 60)
print("SHEET 2: 运费 对比")
print("=" * 60)

ws_ref_f = ref[ref.sheetnames[1]]
ws_gen_f = gen[gen.sheetnames[1]]

ref_freight = {}
for r in range(2, ws_ref_f.max_row + 1):
    wb = ws_ref_f.cell(r, 4).value
    if wb and isinstance(wb, str):
        ref_freight[wb] = {
            'head_weight': ws_ref_f.cell(r, 10).value,
            'tail_qty': ws_ref_f.cell(r, 12).value,
            'head_cost': ws_ref_f.cell(r, 14).value,
            'tail_cost': ws_ref_f.cell(r, 15).value,
            'de_sur': ws_ref_f.cell(r, 16).value,
            'addon': ws_ref_f.cell(r, 17).value,
            'subtotal': ws_ref_f.cell(r, 19).value,
        }

gen_freight = {}
for r in range(2, ws_gen_f.max_row + 1):
    wb = ws_gen_f.cell(r, 4).value
    if wb and isinstance(wb, str):
        gen_freight[wb] = {
            'head_weight': ws_gen_f.cell(r, 10).value,
            'tail_qty': ws_gen_f.cell(r, 12).value,
            'head_cost': ws_gen_f.cell(r, 14).value,
            'tail_cost': ws_gen_f.cell(r, 15).value,
            'de_sur': ws_gen_f.cell(r, 16).value,
            'addon': ws_gen_f.cell(r, 17).value,
            'subtotal': ws_gen_f.cell(r, 19).value,
        }

print(f"参考文件运费记录: {len(ref_freight)}")
print(f"生成文件运费记录: {len(gen_freight)}")

common_f = set(ref_freight.keys()) & set(gen_freight.keys())
print(f"共同运单: {len(common_f)}")

match_f = {'head_weight': 0, 'tail_qty': 0, 'head_cost': 0, 'tail_cost': 0, 'de_sur': 0, 'addon': 0, 'subtotal': 0}
mismatch_f = {}
for wb in common_f:
    r = ref_freight[wb]
    g = gen_freight[wb]
    for field in match_f:
        rv = r[field]
        gv = g[field]
        if rv == gv or (isinstance(rv, (int, float)) and isinstance(gv, (int, float)) and abs(rv - gv) < 0.1):
            match_f[field] += 1
        else:
            if field not in mismatch_f:
                mismatch_f[field] = (wb, rv, gv)

print(f"\n字段匹配率 (共 {len(common_f)} 条):")
for field, count in match_f.items():
    pct = count / len(common_f) * 100 if common_f else 0
    print(f"  {field:12s}: {count}/{len(common_f)} ({pct:.1f}%)")
    if field in mismatch_f:
        wb, rv, gv = mismatch_f[field]
        print(f"    Example mismatch @ {wb}: ref={rv}, gen={gv}")
