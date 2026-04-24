"""检查 C13 公式与目的国的对应关系"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, re

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws = tmpl['20260330期尾程杂费']

# 收集所有 C13 有公式的行，按目的地分组
dest_formula = {}
for r in range(2, 510):
    v13 = ws.cell(r, 13).value
    dest = ws.cell(r, 10).value
    if isinstance(v13, str) and v13.startswith('='):
        # 抽取公式中的参数（基础费和续重费）
        pattern = re.sub(r'[A-Z]\d+', 'X', v13)
        key = (dest, pattern)
        if key not in dest_formula:
            dest_formula[key] = {'count': 0, 'example': v13, 'row': r}
        dest_formula[key]['count'] += 1

print("C13 公式按 (目的地, 公式模式) 分组:")
for (dest, pattern), info in sorted(dest_formula.items(), key=lambda x: x[0]):
    print(f"  {dest}: {info['example']}  (x{info['count']}, 样例R{info['row']})")

# 分别看看 非公式的 C13 行 (R2, R3, R503) 的值
print("\nC13 非公式行:")
for r in [2, 3, 503]:
    v = ws.cell(r, 13).value
    dest = ws.cell(r, 10).value
    wb_id = ws.cell(r, 2).value
    print(f"  R{r}: 运单={wb_id} 目的地={dest} C13={v}")

# 也看一下 C12 上架费的分布 —— 是否都是同样的公式？
print("\nC12 上架费按目的地:")
c12_by_dest = {}
for r in range(2, 510):
    v12 = ws.cell(r, 12).value
    dest = ws.cell(r, 10).value
    if isinstance(v12, str) and v12.startswith('='):
        pattern = v12
        if dest not in c12_by_dest:
            c12_by_dest[dest] = set()
        c12_by_dest[dest].add(pattern)
for dest, formulas in sorted(c12_by_dest.items()):
    print(f"  {dest}: {formulas}")

# 最终看哪些行有 C12 公式但没有 C13 公式 (上架费有但退件费没有)
has_c12_no_c13 = []
for r in range(2, 510):
    v12 = ws.cell(r, 12).value
    v13 = ws.cell(r, 13).value
    if isinstance(v12, str) and v12.startswith('=') and not (isinstance(v13, str) and v13.startswith('=')):
        has_c12_no_c13.append(r)
print(f"\n有C12上架费公式但无C13退件费公式的行: {len(has_c12_no_c13)}行")
if has_c12_no_c13:
    for r in has_c12_no_c13[:5]:
        print(f"  R{r}: 运单={ws.cell(r,2).value} 目的地={ws.cell(r,10).value} C12={ws.cell(r,12).value} C13={ws.cell(r,13).value}")

# 反过来：有 C13 但没有 C12
has_c13_no_c12 = []
for r in range(2, 510):
    v12 = ws.cell(r, 12).value
    v13 = ws.cell(r, 13).value
    if isinstance(v13, str) and v13.startswith('=') and not (isinstance(v12, str) and v12.startswith('=')):
        has_c13_no_c12.append(r)
print(f"\n有C13退件费公式但无C12上架费公式的行: {len(has_c13_no_c12)}行")
if has_c13_no_c12:
    for r in has_c13_no_c12[:5]:
        remark = ws.cell(r, 18).value
        print(f"  R{r}: 运单={ws.cell(r,2).value} 目的地={ws.cell(r,10).value} C13={ws.cell(r,13).value} 备注={remark}")

tmpl.close()
