# -*- coding: utf-8 -*-
"""Check the customer template for how 二派费 is reflected."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

fp = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\反馈客户\原始模板\李志客户账单\20250818-汇森李志（东欧）对账单.xlsx'
if not os.path.exists(fp):
    print(f"File not found: {fp}")
    sys.exit()

wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) ---")
    for row_idx in range(1, min(8, (ws.max_row or 8) + 1)):
        vals = []
        for col_idx in range(1, min(25, (ws.max_column or 25) + 1)):
            v = ws.cell(row_idx, col_idx).value
            vals.append(str(v)[:25] if v else '')
        line = ' | '.join(vals)
        if any(v.strip() for v in vals):
            print(f"  R{row_idx}: {line}")
wb.close()
