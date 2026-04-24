"""Extract remaining rows from 欧洲COD sheet"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

PRICING_FILE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\报价规则\汇森国际-东欧COD报价20260331生效(5).xlsx')
wb = openpyxl.load_workbook(str(PRICING_FILE), data_only=True)
ws = wb['欧洲COD']

for r in range(39, ws.max_row + 1):
    row_vals = []
    for c in range(1, min(ws.max_column + 1, 18)):
        v = ws.cell(r, c).value
        if v is not None:
            row_vals.append(f"C{c}={v}")
    if row_vals:
        print(f"R{r}: {' | '.join(row_vals)}")

wb.close()
