"""检查模板中备注含'上架费'但C12为空的情况"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws = tmpl['20260330期尾程杂费']

# data_only版本用来看计算后的值
tmpl_val = openpyxl.load_workbook(TEMPLATE, data_only=True)
ws_val = tmpl_val['20260330期尾程杂费']

count = 0
for r in range(2, ws.max_row + 1):
    wb_id = ws.cell(r, 2).value
    c12_formula = ws.cell(r, 12).value
    c12_value = ws_val.cell(r, 12).value
    remark = ws.cell(r, 18).value
    
    if remark and '上架费' in str(remark) and c12_formula is None:
        count += 1
        print(f"R{r}: 运单={wb_id} C12公式={c12_formula} C12值={c12_value} 备注={remark}")

print(f"\n共 {count} 条备注含'上架费'但C12为空")

# 反过来看看: C12有公式但备注不含'上架费'
print("\n---")
count2 = 0
for r in range(2, ws.max_row + 1):
    wb_id = ws.cell(r, 2).value
    c12 = ws.cell(r, 12).value
    remark = ws.cell(r, 18).value
    if c12 is not None and remark and '上架费' not in str(remark):
        count2 += 1
        if count2 <= 5:
            print(f"R{r}: 运单={wb_id} C12={c12} 备注={remark}")
print(f"共 {count2} 条C12有值但备注不含'上架费'")

tmpl.close()
tmpl_val.close()
