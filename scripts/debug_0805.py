# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

f = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单\2025.08.05\（李志）鑫腾跃-中文-对账单20250805.xlsx'
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

for sname in ['DE地派服务费', '上架费', '尾程运费', '头程运费', 'COD(2)']:
    if sname not in wb.sheetnames:
        print(f"Sheet [{sname}] 不存在")
        continue
    ws = wb[sname]
    print(f"\n=== Sheet [{sname}] 前15行全部内容 ===")
    for r in range(1, 16):
        vals = []
        for c in range(1, 22):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(f"C{c}={v}")
        if vals:
            print(f"  Row {r:2d}: {' | '.join(vals)}")
        else:
            print(f"  Row {r:2d}: (空行)")
wb.close()
