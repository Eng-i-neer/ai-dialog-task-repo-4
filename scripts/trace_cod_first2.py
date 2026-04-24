"""追踪模板COD回款前两行运单在模板和输入文件中所有出现的位置"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

# 1) 从模板COD回款取前两行运单号
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws_cod = tmpl.worksheets[0]
print(f"模板 Sheet[0]: {ws_cod.title}")
print(f"表头: {[ws_cod.cell(1, c).value for c in range(1, ws_cod.max_column + 1)]}\n")

targets = []
for r in [2, 3]:
    wb_id = None
    for c in range(1, ws_cod.max_column + 1):
        v = ws_cod.cell(r, c).value
        if isinstance(v, str) and v.startswith('DE'):
            wb_id = v
            break
    if not wb_id:
        # 尝试所有列找运单号
        for c in range(1, ws_cod.max_column + 1):
            v = str(ws_cod.cell(r, c).value or '')
            if len(v) > 10 and ('DE' in v or 'IT' in v or 'HR' in v):
                wb_id = v
                break
    targets.append(wb_id)
    print(f"模板 COD R{r} 全部列值:")
    for c in range(1, ws_cod.max_column + 1):
        hdr = ws_cod.cell(1, c).value or f'C{c}'
        val = ws_cod.cell(r, c).value
        is_f = isinstance(val, str) and str(val).startswith('=')
        print(f"  C{c} ({hdr}): {repr(val)} {'[FORMULA]' if is_f else ''}")
    print()

print(f"追踪目标: {targets}\n")
print("="*80)

# 2) 在模板所有Sheet中查找
print("\n>>> 在模板所有Sheet中查找 <<<")
for sname in tmpl.sheetnames:
    ws = tmpl[sname]
    for t in targets:
        if not t:
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 25)):
                if str(ws.cell(r, c).value or '') == t:
                    row_data = []
                    for cc in range(1, min(ws.max_column + 1, 22)):
                        v = ws.cell(r, cc).value
                        if v is not None:
                            hdr = ws.cell(1, cc).value or f'C{cc}'
                            row_data.append(f"{hdr}={v}")
                    print(f"  [{sname}] R{r}: {t}")
                    print(f"    {' | '.join(row_data[:12])}")
                    break
tmpl.close()

# 3) 在输入文件所有Sheet中查找
print("\n" + "="*80)
print("\n>>> 在输入文件所有Sheet中查找 <<<")
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
for sname in inp.sheetnames:
    ws = inp[sname]
    for t in targets:
        if not t:
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 20)):
                if str(ws.cell(r, c).value or '') == t:
                    row_data = []
                    for cc in range(1, min(ws.max_column + 1, 16)):
                        v = ws.cell(r, cc).value
                        if v is not None:
                            row_data.append(f"C{cc}={v}")
                    print(f"  [{sname}] R{r}: {t}")
                    print(f"    {' | '.join(row_data[:12])}")
                    break
inp.close()
