import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
OUTPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
out = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)

# Output sheet 2: 运费
ws_out = out[out.sheetnames[1]]
print("=== OUTPUT 运费 HEADER ===")
for c in range(1, 21):
    v = ws_out.cell(1, c).value
    if v:
        print(f"  Col {c}: {v}")

print("\n=== OUTPUT 运费 Row 2 (first data) ===")
for c in range(1, 21):
    v = ws_out.cell(2, c).value
    if v is not None:
        print(f"  Col {c}: {repr(v)}")

# Input sheets that could feed freight: 头程运费, 尾程运费
for sname in inp.sheetnames:
    if '头程' in sname or '尾程运' in sname:
        ws = inp[sname]
        print(f"\n=== INPUT {sname} HEADER (Row 9) ===")
        for c in range(1, 19):
            v = ws.cell(9, c).value
            if v:
                print(f"  Col {c}: {v}")
        print(f"\n=== INPUT {sname} Row 10 ===")
        for c in range(1, 19):
            v = ws.cell(10, c).value
            if v is not None:
                print(f"  Col {c}: {repr(v)}")

# Try to match output freight row by waybill
out_wb2 = ws_out.cell(2, 4).value
print(f"\n=== Looking for {out_wb2} in input sheets ===")
for sname in inp.sheetnames:
    ws = inp[sname]
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 3).value == out_wb2:
            print(f"  Found in {sname} row {r}")
            for c in range(1, 19):
                v = ws.cell(r, c).value
                if v is not None:
                    print(f"    Col {c}: {repr(v)}")

# Check: which input sheets contribute to output freight
print("\n=== MATCHING ANALYSIS: Output freight rows vs Input sheets ===")
out_waybills = []
for r in range(2, min(ws_out.max_row + 1, 510)):
    wb = ws_out.cell(r, 4).value
    if wb:
        out_waybills.append(wb)

for sname in inp.sheetnames:
    ws = inp[sname]
    in_wbs = set()
    for r in range(10, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if v:
            in_wbs.add(v)
    matched = sum(1 for w in out_waybills if w in in_wbs)
    if matched > 0:
        print(f"  {sname}: {matched}/{len(out_waybills)} matched")
