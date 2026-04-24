"""调查最后两个差异"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

targets = ['DE12603111810003', 'DE12602091410035']

# 1) 看模板中这两行的详细信息
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws_t = tmpl['20260330期尾程杂费']
for t in targets:
    for r in range(2, ws_t.max_row + 1):
        if str(ws_t.cell(r, 2).value or '') == t:
            print(f"\n=== 模板 {t} Row {r} ===")
            for c in range(1, 21):
                v = ws_t.cell(r, c).value
                hdr = ws_t.cell(1, c).value or f'C{c}'
                print(f"  C{c} ({hdr}): {repr(v)}")
            break
tmpl.close()

# 2) 看输入文件中这两个运单的所有费用
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
for t in targets:
    print(f"\n=== 输入文件中 {t} ===")
    for sname in inp.sheetnames:
        ws = inp[sname]
        for r in range(2, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 20)):
                if str(ws.cell(r, c).value or '') == t:
                    vals = [ws.cell(r, cc).value for cc in range(1, ws.max_column + 1)]
                    print(f"  Sheet '{sname}' Row {r}: {vals[:15]}")
                    break
inp.close()
