"""
Deep analysis of head freight:
1. What are GS/SC/IC? How do they map to 普货/特货/敏感货?
2. The pricing file says head freight is 9.7/11.5/12 per KG for most countries.
   But templates use 7.7/9.5/10. Why?
3. Trace the complete chain from input to output.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')

# Check all input files for head freight rates in the C14 "计算公式" column
inp_dir = BASE / '中介提供'
print("="*100)
print("输入文件中的头程运费 — 计算公式列(C14) 分析")
print("="*100)

all_rates = {}
for f in sorted(inp_dir.glob('*.xlsx')):
    wb = openpyxl.load_workbook(str(f), data_only=True)
    for sname in wb.sheetnames:
        if '头程运费' not in sname:
            continue
        ws = wb[sname]
        hr = None
        for r in range(1, 15):
            v = ws.cell(r, 3).value
            if v and '运单' in str(v):
                hr = r
                break
        if not hr:
            continue

        formulas = {}
        for r in range(hr+1, ws.max_row+1):
            wb_id = ws.cell(r, 3).value
            if not wb_id:
                break
            dest = ws.cell(r, 11).value
            formula_str = ws.cell(r, 14).value
            amt = ws.cell(r, 13).value
            weight = ws.cell(r, 8).value
            route = ws.cell(r, 5).value

            if formula_str:
                fs = str(formula_str).strip()
                # extract the rate from "X * Y" pattern
                parts = fs.split('*')
                if len(parts) == 2:
                    try:
                        rate = float(parts[0].strip())
                        w = float(parts[1].strip())
                    except:
                        rate = None
                        w = None
                    if rate:
                        key = (str(dest or '?').strip(), rate)
                        if key not in formulas:
                            formulas[key] = 0
                        formulas[key] += 1

        print(f"\n{f.name} [{sname}]:")
        for (dest, rate), cnt in sorted(formulas.items()):
            print(f"  {dest}: 费率={rate}/KG × {cnt}条")
            all_rates.setdefault(dest, set()).add(rate)
    wb.close()

print(f"\n\n{'='*100}")
print("汇总: 每个国家在输入文件中出现过的头程费率")
print("="*100)
for dest, rates in sorted(all_rates.items()):
    print(f"  {dest}: {sorted(rates)}")


# Now check what the pricing file says
print(f"\n\n{'='*100}")
print("报价文件标准头程费率（按国家和货物类型）")
print("="*100)

pricing_path = BASE / '汇森国际-东欧COD报价20260331生效(5).xlsx'
wb = openpyxl.load_workbook(str(pricing_path), data_only=True)
ws = wb['欧洲COD']
current_country = None
for r in range(6, 48):
    country = ws.cell(r, 1).value
    cat = ws.cell(r, 3).value
    head = ws.cell(r, 4).value
    if country:
        current_country = str(country).strip()
    if cat and head:
        print(f"  {current_country}: {cat} = {head} EUR/KG")
wb.close()


print(f"\n\n{'='*100}")
print("关键发现: 输入费率 vs 报价费率 vs 模板费率 对比")
print("="*100)

# Pricing file rates
pricing_head = {
    '普货': 9.7, '特货': 11.5, '敏感货': 12,
}
# Special: 西班牙/葡萄牙
pricing_head_es = {
    '普货': 10.2, '特货': 12, '敏感货': 13,
}

# Template rates (from IFS formulas)
template_rates = {'GS': 7.7, 'SC': 9.5, 'IC': 10}

# Input rates from the data
input_rates_map = {
    8: '? (most common input rate)',
}

print(f"\n  报价文件(给客户看的报价):")
print(f"    大部分国家: 普货=9.7, 特货=11.5, 敏感货=12 EUR/KG")
print(f"    西班牙/葡萄牙: 普货=10.2, 特货=12, 敏感货=13 EUR/KG")

print(f"\n  模板公式(实际收费):")
print(f"    GS=7.7, SC=9.5, IC=10 EUR/KG")

print(f"\n  输入文件(代理收到的成本):")
for dest, rates in sorted(all_rates.items()):
    print(f"    {dest}: {sorted(rates)} EUR/KG")

print(f"\n  推论:")
print(f"    报价文件(9.7/11.5/12)是给终端客户的'标准报价/牌价'")
print(f"    模板中GS=7.7/SC=9.5/IC=10是'舅妈给客户的实际协议价'(低于牌价)")
print(f"    输入文件中8/9.5/10等是'中邮给鑫腾跃(舅妈公司)的成本价'")
print(f"    差价即为舅妈的利润")
