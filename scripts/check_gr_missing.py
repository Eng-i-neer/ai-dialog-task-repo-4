"""检查GR两条在模板COD中存在但输入COD中不存在的运单"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

targets = ['GR12602021510039', 'GR12602021510041']

# 模板中的详情
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=True)
ws = tmpl.worksheets[0]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("=== 模板 COD 中这两条的详情 ===")
for r in range(2, ws.max_row + 1):
    wb_id = ws.cell(r, 4).value
    if str(wb_id) in targets:
        print(f"\nR{r}: {wb_id}")
        for c in range(1, ws.max_column + 1):
            h = headers[c-1] or f'C{c}'
            v = ws.cell(r, c).value
            if v is not None:
                print(f"  {h}: {repr(v)}")
tmpl.close()

# 输入文件中查找
print("\n=== 输入文件中查找 ===")
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
for t in targets:
    print(f"\n运单: {t}")
    found = False
    for sname in inp.sheetnames:
        ws2 = inp[sname]
        for r in range(1, ws2.max_row + 1):
            for c in range(1, min(ws2.max_column + 1, 20)):
                if str(ws2.cell(r, c).value or '') == t:
                    found = True
                    vals = []
                    for cc in range(1, min(ws2.max_column + 1, 16)):
                        v = ws2.cell(r, cc).value
                        if v is not None:
                            vals.append(f"C{cc}={v}")
                    print(f"  [{sname}] R{r}: {' | '.join(vals[:12])}")
                    break
    if not found:
        print(f"  ** 在输入文件中完全不存在 **")
inp.close()
