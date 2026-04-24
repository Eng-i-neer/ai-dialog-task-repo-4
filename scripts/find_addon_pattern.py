"""找到addon=30 vs 11.91的真正规则: 检查是否与尾程计费重有关"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# 获取尾程运费sheet中的计费重
tail_charge_weight = {}
for sname in inp.sheetnames:
    if '尾程运费' in sname:
        ws = inp[sname]
        for r in range(10, ws.max_row + 1):
            wb = ws.cell(r, 3).value
            cw = ws.cell(r, 8).value
            if wb:
                tail_charge_weight[wb] = cw

ws_freight = ref[ref.sheetnames[1]]

# 分析：当tail_qty=1时，addon分别是30还是11.91
print("=== tail_qty=1时，addon=30 vs 11.91 ===")
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    tail_qty = ws_freight.cell(r, 12).value
    if not wb or not addon or tail_qty != 1:
        continue
    tcw = tail_charge_weight.get(wb, 'N/A')
    actual = ws_freight.cell(r, 9).value
    hw = ws_freight.cell(r, 10).value
    print(f"  {wb}: addon={addon}, actual_w={actual}, head_w={hw}, tail_charge_w={tcw}")

# 分析：当tail_qty=2时，addon=30的少数情况
print("\n=== tail_qty=2时，addon=30的记录 ===")
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    tail_qty = ws_freight.cell(r, 12).value
    if not wb or not addon or tail_qty != 2 or abs(addon - 30) > 0.1:
        continue
    tcw = tail_charge_weight.get(wb, 'N/A')
    actual = ws_freight.cell(r, 9).value
    hw = ws_freight.cell(r, 10).value
    dest = ws_freight.cell(r, 7).value
    print(f"  {wb}: addon={addon}, actual_w={actual}, head_w={hw}, tail_charge_w={tcw}, dest={dest}")
