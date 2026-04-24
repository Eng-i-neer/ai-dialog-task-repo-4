# -*- coding: utf-8 -*-
"""Fast scan: sample files across different periods, only read first 12 rows of each sheet."""
import openpyxl, os, sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单'

files = []
for dp, dn, fns in os.walk(BASE):
    for f in fns:
        if '李志' in f and f.endswith('.xlsx') and not f.startswith('~$'):
            files.append(os.path.join(dp, f))
files.sort()

# Sample: early, mid, late periods
sample_indices = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, len(files)//2, len(files)-3, len(files)-2, len(files)-1]
sample_indices = sorted(set(i for i in sample_indices if 0 <= i < len(files)))
sample_files = [files[i] for i in sample_indices]

print(f"总共 {len(files)} 个文件, 抽样 {len(sample_files)} 个")

HEADER_KEYWORDS = ['序号', '运单号码', '运单号', '寄件日期', '目的地', '原币金额', '科目']

layout_map = defaultdict(list)
all_sheet_names = set()
all_col_names = set()
header_rows = defaultdict(int)

for fpath in sample_files:
    fname = os.path.basename(fpath)
    period = ''
    for part in fname.split('对账单'):
        if len(part) >= 8:
            period = part[:8]
    if not period:
        import re
        m = re.search(r'(\d{8})', fname)
        period = m.group(1) if m else fname

    print(f"\n{'='*80}")
    print(f"[{period}] {fname}")

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        continue

    for sname in wb.sheetnames:
        all_sheet_names.add(sname)
        ws = wb[sname]

        # Read first 12 rows to find header
        rows_data = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=22, values_only=False), 1):
            cells = [(cell.column, cell.value) for cell in row if cell.value is not None]
            rows_data.append((r_idx, cells))

        # Find header row: the one with >=3 header keywords
        hr = None
        hr_cols = []
        for r_idx, cells in rows_data:
            vals = [str(v).strip() for _, v in cells]
            text = ' '.join(vals)
            matches = sum(1 for kw in HEADER_KEYWORDS if kw in text)
            if matches >= 3:
                hr = r_idx
                hr_cols = [(c, str(v).strip()) for c, v in cells]
                break

        if not hr:
            continue

        header_rows[hr] += 1
        col_names = [name for _, name in hr_cols]
        for cn in col_names:
            all_col_names.add(cn)

        layout_key = tuple(col_names)
        layout_map[layout_key].append(f"{period}:{sname}")

        # Also read first data row
        first_data = None
        for r_idx, cells in rows_data:
            if r_idx == hr + 1 and cells:
                first_data = cells
                break

        # Print compact
        col_str = ', '.join(f"C{c}:{n}" for c, n in hr_cols)
        print(f"  [{sname}] hr={hr} | {col_str}")
        if first_data:
            # Show waybill column value
            for c, v in first_data:
                for hc, hn in hr_cols:
                    if hc == c and '运单' in hn:
                        print(f"    首行运单: {v}")

    wb.close()

print(f"\n{'='*80}")
print(f"所有 Sheet 名称 ({len(all_sheet_names)}):")
for s in sorted(all_sheet_names):
    print(f"  {s}")

print(f"\n表头行位置:")
for row, cnt in sorted(header_rows.items()):
    print(f"  Row {row}: {cnt}次")

print(f"\n所有列名:")
for c in sorted(all_col_names):
    print(f"  '{c}'")

print(f"\n列布局种类: {len(layout_map)}")
for i, (layout, occs) in enumerate(sorted(layout_map.items(), key=lambda x: -len(x[1]))):
    print(f"\n--- 布局{i+1} ({len(occs)}次) ---")
    print(f"  列: {list(layout)}")
    for occ in occs[:5]:
        print(f"    {occ}")
    if len(occs) > 5:
        print(f"    ...另外{len(occs)-5}个")
