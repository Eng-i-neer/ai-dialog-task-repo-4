import openpyxl, os

desktop = os.path.expanduser('~/Desktop')
gen_files = sorted([f for f in os.listdir(desktop) if f.startswith('20260330-v4-')], reverse=True)
GEN_FILE = os.path.join(desktop, gen_files[0])
print(f"File: {gen_files[0]}")

wb = openpyxl.load_workbook(GEN_FILE)
ws = wb[wb.sheetnames[0]]

print("\n=== Generated file COD Sheet Row 2-5 ===")
for r in range(2, 6):
    print(f"  Row {r}:")
    for c in range(1, 15):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"    Col {c}: {repr(v)[:80]}")
