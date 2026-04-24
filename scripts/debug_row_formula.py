"""检查生成文件中 HR12603041510002 所在行的公式状态"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

GEN = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\20260330-汇森李志（东欧）对账单-自动生成v3.xlsx'
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

# 生成文件 - 公式模式
gen_wb = openpyxl.load_workbook(GEN, data_only=False)
ws = gen_wb['20260330期尾程杂费']

# 找到 HR12603041510002 在生成文件中的行
target = 'HR12603041510002'
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 2).value or '') == target:
        print(f"生成文件中 {target} 在 Row {r}")
        print(f"  全部列值:")
        for c in range(1, 21):
            v = ws.cell(r, c).value
            hdr = ws.cell(1, c).value or f'C{c}'
            is_f = isinstance(v, str) and v.startswith('=')
            print(f"    C{c} ({hdr}): {repr(v)} {'[FORMULA]' if is_f else ''}")
        
        # 检查模板中同一行号的公式
        tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
        ws_t = tmpl_wb['20260330期尾程杂费']
        print(f"\n  模板同行号 Row {r} 的公式:")
        for c in [12, 13, 17]:
            v_t = ws_t.cell(r, c).value
            hdr = ws_t.cell(1, c).value
            print(f"    C{c} ({hdr}): {repr(v_t)}")
        
        # 检查模板中有公式的行范围
        print(f"\n  模板公式列C12范围检查:")
        has_formula_rows = []
        no_formula_rows = []
        for rr in range(2, min(ws_t.max_row + 1, 510)):
            v12 = ws_t.cell(rr, 12).value
            if isinstance(v12, str) and v12.startswith('='):
                has_formula_rows.append(rr)
            elif v12 is not None:
                no_formula_rows.append((rr, v12))
        print(f"    有公式的行: R{min(has_formula_rows)}-R{max(has_formula_rows)} (共{len(has_formula_rows)}行)")
        print(f"    有值但非公式的行: {no_formula_rows[:10]}")
        
        # 检查该行是否被 clear_data_rows 的逻辑影响
        # clear_data_rows 只清非公式值，公式保留
        # 但如果模板该行C12本来就没有公式也没有值呢？
        print(f"\n  模板 Row {r} C12 原始值: {repr(ws_t.cell(r, 12).value)}")
        
        tmpl_wb.close()
        break

gen_wb.close()
