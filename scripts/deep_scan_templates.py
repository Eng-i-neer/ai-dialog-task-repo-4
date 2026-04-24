"""Deep scan all templates: full column headers, formula positions, data structure.
Outputs a structured JSON config for each template."""
import sys, io, os, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE_DIR = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\反馈客户\原始模板'
OUTPUT_JSON = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\scripts\template_configs.json'

configs = {}

for fname in sorted(os.listdir(TEMPLATE_DIR)):
    if not fname.endswith('.xlsx'):
        continue

    fpath = os.path.join(TEMPLATE_DIR, fname)
    wb_formula = openpyxl.load_workbook(fpath, data_only=False)
    wb_value = openpyxl.load_workbook(fpath, data_only=True)

    # Derive customer name from filename
    if '李志' in fname:
        customer = '李志'
    else:
        m = re.search(r'汇森[—-](.+?)（', fname)
        customer = m.group(1) if m else fname

    template_info = {
        'filename': fname,
        'customer': customer,
        'sheets': [],
    }

    for si, sname in enumerate(wb_formula.sheetnames):
        ws_f = wb_formula[sname]
        ws_v = wb_value[sname]

        # Full headers
        headers = {}
        for c in range(1, ws_f.max_column + 1):
            h = ws_f.cell(1, c).value
            if h is not None:
                headers[str(c)] = str(h)

        # Find data end row
        data_end = 0
        for r in range(2, ws_f.max_row + 1):
            vals = [ws_f.cell(r, c).value for c in range(1, min(ws_f.max_column + 1, 10))]
            if all(v is None for v in vals):
                break
            if any('合计' in str(v or '') for v in vals):
                break
            data_end = r

        data_rows = data_end - 1 if data_end >= 2 else 0

        # Scan formula columns (check rows 2-min(6, data_end))
        formula_cols = {}
        for r in range(2, min(data_end + 1, 8)):
            for c in range(1, ws_f.max_column + 1):
                v = ws_f.cell(r, c).value
                if isinstance(v, str) and v.startswith('='):
                    col_key = str(c)
                    if col_key not in formula_cols:
                        formula_cols[col_key] = {
                            'first_row': r,
                            'sample': v[:80],
                            'header': headers.get(col_key, f'C{c}'),
                        }

        # For COD sheet, extract exchange rate
        rate = None
        if 'COD' in sname:
            for r in range(2, min(data_end + 1, 20)):
                v = ws_v.cell(r, 10).value
                if v and isinstance(v, (int, float)) and v > 1:
                    rate = v
                    break
            if rate is None:
                for r in range(2, min(data_end + 1, 20)):
                    for c in range(1, ws_v.max_column + 1):
                        v = ws_v.cell(r, c).value
                        if isinstance(v, (int, float)) and 7 < v < 9:
                            rate = v
                            break
                    if rate:
                        break

        # For 汇总 sheet, extract rate
        if '汇总' in sname:
            for r in range(1, min(ws_v.max_row + 1, 10)):
                for c in range(1, min(ws_v.max_column + 1, 10)):
                    v = ws_v.cell(r, c).value
                    if isinstance(v, (int, float)) and 7 < v < 9:
                        rate = v

        # Customer code (C1 of COD or 汇总)
        customer_code = None
        if 'COD' in sname or '汇总' in sname:
            v1 = ws_v.cell(2, 1).value
            if isinstance(v1, (int, float)):
                customer_code = v1

        sheet_info = {
            'index': si,
            'name': sname,
            'data_rows': data_rows,
            'max_column': ws_f.max_column,
            'headers': headers,
            'formula_cols': formula_cols,
        }
        if rate:
            sheet_info['exchange_rate'] = rate
        if customer_code:
            sheet_info['customer_code'] = customer_code

        template_info['sheets'].append(sheet_info)

    configs[customer] = template_info
    wb_formula.close()
    wb_value.close()
    print(f"Scanned: {customer} ({fname}) - {len(template_info['sheets'])} sheets")

# Save JSON
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(configs, f, ensure_ascii=False, indent=2)
print(f"\nSaved configs to {OUTPUT_JSON}")

# Print summary comparison
print("\n" + "="*100)
print("TEMPLATE STRUCTURE COMPARISON")
print("="*100)

ref = configs.get('李志')
for customer, info in configs.items():
    print(f"\n--- {customer} ({info['filename']}) ---")
    for sh in info['sheets']:
        sheet_type = ''
        if 'COD' in sh['name']:
            sheet_type = '[COD]'
        elif '运费' in sh['name']:
            sheet_type = '[FREIGHT]'
        elif '杂费' in sh['name']:
            sheet_type = '[SURCHARGE]'
        elif '汇总' in sh['name']:
            sheet_type = '[SUMMARY]'
        elif '增值税' in sh['name']:
            sheet_type = '[VAT]'
        elif 'F' in sh['name'] and '附加' in sh['name']:
            sheet_type = '[F-SURCHARGE]'
        elif '仓储' in sh['name']:
            sheet_type = '[STORAGE]'
        elif '理赔' in sh['name']:
            sheet_type = '[CLAIM]'

        hdrs = list(sh['headers'].values())
        formula_info = ', '.join(f"C{k}={v['header']}" for k, v in sh['formula_cols'].items())
        rate_str = f" rate={sh['exchange_rate']}" if 'exchange_rate' in sh else ''
        print(f"  [{sh['index']}] {sh['name']} {sheet_type}: {sh['data_rows']} rows, {sh['max_column']} cols{rate_str}")
        print(f"      Headers: {hdrs[:14]}")
        if formula_info:
            print(f"      Formulas: {formula_info}")
