"""分析上架费的来源：哪些运单有上架费，来自输入文件的哪个Sheet"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

# 1) 从输入文件的"上架费"Sheet提取所有有上架费的运单
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
print("输入文件所有Sheet:")
for i, name in enumerate(inp.sheetnames):
    print(f"  [{i}] {name}")

shelf_fee_waybills = set()
shelf_sheet = None
for sname in inp.sheetnames:
    if '上架' in sname:
        shelf_sheet = sname
        ws = inp[sname]
        print(f"\n=== 输入文件 Sheet: {sname} ===")
        print(f"  总行数: {ws.max_row}")
        # 看表头
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"  表头: {headers[:15]}")
        # 提取运单号（通常在C列）
        waybill_col = None
        for c in range(1, len(headers) + 1):
            h = str(headers[c-1] or '')
            if '运单' in h or '单号' in h:
                waybill_col = c
                break
        if not waybill_col:
            waybill_col = 3  # 默认C列
        
        print(f"  运单号列: C{waybill_col}")
        print(f"  前5行数据:")
        for r in range(2, min(ws.max_row + 1, 7)):
            wb_id = ws.cell(r, waybill_col).value
            amount = ws.cell(r, 13).value  # M列通常是金额
            fee_name = ws.cell(r, 12).value  # L列通常是费用名称
            print(f"    R{r}: 运单={wb_id} 费用名={fee_name} 金额={amount}")
            if wb_id:
                shelf_fee_waybills.add(str(wb_id))
        
        # 提取所有运单
        for r in range(2, ws.max_row + 1):
            wb_id = ws.cell(r, waybill_col).value
            if wb_id:
                shelf_fee_waybills.add(str(wb_id))

print(f"\n输入文件中有上架费的运单总数: {len(shelf_fee_waybills)}")

# 2) 从模板的"尾程杂费"Sheet看哪些运单有上架费
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws_t = tmpl['20260330期尾程杂费']

tmpl_has_shelf = set()
tmpl_no_shelf = set()
for r in range(2, ws_t.max_row + 1):
    wb_id = ws_t.cell(r, 2).value
    c12 = ws_t.cell(r, 12).value
    if not wb_id or wb_id == '合计：':
        continue
    if c12 is not None:
        tmpl_has_shelf.add(str(wb_id))
    else:
        tmpl_no_shelf.add(str(wb_id))

print(f"\n模板尾程杂费中有上架费(C12非空)的运单: {len(tmpl_has_shelf)}")
print(f"模板尾程杂费中无上架费(C12空)的运单: {len(tmpl_no_shelf)}")

# 3) 交叉对比
in_input_and_template = shelf_fee_waybills & tmpl_has_shelf
in_input_not_template = shelf_fee_waybills - tmpl_has_shelf
in_template_not_input = tmpl_has_shelf - shelf_fee_waybills

print(f"\n交叉对比:")
print(f"  输入有上架费 且 模板也有: {len(in_input_and_template)}")
print(f"  输入有上架费 但 模板没有: {len(in_input_not_template)}")
if in_input_not_template:
    for wb in sorted(in_input_not_template)[:5]:
        print(f"    {wb}")
print(f"  模板有上架费 但 输入没有: {len(in_template_not_input)}")
if in_template_not_input:
    for wb in sorted(in_template_not_input)[:10]:
        # 看看模板中这些运单的备注
        for r in range(2, ws_t.max_row + 1):
            if str(ws_t.cell(r, 2).value or '') == wb:
                remark = ws_t.cell(r, 18).value
                c12 = ws_t.cell(r, 12).value
                c13 = ws_t.cell(r, 13).value
                print(f"    {wb}: C12={c12} C13={c13} 备注={remark}")
                break

# 4) 看看模板中"无上架费"的运单都有什么费用
print(f"\n模板中无上架费的运单的备注分布:")
remark_dist = {}
for r in range(2, ws_t.max_row + 1):
    wb_id = ws_t.cell(r, 2).value
    c12 = ws_t.cell(r, 12).value
    remark = ws_t.cell(r, 18).value
    if wb_id and wb_id != '合计：' and c12 is None:
        remark_dist[remark] = remark_dist.get(remark, 0) + 1
for remark, count in sorted(remark_dist.items(), key=lambda x: -x[1]):
    print(f"  {remark}: {count}条")

inp.close()
tmpl.close()
