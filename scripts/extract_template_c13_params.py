"""从模板的 C13 公式中提取实际使用的参数，按国家分组"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws = tmpl['20260330期尾程杂费']

# 提取公式中的 base 和 step
# 格式: =ROUND(IF(G_>2,BASE+(ROUNDUP(G_,0)-2)*STEP,BASE)*RATE[*70%],2)
pattern = re.compile(
    r'=ROUND\(IF\(G\d+>2,([0-9.]+)\+\(ROUNDUP\(G\d+,0\)-2\)\*([0-9.]+),([0-9.]+)\)\*([0-9.]+)(\*70%)?,2\)'
)

by_dest = {}
for r in range(2, 510):
    v13 = ws.cell(r, 13).value
    dest = ws.cell(r, 10).value
    if not isinstance(v13, str) or not v13.startswith('='):
        continue
    
    m = pattern.match(v13)
    if m:
        base = float(m.group(1))
        step = float(m.group(2))
        rate = float(m.group(4))
        is_70 = m.group(5) is not None
        key = (dest, base, step, is_70)
        if key not in by_dest:
            by_dest[key] = 0
        by_dest[key] += 1
    else:
        print(f"  R{r} 无法解析: {v13}")

print(f"{'目的地':<12} {'首2KG':>8} {'续1KG':>8} {'70%?':>6} {'行数':>6}")
print("-"*50)
for (dest, base, step, is_70), count in sorted(by_dest.items()):
    pct = '是' if is_70 else '否'
    print(f"{dest:<12} {base:>8} {step:>8} {pct:>6} {count:>6}")

tmpl.close()
