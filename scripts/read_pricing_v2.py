"""读取报价文件 - 用 UTF-8 输出"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
wb = openpyxl.load_workbook(PRICING, data_only=True)

print(f"Sheets: {wb.sheetnames}")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*70}")
    print(f"  Sheet: {sname}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"{'='*70}")
    
    for r in range(1, min(ws.max_row + 1, 100)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = str(v).replace('\n', '\\n')[:80]
                row_data.append(f"C{c}={val_str}")
        if row_data:
            print(f"  R{r}: {' | '.join(row_data)}")
    
    if ws.max_row > 99:
        print(f"  ... (showing first 99 of {ws.max_row} rows)")
