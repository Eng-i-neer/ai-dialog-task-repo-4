"""重新检查模板的真实结构 — 精确到每一个单元格"""
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
wb = openpyxl.load_workbook(TEMPLATE)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*70}")
    print(f"  Sheet: {sname}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"{'='*70}")

    # Print header row 1 - exact values
    print("  ROW 1 (header):")
    for c in range(1, min(ws.max_column + 1, 25)):
        v = ws.cell(1, c).value
        if v is not None:
            # Show unicode codepoints
            if isinstance(v, str):
                codes = ' '.join(f'U+{ord(ch):04X}' for ch in v)
                print(f"    Col {c}: '{v}' [{codes}]")
            else:
                print(f"    Col {c}: {repr(v)}")

    # Print first 3 data rows
    for dr in range(2, min(5, ws.max_row + 1)):
        print(f"  ROW {dr} (data):")
        for c in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(dr, c).value
            if v is not None:
                print(f"    Col {c}: {repr(v)}")

    # Check if there's a DIFFERENT header structure (some templates use row 0 differently)
    # Also check date format
    for dr in [2, 3]:
        cell_c = ws.cell(dr, 3)
        if cell_c.value:
            print(f"  ROW {dr} Col C number_format: {cell_c.number_format}")
            print(f"  ROW {dr} Col C value type: {type(cell_c.value)}")
