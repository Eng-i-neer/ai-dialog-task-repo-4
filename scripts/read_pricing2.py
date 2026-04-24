# -*- coding: utf-8 -*-
"""Read more rows from the pricing table."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'web'))
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

fp = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\汇森国际-东欧COD报价20250121生效.xlsx'
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
ws = wb['欧洲COD']
for row_idx, row in enumerate(ws.iter_rows(min_row=50, max_row=ws.max_row or 100, values_only=True), 50):
    vals = [str(c) if c else '' for c in row[:15]]
    line = ' | '.join(v[:50] for v in vals)
    if any(v.strip() for v in vals):
        print(f"  R{row_idx}: {line}")
wb.close()
