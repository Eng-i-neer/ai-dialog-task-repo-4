b1 = bytes.fromhex('e8bdace5af84')
b2 = bytes.fromhex('e8bdace58f91')
c1 = b1.decode('utf-8')
c2 = b2.decode('utf-8')
print(f"ref char: {c1} (U+{ord(c1[0]):04X} U+{ord(c1[1]):04X})")
print(f"gen char: {c2} (U+{ord(c2[0]):04X} U+{ord(c2[1]):04X})")

import openpyxl
ref = openpyxl.load_workbook(r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx', data_only=True)
ws = ref[ref.sheetnames[0]]
for r in range(2, 10):
    v = ws.cell(r, 2).value
    if v:
        codes = ' '.join(f'U+{ord(ch):04X}' for ch in v)
        print(f"  Row {r}: {codes}")
