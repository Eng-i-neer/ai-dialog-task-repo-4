"""分析模板中尾程杂费表的合并规则"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
tmpl_f_wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
ws = tmpl_wb['20260330期尾程杂费']
ws_f = tmpl_f_wb['20260330期尾程杂费']

print("="*80)
print("尾程杂费表 - 列结构")
print("="*80)
for c in range(1, ws.max_column + 1):
    v = ws.cell(1, c).value
    print(f"  C{c}: {v}")

# 分析每行的填写模式
print("\n" + "="*80)
print("每行的填充列分析 (非空列)")
print("="*80)

fee_col_patterns = {}
for r in range(2, ws.max_row + 1):
    waybill = ws.cell(r, 2).value
    if not waybill:
        continue
    
    filled_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            hdr = ws.cell(1, c).value or f'C{c}'
            filled_cols[c] = {'header': hdr, 'value': v}
    
    # 特别关注 C12-C16 哪些列有值
    fee_pattern = []
    for c in [12, 13, 14, 15, 16]:
        v = ws.cell(r, c).value
        if v is not None:
            hdr = ws.cell(1, c).value
            fee_pattern.append(f"{hdr}={v}")
    
    pattern_key = tuple(c for c in [12, 13, 14, 15, 16] if ws.cell(r, c).value is not None)
    if pattern_key not in fee_col_patterns:
        fee_col_patterns[pattern_key] = []
    fee_col_patterns[pattern_key].append({
        'row': r,
        'waybill': waybill,
        'fees': fee_pattern,
        'remark': ws.cell(r, 18).value,
    })

print("\n费用列组合模式:")
for pattern, items in sorted(fee_col_patterns.items(), key=lambda x: -len(x[1])):
    col_names = [ws.cell(1, c).value for c in pattern]
    print(f"\n  模式 {pattern} = {col_names}: {len(items)} 条")
    for item in items[:5]:
        print(f"    R{item['row']}: {item['waybill']} | {item['fees']} | 备注={item['remark']}")

# 检查同一运单是否出现多行
print("\n" + "="*80)
print("同一运单出现多行的情况")
print("="*80)
waybill_rows = {}
for r in range(2, ws.max_row + 1):
    wb = ws.cell(r, 2).value
    if wb:
        if wb not in waybill_rows:
            waybill_rows[wb] = []
        waybill_rows[wb].append(r)

multi_row = {k: v for k, v in waybill_rows.items() if len(v) > 1}
print(f"  总运单数: {len(waybill_rows)}")
print(f"  出现多行的运单: {len(multi_row)}")
for wb, rows in list(multi_row.items())[:10]:
    print(f"\n  {wb} 出现在 {len(rows)} 行:")
    for r in rows:
        vals = []
        for c in range(12, 20):
            v = ws.cell(r, c).value
            if v is not None:
                hdr = ws.cell(1, c).value or f'C{c}'
                vals.append(f"{hdr}={v}")
        print(f"    R{r}: {vals}")

# 公式分析
print("\n" + "="*80)
print("公式列分析 (R4-R10)")
print("="*80)
for r in range(2, min(15, ws.max_row + 1)):
    formulas = []
    for c in range(1, ws.max_column + 1):
        v = ws_f.cell(r, c).value
        if isinstance(v, str) and v.startswith('='):
            formulas.append(f"C{c}={v}")
    if formulas:
        print(f"  R{r}: {formulas}")

# 输入文件中看看这些运单的费用来源
print("\n" + "="*80)
print("输入文件费用sheets结构")
print("="*80)
inp_wb = openpyxl.load_workbook(INPUT, data_only=True)
for sname in inp_wb.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws_in = inp_wb[sname]
    header_row = None
    for r in range(1, 15):
        v = ws_in.cell(r, 3).value
        if v and '运单' in str(v):
            header_row = r
            break
    if header_row:
        data_count = 0
        for r in range(header_row + 1, ws_in.max_row + 1):
            if ws_in.cell(r, 3).value:
                data_count += 1
        print(f"  {sname}: {data_count} 条数据")

inp_wb.close()
tmpl_wb.close()
tmpl_f_wb.close()
