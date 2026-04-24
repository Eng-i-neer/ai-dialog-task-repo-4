"""
科目驱动转换器 — 按输入科目决定运单归属，按模板结构自动适配列位置

核心原则:
  1. 输入科目决定运单出现在哪个输出 Sheet（COD→COD回款, 尾程运费→运费Sheet, etc.）
  2. 输出金额由模板内嵌公式计算（客户报价），与代理输入金额无关
  3. 代码只填元数据（日期、运单号、重量、目的地等），公式列保持不动
  4. 列位置按模板表头自动检测，而非硬编码
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from datetime import datetime
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from convert_bill import (
    parse_input, determine_ship_type, is_formula, clear_data_rows,
    ensure_formulas, load_remote_zipcodes, calc_remote_fee_rmb,
    DELIVERY_PRICING, _get_delivery_params, _calc_return_fee_rmb,
    _build_return_fee_formula,
)

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
INPUT_DIR = BASE / '中介提供'
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'
OUTPUT_BASE = BASE / '反馈客户' / '自动生成'
PRICING_FILE = BASE / '报价规则' / '汇森国际-东欧COD报价20260331生效(5).xlsx'

EXCHANGE_RATE = 7.9342
PRICING_VERSION = 'old'

CUSTOMER_MAP = {
    '中文':   '李志',
    '中文1':  '君悦',
    '中文3':  '小美',
    '中文5':  'J',
    '中文6':  '涵江',
    '中文7':  '阿甘',
    '中文9':  '威总',
    '中文12': '峰总',
}

# ── Template structure detection ──

def detect_c5c6_fields(ws):
    """Detect what C5 and C6 represent from headers."""
    h5 = str(ws.cell(1, 5).value or '').strip()
    h6 = str(ws.cell(1, 6).value or '').strip()
    if '转单' in h5:
        return 'transfer', 'cust_ref'
    if '客户' in h5 or '订单' in h5:
        return 'cust_ref', 'transfer'
    return 'cust_ref', 'transfer'


def detect_weight_layout(ws):
    """Detect weight column layout in freight sheet.
    李志: C9=实重 C10=头程计费重 C11=尺寸 C12=尾程计费重 C13=普特敏货
    通用: C9=重量(KG) C10=尺寸 C11=尾程计费重 C12=普特敏货
    """
    h9 = str(ws.cell(1, 9).value or '')
    h10 = str(ws.cell(1, 10).value or '')
    if '实重' in h9 or '头程' in h10:
        return 'lizhi'
    return 'generic'


def find_sheet_by_type(wb, stype):
    """Find a sheet index by type keyword."""
    for i, name in enumerate(wb.sheetnames):
        if stype == 'COD' and 'COD' in name:
            return i
        if stype == '运费' and '运费' in name and '杂费' not in name:
            return i
        if stype == '增值税' and '增值税' in name:
            return i
        if stype == '杂费' and '杂费' in name:
            return i
        if stype == 'F附加费' and ('F' in name.upper() or 'f' in name) and '附加' in name:
            return i
        if stype == '汇总' and '汇总' in name:
            return i
    return None


def find_data_end(ws, check_cols=10):
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, check_cols + 1))]
        if all(v is None for v in vals):
            return r - 1
        if any('合计' in str(v or '') for v in vals):
            return r - 1
    return ws.max_row


def find_col_by_keywords(ws, *keywords):
    """Find column number whose header contains any of the keywords."""
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '')
        if any(kw in h for kw in keywords):
            return c
    return None


# ── Per-category fill functions ──

def fill_cod(wb, parcels, exchange_rate):
    """COD回款: 输入有COD科目的运单 → 填入COD Sheet"""
    idx = find_sheet_by_type(wb, 'COD')
    if idx is None:
        return {'sheet': None, 'rows': 0, 'skip': 'no COD sheet'}

    ws = wb.worksheets[idx]
    data_end = find_data_end(ws, 14)
    clear_data_rows(ws, 2, data_end)

    c5_type, c6_type = detect_c5c6_fields(ws)

    cod_parcels = sorted(
        [p for p in parcels.values() if 'cod' in p['fees']],
        key=lambda p: p.get('date') or datetime.min
    )

    for i, p in enumerate(cod_parcels):
        row = 2 + i
        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        ws.cell(row, 5, p.get(c5_type, ''))
        ws.cell(row, 6, p.get(c6_type, ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))
        ws.cell(row, 9, p['fees']['cod']['amount_eur'])
        ws.cell(row, 10, exchange_rate)

    if cod_parcels:
        formula_source = None
        for r in range(2, min(data_end + 1, 20)):
            if is_formula(ws.cell(r, 11).value):
                formula_source = r
                break
        if formula_source:
            all_rows = list(range(2, 2 + len(cod_parcels)))
            ensure_formulas(ws, formula_source, all_rows, [11, 12, 13])

    return {'sheet': wb.sheetnames[idx], 'rows': len(cod_parcels)}


def fill_freight(wb, parcels, exchange_rate):
    """运费: 输入有 尾程运费 科目的运单 → 填入运费 Sheet
    只填元数据列，公式列（头程/尾程运费/COD手续费/附加费/小计）由模板公式自动计算
    """
    idx = find_sheet_by_type(wb, '运费')
    if idx is None:
        return {'sheet': None, 'rows': 0, 'skip': 'no freight sheet'}

    ws = wb.worksheets[idx]
    data_end = find_data_end(ws, 12)
    clear_data_rows(ws, 2, data_end)

    c5_type, c6_type = detect_c5c6_fields(ws)
    weight_layout = detect_weight_layout(ws)

    freight_parcels = sorted(
        [p for p in parcels.values() if 'tail_freight' in p['fees']],
        key=lambda p: p.get('date') or datetime.min
    )

    for i, p in enumerate(freight_parcels):
        row = 2 + i
        head_w = p.get('head_charge_weight') or p.get('charge_weight') or 0
        tail_q = p.get('tail_charge_weight') or p.get('charge_weight') or 0

        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        ws.cell(row, 5, p.get(c5_type, ''))
        ws.cell(row, 6, p.get(c6_type, ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))

        if weight_layout == 'lizhi':
            ws.cell(row, 9, p.get('actual_weight'))
            ws.cell(row, 10, head_w)
            ws.cell(row, 11, p.get('dimensions', ''))
            ws.cell(row, 12, tail_q)
            ws.cell(row, 13, 'IC')
        else:
            ws.cell(row, 9, p.get('actual_weight'))
            ws.cell(row, 10, p.get('dimensions', ''))
            ws.cell(row, 11, tail_q)
            ws.cell(row, 12, 'IC')

    return {'sheet': wb.sheetnames[idx], 'rows': len(freight_parcels)}


def fill_vat(wb, parcels, exchange_rate):
    """增值税&清关费: 输入有 目的地增值税 科目的运单 → 填入增值税 Sheet
    增值税公式列由模板自行计算，只填元数据
    """
    idx = find_sheet_by_type(wb, '增值税')
    if idx is None:
        return {'sheet': None, 'rows': 0, 'skip': 'no VAT sheet'}

    ws = wb.worksheets[idx]
    data_end = find_data_end(ws, 10)
    clear_data_rows(ws, 2, data_end)

    c5_type, c6_type = detect_c5c6_fields(ws)

    vat_parcels = sorted(
        [p for p in parcels.values() if 'vat' in p['fees']],
        key=lambda p: p.get('date') or datetime.min
    )

    for i, p in enumerate(vat_parcels):
        row = 2 + i
        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        ws.cell(row, 5, p.get(c5_type, ''))
        ws.cell(row, 6, p.get(c6_type, ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))
        ws.cell(row, 9, p.get('actual_weight'))
        ws.cell(row, 10, p.get('dimensions', ''))

    return {'sheet': wb.sheetnames[idx], 'rows': len(vat_parcels)}


def fill_surcharge(wb, parcels, exchange_rate, pricing_version,
                   remote_zips=None, parcel_zipcodes=None):
    """尾程杂费: 输入有 上架费/退件费/拒收返程费 科目的运单 → 填入杂费 Sheet
    同一运单合并为一行，不同费用填到不同列
    """
    idx = find_sheet_by_type(wb, '杂费')
    if idx is None:
        return {'sheet': None, 'rows': 0, 'skip': 'no surcharge sheet'}

    ws = wb.worksheets[idx]
    data_end = find_data_end(ws, 10)
    clear_data_rows(ws, 2, data_end)

    # Auto-detect fee columns
    shelf_col = find_col_by_keywords(ws, '上架')
    return_col = find_col_by_keywords(ws, '退件', '返程', '拒收')
    remote_col = find_col_by_keywords(ws, '偏远')
    vat_col = find_col_by_keywords(ws, '增值税')
    subtotal_col = find_col_by_keywords(ws, '小计')
    remark_col = find_col_by_keywords(ws, '备注')
    zip_col = find_col_by_keywords(ws, '邮编')

    if not remark_col and subtotal_col:
        remark_col = subtotal_col + 1
    if not zip_col and remark_col:
        zip_col = remark_col + 1

    surcharge_triggers = {'shelf_fee', 'tail_return_fee', 'tail_return_refund', 'vat'}
    remote_set = set()
    if remote_zips and parcel_zipcodes:
        for wb_id, zc in parcel_zipcodes.items():
            if str(zc).strip() in remote_zips:
                remote_set.add(wb_id)

    merged = {}
    for p in parcels.values():
        wb_id = p['waybill']
        has_surcharge = any(ft in p['fees'] for ft in surcharge_triggers)
        has_remote = wb_id in remote_set
        if not has_surcharge and not has_remote:
            continue

        if wb_id not in merged:
            merged[wb_id] = {
                'date': p['date'], 'waybill': wb_id,
                'cust_ref': p.get('cust_ref', ''), 'transfer_no': p.get('transfer_no', ''),
                'category': p.get('category', ''), 'pieces': p.get('pieces'),
                'charge_weight': p.get('charge_weight'),
                'dimensions': p.get('dimensions', ''), 'dest': p.get('dest', ''),
                'remark_parts': [], 'fees': {},
            }

        rec = merged[wb_id]
        rec['fees'].update(p.get('fees', {}))

        if 'shelf_fee' in p['fees'] and '上架费' not in rec['remark_parts']:
            rec['remark_parts'].append('上架费')
        if 'tail_return_fee' in p['fees'] and '退件操作费' not in rec['remark_parts']:
            rec['remark_parts'].append('退件操作费')
        if 'tail_return_refund' in p['fees'] and '签收件，拒收返程费补退' not in rec['remark_parts']:
            rec['remark_parts'].append('签收件，拒收返程费补退')
        if 'vat' in p['fees'] and '增值税' not in rec['remark_parts']:
            rec['remark_parts'].append('增值税')
        if has_remote and '偏远费' not in rec['remark_parts']:
            rec['remark_parts'].append('偏远费')

    sur_list = sorted(merged.values(), key=lambda x: x.get('date') or datetime.min)

    for i, item in enumerate(sur_list):
        row = 2 + i
        ws.cell(row, 1, item['date'])
        ws.cell(row, 2, item['waybill'])
        ws.cell(row, 3, item['cust_ref'])
        ws.cell(row, 4, item['transfer_no'])
        ws.cell(row, 5, item['category'])
        ws.cell(row, 6, item['pieces'])
        ws.cell(row, 7, item['charge_weight'])
        ws.cell(row, 8, item['dimensions'])
        ws.cell(row, 10, item['dest'])

        fees = item.get('fees', {})

        if shelf_col and 'shelf_fee' in fees:
            ws.cell(row, shelf_col, f'=ROUND(1.5*{exchange_rate},2)')

        if return_col and 'tail_return_fee' in fees:
            dest = str(item['dest'] or '')
            ws.cell(row, return_col,
                    _build_return_fee_formula(dest, row, exchange_rate, pricing_version))
        if return_col and 'tail_return_refund' in fees:
            dest = str(item['dest'] or '')
            ws.cell(row, return_col, -abs(_calc_return_fee_rmb(dest, exchange_rate, pricing_version)))

        if vat_col and 'vat' in fees:
            vat_eur = fees['vat'].get('amount_eur', 0) or 0
            if 'tail_return_refund' in fees:
                ws.cell(row, vat_col, -abs(round(vat_eur * exchange_rate, 2)))
            else:
                ws.cell(row, vat_col, f'=ROUNDUP({vat_eur}*{exchange_rate},2)')

        if remote_col and remote_zips and parcel_zipcodes:
            zc = parcel_zipcodes.get(item['waybill'])
            if zc and str(zc).strip() in remote_zips:
                zi = remote_zips[str(zc).strip()]
                ws.cell(row, remote_col, calc_remote_fee_rmb(zi['region'], item.get('charge_weight') or 1, exchange_rate))
                if zip_col:
                    ws.cell(row, zip_col, str(zc).strip())

        if remark_col and item['remark_parts']:
            rp = item['remark_parts']
            if '签收件，拒收返程费补退' in rp:
                ws.cell(row, remark_col, '签收件，拒收返程费补退')
            elif len(rp) == 1:
                display = {'上架费': '上架费', '退件操作费': '尾程退件操作费',
                           '增值税': '目的地增值税', '偏远费': '偏远费'}
                ws.cell(row, remark_col, display.get(rp[0], rp[0]))
            else:
                ws.cell(row, remark_col, '+'.join(rp))

    if subtotal_col and sur_list:
        formula_src = min(data_end, 4)
        ensure_formulas(ws, formula_src, list(range(2, 2 + len(sur_list))), [subtotal_col])

    return {'sheet': wb.sheetnames[idx], 'rows': len(sur_list)}


def fill_f_surcharge(wb, parcels, exchange_rate):
    """F货附加费: 与运费Sheet完全相同的运单集，每单填入固定数据
    附加费列由模板公式计算（通常 =ROUNDUP(2*汇率,2)）
    """
    idx = find_sheet_by_type(wb, 'F附加费')
    if idx is None:
        return {'sheet': None, 'rows': 0, 'skip': 'no F-surcharge sheet'}

    ws = wb.worksheets[idx]
    data_end = find_data_end(ws, 8)
    clear_data_rows(ws, 2, data_end)

    c5_type, c6_type = detect_c5c6_fields(ws)

    f_parcels = sorted(
        [p for p in parcels.values() if 'tail_freight' in p['fees']],
        key=lambda p: p.get('date') or datetime.min
    )

    for i, p in enumerate(f_parcels):
        row = 2 + i
        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        ws.cell(row, 5, p.get(c5_type, ''))
        ws.cell(row, 6, p.get(c6_type, ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))

    return {'sheet': wb.sheetnames[idx], 'rows': len(f_parcels)}


# ── Main conversion flow ──

def convert_customer(code, customer, remote_zips):
    """Convert one customer: parse input, fill all applicable category sheets."""
    input_path = None
    for f in os.listdir(INPUT_DIR):
        if f'-{code}-' in f and f.endswith('.xlsx'):
            input_path = INPUT_DIR / f
    template_path = None
    for f in os.listdir(TEMPLATE_DIR):
        if customer in f and f.endswith('.xlsx'):
            template_path = TEMPLATE_DIR / f

    if not input_path or not template_path:
        print(f"  SKIP {code}->{customer}: missing files")
        return None

    parcels = parse_input(str(input_path))
    parcel_zipcodes = {}
    try:
        wb_tmp = openpyxl.load_workbook(str(template_path), data_only=True)
        sur_idx = find_sheet_by_type(wb_tmp, '杂费')
        if sur_idx is not None:
            ws_tmp = wb_tmp.worksheets[sur_idx]
            zc_col = find_col_by_keywords(ws_tmp, '邮编')
            wb_col = 2
            if zc_col:
                for r in range(2, ws_tmp.max_row + 1):
                    wb_id = ws_tmp.cell(r, wb_col).value
                    zc = ws_tmp.cell(r, zc_col).value
                    if wb_id and zc:
                        parcel_zipcodes[str(wb_id)] = str(zc)
        wb_tmp.close()
    except Exception:
        pass

    out_dir = OUTPUT_BASE / customer
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f'20260330-{customer}-自动生成.xlsx'

    wb = openpyxl.load_workbook(str(template_path))

    results = {}
    results['COD回款'] = fill_cod(wb, parcels, EXCHANGE_RATE)
    results['运费'] = fill_freight(wb, parcels, EXCHANGE_RATE)
    results['增值税'] = fill_vat(wb, parcels, EXCHANGE_RATE)
    results['尾程杂费'] = fill_surcharge(wb, parcels, EXCHANGE_RATE, PRICING_VERSION,
                                         remote_zips, parcel_zipcodes)
    results['F附加费'] = fill_f_surcharge(wb, parcels, EXCHANGE_RATE)

    wb.save(str(out_file))
    wb.close()

    return {
        'input': input_path.name,
        'template': template_path.name,
        'output': out_file.name,
        'parcels': len(parcels),
        'categories': results,
    }


def main():
    print("Loading remote zipcodes...")
    remote_zips = load_remote_zipcodes(str(PRICING_FILE))
    print(f"  Loaded {len(remote_zips)} entries\n")

    all_results = {}

    for code, customer in CUSTOMER_MAP.items():
        print(f"{'='*60}")
        print(f"{code} -> {customer}")
        result = convert_customer(code, customer, remote_zips)
        if result:
            all_results[customer] = result
            for cat, info in result['categories'].items():
                status = f"{info['rows']} rows -> {info['sheet']}" if info['sheet'] else f"SKIP ({info.get('skip', 'N/A')})"
                print(f"  [{cat}] {status}")

    # Summary by category
    print(f"\n\n{'='*60}")
    print("SUMMARY BY CATEGORY")
    print(f"{'='*60}")

    categories = ['COD回款', '运费', '增值税', '尾程杂费', 'F附加费']
    for cat in categories:
        print(f"\n  {cat}:")
        for customer, result in all_results.items():
            info = result['categories'].get(cat, {})
            if info.get('rows', 0) > 0:
                print(f"    {customer}: {info['rows']} rows")
            elif info.get('sheet'):
                print(f"    {customer}: 0 rows (empty)")
            else:
                print(f"    {customer}: N/A")


if __name__ == '__main__':
    main()
