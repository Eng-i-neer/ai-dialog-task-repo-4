"""分析不同目的地国家的运费计费规则"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
ref = openpyxl.load_workbook(REF_FILE, data_only=True)
ws = ref[ref.sheetnames[1]]

# Group by dest country
by_country = {}
for r in range(2, ws.max_row + 1):
    dest = ws.cell(r, 7).value
    if not dest:
        continue
    head_w = ws.cell(r, 10).value or 0
    tail_qty = ws.cell(r, 12).value or 0
    head_cost = ws.cell(r, 14).value or 0
    tail_cost = ws.cell(r, 15).value or 0
    de_sur = ws.cell(r, 16).value or 0
    addon = ws.cell(r, 17).value or 0
    de_addon = ws.cell(r, 18).value  # col R: 德国增值加费(second)
    subtotal = ws.cell(r, 19).value or 0

    if dest not in by_country:
        by_country[dest] = []
    by_country[dest].append({
        'waybill': ws.cell(r, 4).value,
        'head_w': head_w,
        'tail_qty': tail_qty,
        'head_cost': head_cost,
        'tail_cost': tail_cost,
        'de_sur': de_sur,
        'addon': addon,
        'de_addon': de_addon,
        'subtotal': subtotal,
    })

for country, items in sorted(by_country.items()):
    print(f"\n{'='*60}")
    print(f"  目的地: {country}  ({len(items)} 条)")
    print(f"{'='*60}")

    # head unit price
    head_prices = set()
    for it in items:
        if it['head_w'] and it['head_w'] > 0 and it['head_cost'] > 0:
            up = round(it['head_cost'] / it['head_w'], 2)
            head_prices.add(up)

    print(f"  头程单价: {sorted(head_prices)}")

    # tail prices
    tail_by_qty = {}
    for it in items:
        q = it['tail_qty']
        if q not in tail_by_qty:
            tail_by_qty[q] = set()
        tail_by_qty[q].add(it['tail_cost'])
    print(f"  尾程运费: { {q: sorted(v) for q, v in sorted(tail_by_qty.items())} }")

    # de_sur values
    de_sur_vals = set(it['de_sur'] for it in items)
    print(f"  德国增值加费(col P): {sorted(de_sur_vals)}")

    # addon values
    addon_vals = {}
    for it in items:
        q = it['tail_qty']
        if q not in addon_vals:
            addon_vals[q] = set()
        addon_vals[q].add(it['addon'])
    print(f"  加费(col Q): { {q: sorted(v) for q, v in sorted(addon_vals.items())} }")

    # de_addon (col R)
    de_addon_vals = set(it['de_addon'] for it in items if it['de_addon'] is not None)
    print(f"  德国增值加费(col R): {sorted(de_addon_vals) if de_addon_vals else 'None'}")

    # Show a few examples
    print(f"  示例:")
    for it in items[:3]:
        print(f"    hw={it['head_w']}, tq={it['tail_qty']}, "
              f"hc={it['head_cost']}, tc={it['tail_cost']}, "
              f"ds={it['de_sur']}, ad={it['addon']}, da={it['de_addon']}, "
              f"sub={it['subtotal']}")

    # Verify subtotal formula
    ok = 0
    for it in items:
        calc = it['head_cost'] + it['tail_cost'] + it['de_sur'] + it['addon']
        if it['de_addon']:
            calc += it['de_addon']
        if abs(calc - it['subtotal']) < 0.1:
            ok += 1
    print(f"  小计公式验证: {ok}/{len(items)} 匹配")
