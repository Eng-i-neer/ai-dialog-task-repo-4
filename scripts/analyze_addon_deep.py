"""检查 addon=30 vs 11.91 到底跟什么因素关联"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# 从输入建立保险费lookup
ins_data = {}
for sname in inp.sheetnames:
    ws_tmp = inp[sname]
    for r in range(10, ws_tmp.max_row + 1):
        wb = ws_tmp.cell(r, 3).value
        item_type = ws_tmp.cell(r, 12).value
        if wb and item_type:
            if wb not in ins_data:
                ins_data[wb] = set()
            ins_data[wb].add(item_type)

# Check: 直发/转发 vs addon
ws_freight = ref[ref.sheetnames[1]]
ship_addon = {}
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    ship_type = ws_freight.cell(r, 2).value
    tail_qty = ws_freight.cell(r, 12).value
    dest = ws_freight.cell(r, 7).value
    if not wb or not addon:
        continue
    key = (ship_type, tail_qty, round(addon, 0))
    ship_addon[key] = ship_addon.get(key, 0) + 1

print("=== (ship_type, tail_qty, addon) distribution ===")
for k, v in sorted(ship_addon.items()):
    print(f"  type={k[0]}, tail_qty={k[1]}, addon={k[2]}: {v} records")

# Check: does addon=30 match with 直发?
print("\n=== Check if addon=30 correlates with 直发 ===")
direct_count_30 = 0
direct_count_12 = 0
transfer_count_30 = 0
transfer_count_12 = 0
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    ship_type = ws_freight.cell(r, 2).value
    if not wb or not addon or not ship_type:
        continue
    is_direct = '直' in str(ship_type) or 'ֱ' in str(ship_type)
    if abs(addon - 30) < 0.1:
        if is_direct:
            direct_count_30 += 1
        else:
            transfer_count_30 += 1
    elif abs(addon - 11.91) < 0.1:
        if is_direct:
            direct_count_12 += 1
        else:
            transfer_count_12 += 1

print(f"  直发 addon=30: {direct_count_30}")
print(f"  直发 addon=11.91: {direct_count_12}")
print(f"  转发 addon=30: {transfer_count_30}")
print(f"  转发 addon=11.91: {transfer_count_12}")
