import openpyxl

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
OUTPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)
out = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)

# For output freight row 2, waybill = DE12602241410017
# Found in input: 保险费(10), DE清关(3.05), 尾程运费(5.04), 目的地增值税(1.2), 头程运费(12.8=8*1.6)
# Output: 头程运费=126.95, 尾程运费=63.48, 德国增加加费=55.54, 加费=11.91, 小计=257.88

# These are clearly DIFFERENT numbers - input is in EUR, output is in CNY
# But the output values don't match EUR*rate either
# Output freight values look like FIXED rates based on weight tiers

# Let's check if 头程运费(output col N) correlates with 头程计费重(output col J)
ws_out = out[out.sheetnames[1]]
print("=== Output Freight: 头程运费 vs 头程计费重 ===")
weight_price = {}
for r in range(2, min(ws_out.max_row + 1, 510)):
    w = ws_out.cell(r, 10).value  # 头程计费重量
    p = ws_out.cell(r, 14).value  # 头程运费
    tail = ws_out.cell(r, 15).value  # 尾程运费
    surcharge = ws_out.cell(r, 16).value  # 德国增加加费
    addon = ws_out.cell(r, 17).value  # 加费
    tail_charge = ws_out.cell(r, 12).value  # 尾程计费量
    svc = ws_out.cell(r, 13).value  # 服务类型
    if w and p:
        key = (w, tail_charge, svc)
        if key not in weight_price:
            weight_price[key] = {'head': p, 'tail': tail, 'sur': surcharge, 'addon': addon, 'count': 0}
        weight_price[key]['count'] += 1

print(f"Unique (weight, tail_charge, svc) combos: {len(weight_price)}")
for k, v in sorted(weight_price.items()):
    print(f"  weight={k[0]}, tail_qty={k[1]}, svc={k[2]}: "
          f"head={v['head']}, tail={v['tail']}, sur={v['sur']}, addon={v['addon']} "
          f"({v['count']} records)")

# Check output sheet 3: 尾程加费
ws_sur = out[out.sheetnames[2]]
print(f"\n=== OUTPUT 尾程加费 HEADER ===")
for c in range(1, 21):
    v = ws_sur.cell(1, c).value
    if v:
        print(f"  Col {c}: {v}")

print(f"\n=== OUTPUT 尾程加费 Row 2 ===")
for c in range(1, 21):
    v = ws_sur.cell(2, c).value
    if v is not None:
        print(f"  Col {c}: {repr(v)}")

# Check output sheet 4: 在库存储
ws_store = out[out.sheetnames[3]]
print(f"\n=== OUTPUT 在库存储 HEADER ===")
for c in range(1, 19):
    v = ws_store.cell(1, c).value
    if v:
        print(f"  Col {c}: {v}")
print(f"  Rows: {ws_store.max_row}")
for r in range(2, ws_store.max_row + 1):
    vals = [ws_store.cell(r, c).value for c in range(1, 19)]
    if any(v is not None for v in vals):
        print(f"  Row {r}: {[v for v in vals if v is not None]}")

# Check output sheet 5: 保险
ws_ins = out[out.sheetnames[4]]
print(f"\n=== OUTPUT 保险 HEADER ===")
for c in range(1, 17):
    v = ws_ins.cell(1, c).value
    if v:
        print(f"  Col {c}: {v}")
print(f"  Rows: {ws_ins.max_row}")
for r in range(2, ws_ins.max_row + 1):
    vals = [ws_ins.cell(r, c).value for c in range(1, 17)]
    if any(v is not None for v in vals):
        print(f"  Row {r}: {[v for v in vals if v is not None]}")

# Summary totals from output COD sheet bottom
ws_cod = out[out.sheetnames[0]]
print(f"\n=== OUTPUT COD Sheet bottom summary ===")
for r in range(190, ws_cod.max_row + 1):
    vals = []
    for c in range(1, 18):
        v = ws_cod.cell(r, c).value
        if v is not None:
            vals.append(f"Col{c}={repr(v)}")
    if vals:
        print(f"  Row {r}: {vals}")
