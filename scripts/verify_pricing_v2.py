"""用报价文件验证输出文件中的费用计算规则 - v2"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
OUTPUT  = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT   = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

output_wb  = openpyxl.load_workbook(OUTPUT, data_only=True)
input_wb   = openpyxl.load_workbook(INPUT, data_only=True)

print(f"输出文件 sheets: {output_wb.sheetnames}")
print(f"输入文件 sheets: {input_wb.sheetnames}")

# 找到正确的sheet名
cod_sheet = None
freight_sheet = None
addon_sheet = None
for sn in output_wb.sheetnames:
    sn_lower = sn.lower()
    if 'cod' in sn_lower:
        cod_sheet = sn
    elif '运费' in sn or 'freight' in sn_lower:
        freight_sheet = sn
    elif '加费' in sn or '尾程' in sn:
        addon_sheet = sn

print(f"\nCOD sheet: {cod_sheet}")
print(f"运费 sheet: {freight_sheet}")
print(f"加费 sheet: {addon_sheet}")

# --- COD 手续费验证 ---
if cod_sheet:
    ws = output_wb[cod_sheet]
    print(f"\n{'='*70}")
    print(f"  验证 COD 手续费 (Sheet: {cod_sheet})")
    print(f"{'='*70}")
    
    # Headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[c] = str(v).strip()
            print(f"  C{c}: {v}")
    
    print(f"\n  验证每条数据的COD手续费:")
    for r in range(2, ws.max_row + 1):
        waybill = ws.cell(r, 4).value
        if not waybill:
            continue
        dest = ws.cell(r, 7).value or ''
        cod_amount = ws.cell(r, 9).value  # I: 代收金额
        cod_fee_eur = ws.cell(r, 11).value  # K: COD手续费(EUR)
        rate = ws.cell(r, 10).value  # J: 汇率
        cod_fee_cny = ws.cell(r, 12).value  # L
        total_cny = ws.cell(r, 13).value  # M
        
        if cod_amount is not None and cod_fee_eur is not None:
            pct3 = cod_amount * 0.03
            
            # 德国规则: max(3%, 7.0)
            rule_de = max(pct3, 7.0)
            match_de = abs(cod_fee_eur - rule_de) < 0.01
            
            # 检查 CNY = EUR * rate
            calc_cny = cod_fee_eur * rate if rate else 0
            cny_match = abs((cod_fee_cny or 0) - calc_cny) < 0.1
            
            # total = cod_amount * rate - cod_fee_cny
            calc_total = cod_amount * rate - (cod_fee_cny or 0) if rate else 0
            total_match = abs((total_cny or 0) - calc_total) < 0.1
            
            status = "OK" if match_de else "MISMATCH"
            print(f"  R{r}: {waybill} | dest={dest} | 代收={cod_amount:.2f} | "
                  f"3%={pct3:.2f} | max(3%,7)={rule_de:.2f} | 实际={cod_fee_eur:.2f} | "
                  f"{status} | CNY={cod_fee_cny} calc={calc_cny:.2f} {'OK' if cny_match else 'X'} | "
                  f"合计={total_cny} calc={calc_total:.2f} {'OK' if total_match else 'X'}")

# --- 运费验证 ---
if freight_sheet:
    ws = output_wb[freight_sheet]
    print(f"\n{'='*70}")
    print(f"  验证运费 (Sheet: {freight_sheet})")
    print(f"{'='*70}")
    
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            print(f"  C{c}: {v}")
    
    # 从输入文件获取重量信息
    # 找DE头程
    weight_data = {}
    for sname in input_wb.sheetnames:
        if '头程' in sname:
            ws_in = input_wb[sname]
            # 找运单号和重量列
            for r in range(2, ws_in.max_row + 1):
                for c in range(1, ws_in.max_column + 1):
                    v = ws_in.cell(r, c).value
                    if v and str(v).startswith('DE') and len(str(v)) > 10:
                        waybill = str(v).strip()
                        # 尝试找同行的重量
                        row_vals = {}
                        for cc in range(1, ws_in.max_column + 1):
                            vv = ws_in.cell(r, cc).value
                            row_vals[cc] = vv
                        weight_data[waybill] = row_vals
                        break
    
    print(f"\n  从输入文件找到 {len(weight_data)} 条头程重量数据")
    
    # 验证运费
    print(f"\n  运费验证 (前20条):")
    for r in range(2, min(ws.max_row + 1, 22)):
        waybill = ws.cell(r, 4).value
        if not waybill:
            continue
        dest = ws.cell(r, 7).value or ''
        category = ws.cell(r, 8).value or ''
        freight_eur = ws.cell(r, 9).value
        rate = ws.cell(r, 10).value
        freight_cny = ws.cell(r, 11).value
        ship_type = ws.cell(r, 2).value or ''
        
        # 德国普货 报价: 头程9.7, 首2KG尾程8.3, 续1KG 1.5
        # 运费 = 头程 + 尾程(首2KG + 续重)
        # 但输出中运费列只有一个值... 需要理解这个值是什么
        
        calc_cny = freight_eur * rate if freight_eur and rate else 0
        cny_match = abs((freight_cny or 0) - calc_cny) < 0.1 if freight_cny else False
        
        print(f"  R{r}: {waybill} | {ship_type} | dest={dest} | 品名={category} | "
              f"运费EUR={freight_eur} | 汇率={rate} | "
              f"运费CNY={freight_cny} calc={calc_cny:.2f} {'OK' if cny_match else 'X'}")

# --- 尾程加费验证 ---
if addon_sheet:
    ws = output_wb[addon_sheet]
    print(f"\n{'='*70}")
    print(f"  验证尾程加费 (Sheet: {addon_sheet})")
    print(f"{'='*70}")
    
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            print(f"  C{c}: {v}")
    
    print(f"\n  全部尾程加费数据:")
    for r in range(2, ws.max_row + 1):
        vals = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                vals[c] = v
        if vals:
            waybill = vals.get(4, '')
            fee_eur = vals.get(9, 0)
            rate = vals.get(10, 0)
            fee_cny = vals.get(11, 0)
            fee_name = vals.get(8, '')
            
            calc_cny = fee_eur * rate if fee_eur and rate else 0
            cny_match = abs((fee_cny or 0) - calc_cny) < 0.1 if fee_cny is not None else False
            
            print(f"  R{r}: {waybill} | 费用={fee_name} | EUR={fee_eur} | "
                  f"rate={rate} | CNY={fee_cny} calc={calc_cny:.2f} {'OK' if cny_match else 'X'} | "
                  f"all={vals}")

output_wb.close()
input_wb.close()
