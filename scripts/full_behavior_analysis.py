"""
完整数据行为分析：
1. 输入文件：每个运单出现在哪些Sheet，各Sheet的费用金额
2. 输出文件：每条数据的字段来源追溯
3. 输入→输出的完整转换链路
"""
import openpyxl
from collections import defaultdict

INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
OUTPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

inp = openpyxl.load_workbook(INPUT, data_only=True)
out = openpyxl.load_workbook(OUTPUT, data_only=True)

# ═══════════════════════════════════════════════════
# Part 1: 输入文件 — 建立运单号全量索引
# ═══════════════════════════════════════════════════

# waybill -> { sheet_name: {amount, charge_weight, ...} }
waybill_index = defaultdict(dict)
# 每个sheet的中文名
sheet_labels = {}

for sname in inp.sheetnames:
    ws = inp[sname]
    if sname == '\u603b\u8868':  # 总表
        continue

    # 获取表头 row 9 col L (项目名)
    header_l = ws.cell(9, 12).value or sname
    sheet_labels[sname] = header_l

    for r in range(10, ws.max_row + 1):
        waybill = ws.cell(r, 3).value
        if not waybill or not isinstance(waybill, str):
            continue

        waybill_index[waybill][sname] = {
            'date': ws.cell(r, 2).value,
            'transfer_no': ws.cell(r, 4).value,
            'route': ws.cell(r, 5).value,
            'pieces': ws.cell(r, 6).value,
            'actual_weight': ws.cell(r, 7).value,
            'charge_weight': ws.cell(r, 8).value,
            'dimensions': ws.cell(r, 9).value,
            'dest': ws.cell(r, 11).value,
            'fee_type': ws.cell(r, 12).value,
            'amount_eur': ws.cell(r, 13).value,
            'formula': ws.cell(r, 14).value,
            'cust_ref': ws.cell(r, 16).value,
            'category': ws.cell(r, 17).value,
        }

print(f"Total unique waybills in input: {len(waybill_index)}")
print(f"Sheets (excl 总表): {len(sheet_labels)}")

# 统计每个运单出现在几个Sheet中
appearance_count = defaultdict(int)
for wb, sheets in waybill_index.items():
    appearance_count[len(sheets)] += 1

print(f"\nWaybill appearance distribution:")
for cnt in sorted(appearance_count):
    print(f"  Appears in {cnt} sheets: {appearance_count[cnt]} waybills")

# 统计每种Sheet组合
combo_count = defaultdict(int)
for wb, sheets in waybill_index.items():
    key = tuple(sorted(sheets.keys()))
    combo_count[key] += 1

print(f"\nTop sheet combinations:")
for combo, cnt in sorted(combo_count.items(), key=lambda x: -x[1])[:15]:
    short = [s[:6] for s in combo]
    print(f"  {cnt:>4} waybills: {short}")

# ═══════════════════════════════════════════════════
# Part 2: 输出文件每个Sheet — 数据来源追溯
# ═══════════════════════════════════════════════════

print(f"\n{'='*70}")
print("OUTPUT FILE ANALYSIS")
print(f"{'='*70}")

for si, sname in enumerate(out.sheetnames):
    ws = out[sname]
    print(f"\n--- Output Sheet {si+1}: {sname} ---")
    print(f"    Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Collect all waybills in this output sheet
    waybill_col = None
    for c in range(1, ws.max_column + 1):
        for r in range(2, min(10, ws.max_row + 1)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and len(v) > 10 and v[:2] in ('DE', 'IT', 'HR', 'GR', 'PL'):
                waybill_col = c
                break
        if waybill_col:
            break

    if not waybill_col:
        print(f"    No waybill column found, skipping")
        continue

    print(f"    Waybill column: {waybill_col}")

    out_waybills = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, waybill_col).value
        if isinstance(v, str) and len(v) > 10:
            out_waybills.append(v)

    print(f"    Total data rows: {len(out_waybills)}")

    # For each output waybill, which input sheets does it appear in?
    source_sheets = defaultdict(int)
    not_found = []
    for wb in out_waybills:
        if wb in waybill_index:
            for s in waybill_index[wb]:
                source_sheets[s] += 1
        else:
            not_found.append(wb)

    print(f"    Source sheets (input):")
    for s, cnt in sorted(source_sheets.items(), key=lambda x: -x[1]):
        pct = cnt / len(out_waybills) * 100
        print(f"      {s}: {cnt}/{len(out_waybills)} ({pct:.0f}%)")
    if not_found:
        print(f"    Not found in input: {len(not_found)} -> {not_found[:3]}")

# ═══════════════════════════════════════════════════
# Part 3: 逐字段来源追溯 (Output Sheet 1: COD)
# ═══════════════════════════════════════════════════

print(f"\n{'='*70}")
print("FIELD-BY-FIELD SOURCE TRACING: COD Sheet")
print(f"{'='*70}")

ws_out_cod = out[out.sheetnames[0]]
ws_in_cod = inp['COD']

# Build input COD lookup
in_cod = {}
for r in range(10, ws_in_cod.max_row + 1):
    wb = ws_in_cod.cell(r, 3).value
    if wb:
        in_cod[wb] = r

# Check each output field
print("\nFor waybills that ARE in input COD sheet:")
sample_wb = None
for r in range(2, ws_out_cod.max_row + 1):
    wb = ws_out_cod.cell(r, 4).value
    if wb and wb in in_cod:
        sample_wb = wb
        ir = in_cod[wb]
        print(f"\n  Example: {wb} (output row {r}, input COD row {ir})")
        out_vals = {c: ws_out_cod.cell(r, c).value for c in range(1, 15)}
        in_vals = {c: ws_in_cod.cell(ir, c).value for c in range(1, 19)}
        
        mappings = [
            ('A: account_period', out_vals[1], 'FIXED constant 46111'),
            ('B: ship_type', out_vals[2], 'DERIVED from cust_ref pattern'),
            ('C: ship_date', out_vals[3], f'input COD col B = {in_vals[2]}'),
            ('D: waybill', out_vals[4], f'input COD col C = {in_vals[3]}'),
            ('E: order_no', out_vals[5], f'input COD col P = {in_vals[16]}'),
            ('F: transfer_no', out_vals[6], f'input COD col D = {in_vals[4]}'),
            ('G: dest', out_vals[7], f'input COD col K = {in_vals[11]}'),
            ('H: category', out_vals[8], f'input COD col Q = {in_vals[17]}'),
            ('I: cod_eur', out_vals[9], f'input COD col M = {in_vals[13]}'),
            ('J: rate', out_vals[10], 'EXTERNAL exchange rate'),
            ('K: cny', out_vals[11], 'FORMULA =ROUND(I*J,2)'),
            ('L: cod_fee', out_vals[12], 'FORMULA (complex, varies per row)'),
            ('M: subtotal', out_vals[13], 'FORMULA =K-L'),
        ]
        for label, oval, source in mappings:
            print(f"    {label}: output={repr(oval)[:40]} <- {source}")
        break

print("\nFor waybills NOT in input COD sheet (historical):")
for r in range(2, ws_out_cod.max_row + 1):
    wb = ws_out_cod.cell(r, 4).value
    if wb and wb not in in_cod:
        print(f"  {wb}: note={ws_out_cod.cell(r, 14).value}")
        if wb in waybill_index:
            print(f"    Found in input sheets: {list(waybill_index[wb].keys())}")
        else:
            print(f"    NOT FOUND in input at all")

# ═══════════════════════════════════════════════════
# Part 4: 逐字段来源追溯 (Output Sheet 2: 运费)
# ═══════════════════════════════════════════════════

print(f"\n{'='*70}")
print("FIELD-BY-FIELD SOURCE TRACING: Freight Sheet")
print(f"{'='*70}")

ws_out_f = out[out.sheetnames[1]]
# Read formulas (not data_only)
out_formulas = openpyxl.load_workbook(OUTPUT)
ws_out_f_formulas = out_formulas[out_formulas.sheetnames[1]]

# Get one example row with formulas
print("\nFormula patterns from template row 2:")
for c in range(1, 21):
    v = ws_out_f_formulas.cell(2, c).value
    h = ws_out_f_formulas.cell(1, c).value
    if v is not None:
        print(f"  Col {c} ({h}): {repr(v)[:80]}")

# Track where each field comes from
print("\nField source mapping:")
wb_example = ws_out_f.cell(2, 4).value
if wb_example and wb_example in waybill_index:
    sheets = waybill_index[wb_example]
    print(f"  Example waybill: {wb_example}")
    print(f"  Appears in input sheets: {list(sheets.keys())}")
    for s, data in sheets.items():
        print(f"    {s}: amount={data['amount_eur']}, charge_w={data['charge_weight']}")

# ═══════════════════════════════════════════════════
# Part 5: 逐字段来源追溯 (Output Sheet 3: 尾程加费)
# ═══════════════════════════════════════════════════

print(f"\n{'='*70}")
print("FIELD-BY-FIELD SOURCE TRACING: Surcharge Sheet")
print(f"{'='*70}")

ws_out_s = out[out.sheetnames[2]]
ws_out_s_formulas = out_formulas[out_formulas.sheetnames[2]]

print("\nFormula patterns from template rows 2-5:")
for r in [2, 3, 4, 5]:
    has_formula = False
    for c in range(1, 21):
        v = ws_out_s_formulas.cell(r, c).value
        if v and isinstance(v, str) and v.startswith('='):
            h = ws_out_s_formulas.cell(1, c).value
            print(f"  Row {r} Col {c} ({h}): {v[:80]}")
            has_formula = True
    if has_formula:
        break

# Check which fee types appear in surcharge sheet
print("\nSurcharge waybills - input sheet sources:")
sur_sources = defaultdict(int)
for r in range(2, ws_out_s.max_row + 1):
    wb = ws_out_s.cell(r, 2).value
    if wb and wb in waybill_index:
        for s in waybill_index[wb]:
            sur_sources[s] += 1

for s, cnt in sorted(sur_sources.items(), key=lambda x: -x[1]):
    print(f"  {s}: {cnt}")

# Check col R (备注) patterns
print("\nSurcharge 备注 patterns:")
notes = defaultdict(int)
for r in range(2, ws_out_s.max_row + 1):
    n = ws_out_s.cell(r, 18).value
    if n:
        notes[n] += 1
for n, cnt in sorted(notes.items(), key=lambda x: -x[1]):
    print(f"  '{n}': {cnt}")

# Check col K (科目) patterns  
print("\nSurcharge 科目 patterns:")
items = defaultdict(int)
for r in range(2, ws_out_s.max_row + 1):
    n = ws_out_s.cell(r, 11).value
    if n:
        items[n] += 1
for n, cnt in sorted(items.items(), key=lambda x: -x[1]):
    print(f"  '{n}': {cnt}")
