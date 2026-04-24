"""用报价文件验证输出文件中的费用计算规则"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
INPUT   = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
OUTPUT  = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

pricing_wb = openpyxl.load_workbook(PRICING, data_only=True)
input_wb   = openpyxl.load_workbook(INPUT, data_only=True)
output_wb  = openpyxl.load_workbook(OUTPUT, data_only=True)

# 1. 从报价表提取每个国家的报价规则
print("="*70)
print("  报价表 - 欧洲COD 费率汇总")
print("="*70)

ws_price = pricing_wb['欧洲COD']
countries = {}
current_country = None
for r in range(6, 48):
    c1 = ws_price.cell(r, 1).value
    c3 = ws_price.cell(r, 3).value
    c4 = ws_price.cell(r, 4).value
    c5 = ws_price.cell(r, 5).value
    c6 = ws_price.cell(r, 6).value
    c7 = ws_price.cell(r, 7).value
    c8 = ws_price.cell(r, 8).value
    
    if c1 and str(c1).strip():
        current_country = str(c1).strip()
        countries[current_country] = {
            'carrier': str(ws_price.cell(r, 2).value or '').strip(),
            'cod_fee_rule': str(c8 or '').strip(),
            'return_fee': str(c7 or '').strip(),
            'rates': {}
        }
    
    if current_country and c3:
        cargo_type = str(c3).strip()
        countries[current_country]['rates'][cargo_type] = {
            'head_freight': c4,
            'first_2kg': c5,
            'per_1kg': c6
        }

for cname, info in countries.items():
    print(f"\n  {cname}")
    print(f"    承运商: {info['carrier']}")
    print(f"    COD手续费: {info['cod_fee_rule']}")
    print(f"    拒收返程费: {info['return_fee']}")
    for cargo, rates in info['rates'].items():
        print(f"    {cargo}: 头程={rates['head_freight']}, 首2KG={rates['first_2kg']}, 续1KG={rates['per_1kg']}")

# 2. 从输出文件读取COD表，验证COD手续费计算
print("\n" + "="*70)
print("  验证COD手续费计算")
print("="*70)

ws_cod = output_wb['COD回款']
print(f"\n  COD回款 Sheet, max_row={ws_cod.max_row}")
print(f"  Headers: ", end='')
for c in range(1, ws_cod.max_column + 1):
    v = ws_cod.cell(1, c).value
    if v:
        print(f"C{c}={v} ", end='')
print()

# 读取COD数据行并分析手续费
cod_data = []
for r in range(2, ws_cod.max_row + 1):
    waybill = ws_cod.cell(r, 4).value
    if not waybill:
        continue
    cod_amount = ws_cod.cell(r, 9).value  # I: 代收金额
    rate = ws_cod.cell(r, 10).value       # J: 汇率
    cod_fee_eur = ws_cod.cell(r, 11).value  # K: COD手续费(EUR)
    cod_fee_cny = ws_cod.cell(r, 12).value  # L: COD手续费(CNY)
    total_cny = ws_cod.cell(r, 13).value    # M: 总计(CNY)
    dest = ws_cod.cell(r, 7).value          # G: 目的地
    
    # 计算 3% * cod_amount
    if cod_amount and cod_fee_eur:
        calc_3pct = round(cod_amount * 0.03, 2)
        diff = round(cod_fee_eur - calc_3pct, 4) if cod_fee_eur else None
        cod_data.append({
            'row': r,
            'waybill': waybill,
            'dest': dest,
            'cod_amount': cod_amount,
            'cod_fee_eur': cod_fee_eur,
            'calc_3pct': calc_3pct,
            'diff': diff,
            'rate': rate,
            'cod_fee_cny': cod_fee_cny,
            'total_cny': total_cny,
        })

print(f"\n  共 {len(cod_data)} 条COD数据")

# 德国 COD手续费规则: 3%，最低7.0EUR；不签收也按7.0欧/票收
# 需要看不同目的国
dest_groups = {}
for d in cod_data:
    dest = str(d['dest'] or '').strip()
    if dest not in dest_groups:
        dest_groups[dest] = []
    dest_groups[dest].append(d)

for dest, items in sorted(dest_groups.items()):
    print(f"\n  目的地: {dest} ({len(items)} 条)")
    
    # 找到报价中对应国家的规则
    matched_country = None
    dest_upper = dest.upper()
    for cname in countries:
        if dest_upper in cname.upper() or dest.lower() in cname.lower():
            matched_country = cname
            break
        # 尝试国家代码匹配
        import re
        code_match = re.search(r'\((\w+)\)', cname)
        if code_match and code_match.group(1).upper() == dest_upper:
            matched_country = cname
            break
    
    if matched_country:
        print(f"    匹配报价: {matched_country}")
        print(f"    COD手续费规则: {countries[matched_country]['cod_fee_rule']}")
    
    # 分析实际数据
    for item in items[:5]:
        cod_amt = item['cod_amount']
        fee = item['cod_fee_eur']
        pct3 = item['calc_3pct']
        
        # 尝试各种规则
        # 德国: 3% 最低7.0
        min7 = max(pct3, 7.0)
        # 意大利: 3% 最低2.0
        min2 = max(pct3, 2.0)
        # 波兰: 3% 最低1.5
        min15 = max(pct3, 1.5)
        # 奥地利: 3% 最低5.0
        min5 = max(pct3, 5.0)
        
        print(f"    单号={item['waybill']}, 代收={cod_amt}, 实际手续费={fee}, "
              f"3%={pct3}, max(3%,7)={min7}, max(3%,2)={min2}, max(3%,1.5)={min15}")

# 3. 验证运费表的头程运费计算
print("\n" + "="*70)
print("  验证运费表的头程运费")
print("="*70)

ws_freight = output_wb['运费']
print(f"\n  运费 Sheet, max_row={ws_freight.max_row}")
print(f"  Headers: ", end='')
for c in range(1, ws_freight.max_column + 1):
    v = ws_freight.cell(1, c).value
    if v:
        print(f"C{c}={v} ", end='')
print()

# 读前10条运费数据做验证
print("\n  前15条运费数据:")
for r in range(2, min(ws_freight.max_row + 1, 17)):
    waybill = ws_freight.cell(r, 4).value
    if not waybill:
        continue
    dest = ws_freight.cell(r, 7).value
    category = ws_freight.cell(r, 8).value
    freight_eur = ws_freight.cell(r, 9).value
    rate = ws_freight.cell(r, 10).value
    freight_cny = ws_freight.cell(r, 11).value
    
    print(f"  R{r}: 单号={waybill}, 目的地={dest}, 品名={category}, "
          f"运费EUR={freight_eur}, 汇率={rate}, 运费CNY={freight_cny}")

# 从输入文件查找这些运单对应的重量信息
print("\n  对比输入文件中的重量数据:")
input_sheets = input_wb.sheetnames
# 先搜DE头程 sheet
for sname in input_sheets:
    if '头程' in sname and 'DE' in sname.upper():
        ws_in = input_wb[sname]
        print(f"\n  输入Sheet: {sname}")
        # 打印header
        headers = []
        for c in range(1, ws_in.max_column + 1):
            v = ws_in.cell(1, c).value
            if v:
                headers.append(f"C{c}={v}")
        print(f"  Headers: {headers}")
        # 前5条数据
        for r in range(2, min(ws_in.max_row + 1, 7)):
            vals = []
            for c in range(1, ws_in.max_column + 1):
                v = ws_in.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}={v}")
            print(f"  R{r}: {vals}")
        break

# 4. 检验"普特敏货"分类
print("\n" + "="*70)
print("  验证品名/货物类型分类")
print("="*70)
categories = set()
for r in range(2, ws_freight.max_row + 1):
    cat = ws_freight.cell(r, 8).value
    if cat:
        categories.add(str(cat).strip())
print(f"  运费表中出现的品名类别: {categories}")

categories_cod = set()
for r in range(2, ws_cod.max_row + 1):
    cat = ws_cod.cell(r, 8).value
    if cat:
        categories_cod.add(str(cat).strip())
print(f"  COD表中出现的品名类别: {categories_cod}")

# 5. 验证尾程加费
print("\n" + "="*70)
print("  验证尾程加费")
print("="*70)

ws_addon = output_wb['尾程加费']
print(f"\n  尾程加费 Sheet, max_row={ws_addon.max_row}")
print(f"  Headers: ", end='')
for c in range(1, ws_addon.max_column + 1):
    v = ws_addon.cell(1, c).value
    if v:
        print(f"C{c}={v} ", end='')
print()

# 读取全部尾程加费数据
print("\n  尾程加费数据:")
for r in range(2, ws_addon.max_row + 1):
    vals = []
    for c in range(1, ws_addon.max_column + 1):
        v = ws_addon.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"  R{r}: {vals}")

# 从输入文件找尾程加费的来源
print("\n  输入文件中的尾程加费相关sheets:")
for sname in input_sheets:
    if '尾程' in sname or '加费' in sname or '退件' in sname or '上架' in sname:
        ws_in = input_wb[sname]
        print(f"\n  Input Sheet: {sname}, rows={ws_in.max_row}")
        for r in range(1, min(ws_in.max_row + 1, 6)):
            vals = []
            for c in range(1, ws_in.max_column + 1):
                v = ws_in.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}={v}")
            if vals:
                print(f"    R{r}: {vals}")

pricing_wb.close()
input_wb.close()
output_wb.close()
