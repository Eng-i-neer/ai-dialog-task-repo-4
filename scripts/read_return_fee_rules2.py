"""读取报价文件 R30+ 的完整数据"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

PRICING = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'

wb = openpyxl.load_workbook(PRICING, data_only=True)
ws = wb['欧洲COD']

print("欧洲COD Sheet R30-R85:")
for r in range(30, min(ws.max_row + 1, 86)):
    row_vals = []
    for c in range(1, min(ws.max_column + 1, 15)):
        v = ws.cell(r, c).value
        if v is not None:
            row_vals.append(f"C{c}={v}")
    if row_vals:
        print(f"  R{r}: {' | '.join(row_vals)}")

print("\n\n汇总: 各国尾程派送费与退件费规则")
print("-"*80)
# 重新遍历提取结构化数据
countries = []
current = None
for r in range(6, 70):
    c1 = ws.cell(r, 1).value
    c3 = ws.cell(r, 3).value
    c4 = ws.cell(r, 4).value
    c5 = ws.cell(r, 5).value
    c6 = ws.cell(r, 6).value
    c7 = ws.cell(r, 7).value
    
    if c1 and str(c1).strip():
        current = str(c1).strip()
    
    if c3 and str(c3).strip() == '普货' and current:
        countries.append({
            'country': current,
            'head_freight': c4,
            'delivery_first2kg': c5,
            'delivery_extra_1kg': c6,
            'return_rule': str(c7 or '').strip(),
        })

print(f"{'国家':<20} {'首2KG':>8} {'续1KG':>8} {'退件费规则':<30}")
print("-"*80)
for c in countries:
    print(f"{c['country']:<20} {c['delivery_first2kg']:>8} {c['delivery_extra_1kg']:>8} {c['return_rule']:<30}")

wb.close()
