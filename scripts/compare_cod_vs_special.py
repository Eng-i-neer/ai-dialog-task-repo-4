"""对比模板COD回款中这两条特殊记录 vs 正常COD记录的区别
并在输入文件中寻找区分线索"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

specials = {'DE12510231810002', 'DE12511241510023'}

# === 模板分析 ===
tmpl = openpyxl.load_workbook(TEMPLATE, data_only=True)
ws = tmpl.worksheets[0]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

print("=== 模板 COD 回款：特殊 vs 正常 ===\n")

# 收集特殊行和正常行
special_rows = []
normal_rows = []
for r in range(2, ws.max_row + 1):
    wb_id = ws.cell(r, 4).value  # C4 运单号码
    if not wb_id or wb_id == '合计：':
        continue
    row_data = {}
    for c in range(1, ws.max_column + 1):
        row_data[headers[c-1] or f'C{c}'] = ws.cell(r, c).value
    row_data['_row'] = r
    row_data['_waybill'] = wb_id
    if str(wb_id) in specials:
        special_rows.append(row_data)
    else:
        normal_rows.append(row_data)

print(f"特殊行: {len(special_rows)} 条, 正常行: {len(normal_rows)} 条\n")

# 逐列对比
print("--- 逐列对比 ---")
for h in headers:
    if h is None:
        continue
    sp_vals = set()
    for sr in special_rows:
        sp_vals.add(repr(sr.get(h)))
    
    nm_vals = set()
    nm_none_count = 0
    nm_total = len(normal_rows)
    for nr in normal_rows:
        v = nr.get(h)
        if v is None:
            nm_none_count += 1
        nm_vals.add(type(v).__name__)
    
    sp_str = ', '.join(str(sr.get(h)) for sr in special_rows)
    
    # 只显示有差异的列
    all_normal_have = nm_none_count == 0
    all_special_none = all(sr.get(h) is None for sr in special_rows)
    
    if all_special_none and all_normal_have:
        print(f"  {h}: 特殊全None, 正常全有值")
    elif all_special_none and not all_normal_have:
        print(f"  {h}: 特殊全None, 正常{nm_total - nm_none_count}/{nm_total}有值")
    else:
        print(f"  {h}: 特殊=[{sp_str}], 正常None率={nm_none_count}/{nm_total}")

# 详细看特殊行
print("\n--- 特殊行详情 ---")
for sr in special_rows:
    print(f"\n  R{sr['_row']}: {sr['_waybill']}")
    for h in headers:
        if h is None:
            continue
        print(f"    {h}: {repr(sr.get(h))}")

# 看几条正常行作为对照
print("\n--- 正常行示例(前3条) ---")
for nr in normal_rows[:3]:
    print(f"\n  R{nr['_row']}: {nr['_waybill']}")
    for h in headers:
        if h is None:
            continue
        print(f"    {h}: {repr(nr.get(h))}")

tmpl.close()

# === 输入文件分析 ===
print("\n\n" + "="*80)
print("=== 输入文件：这两条 vs 其他COD运单 ===\n")

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# 1) 检查COD Sheet
print("--- 输入文件 COD Sheet ---")
ws_cod = inp['COD']
cod_headers = [ws_cod.cell(1, c).value for c in range(1, ws_cod.max_column + 1)]
print(f"  表头: {cod_headers[:16]}")

# 搜索这两条
found_in_cod = False
for r in range(2, ws_cod.max_row + 1):
    for c in range(1, min(ws_cod.max_column + 1, 20)):
        if str(ws_cod.cell(r, c).value or '') in specials:
            found_in_cod = True
            vals = [ws_cod.cell(r, cc).value for cc in range(1, min(ws_cod.max_column + 1, 16))]
            print(f"  R{r}: {vals}")
if not found_in_cod:
    print(f"  ** 这两条运单在COD Sheet中不存在 **")

# 2) 这两条出现的所有Sheet
print("\n--- 这两条在输入文件出现的完整记录 ---")
for t in sorted(specials):
    print(f"\n  运单: {t}")
    for sname in inp.sheetnames:
        ws2 = inp[sname]
        for r in range(1, ws2.max_row + 1):
            for c in range(1, min(ws2.max_column + 1, 20)):
                if str(ws2.cell(r, c).value or '') == t:
                    vals = {f'C{cc}': ws2.cell(r, cc).value for cc in range(1, min(ws2.max_column + 1, 16)) if ws2.cell(r, cc).value is not None}
                    print(f"    [{sname}] R{r}: {vals}")
                    break

# 3) 看看正常COD运单，随机取3条，在输入文件中出现的情况对比
print("\n--- 正常COD运单在输入文件的出现情况(对照) ---")
# 取模板正常COD的前3个运单号
sample_normals = [nr['_waybill'] for nr in normal_rows[:3]]
for t in sample_normals:
    print(f"\n  运单: {t}")
    for sname in inp.sheetnames:
        ws2 = inp[sname]
        for r in range(1, ws2.max_row + 1):
            for c in range(1, min(ws2.max_column + 1, 20)):
                if str(ws2.cell(r, c).value or '') == str(t):
                    fee_name = ws2.cell(r, 12).value
                    amount = ws2.cell(r, 13).value
                    print(f"    [{sname}] R{r}: 费用={fee_name} 金额={amount}")
                    break

# 4) 关键区分：正常COD运单是否都在COD Sheet中？
print("\n--- 正常COD运单是否都在输入文件COD Sheet中？ ---")
normal_waybills = set(nr['_waybill'] for nr in normal_rows)
cod_waybills_in_input = set()
for r in range(2, ws_cod.max_row + 1):
    for c in [3, 4]:  # 通常运单号在C3或C4
        v = ws_cod.cell(r, c).value
        if v and str(v).startswith(('DE', 'IT', 'HR')):
            cod_waybills_in_input.add(str(v))

overlap = normal_waybills & cod_waybills_in_input
print(f"  模板正常COD运单数: {len(normal_waybills)}")
print(f"  输入COD Sheet运单数: {len(cod_waybills_in_input)}")
print(f"  交集: {len(overlap)}")
print(f"  模板有但输入COD没有: {len(normal_waybills - cod_waybills_in_input)}")
if normal_waybills - cod_waybills_in_input:
    for w in sorted(normal_waybills - cod_waybills_in_input)[:5]:
        print(f"    {w}")

inp.close()
