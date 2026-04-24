# -*- coding: utf-8 -*-
"""Debug specific orders across periods."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'web'))
sys.stdout.reconfigure(encoding='utf-8')
os.environ['FLASK_ENV'] = 'testing'

from app import create_app, db
app = create_app()

WAYBILLS = ['IT12507101310001', 'IT12507071310019']

with app.app_context():
    from app.models import Order, OrderFee, FeeCategory

    for wb_no in WAYBILLS:
        print(f"\n{'='*80}")
        print(f"运单号: {wb_no}")
        print(f"{'='*80}")

        orders = Order.query.filter_by(waybill_no=wb_no).order_by(Order.bill_period).all()
        if not orders:
            print("  未找到订单!")
            continue

        print(f"  共 {len(orders)} 条记录:")
        for o in orders:
            period = o.bill_period.strftime('%Y%m%d') if o.bill_period else '?'
            print(f"\n  --- 期次: {period} (order.id={o.id}) ---")
            print(f"    import_log_id: {o.import_log_id}")
            print(f"    source_file: {o.source_file}")
            print(f"    customer_id: {o.customer_id}")
            print(f"    region: {o.region_id}")
            print(f"    import_sheets: {o.import_sheets}")
            print(f"    has_head_freight: {o.has_head_freight}")
            print(f"    has_tail_freight: {o.has_tail_freight}")
            print(f"    cod_amount: {o.cod_amount} {o.cod_currency}")
            print(f"    needs_return_fee: {o.needs_return_fee}")
            print(f"    needs_shelf_fee: {o.needs_shelf_fee}")
            print(f"    needs_vat: {o.needs_vat}")
            print(f"    product_name: {o.product_name}")
            print(f"    cargo_type: {o.cargo_type}")
            print(f"    actual_weight: {o.actual_weight}")
            print(f"    charge_weight_head: {o.charge_weight_head}")
            print(f"    charge_weight_tail: {o.charge_weight_tail}")

            fees = OrderFee.query.filter_by(order_id=o.id).all()
            if fees:
                print(f"    费用 ({len(fees)} 条):")
                for f in fees:
                    cat_name = f.category.name if f.category else '?'
                    cat_code = f.category.code if f.category else '?'
                    print(f"      {cat_code} ({cat_name}): input={f.input_amount}, calc={f.calculated_amount}, "
                          f"override={f.override_amount}, period={f.import_period}, sheet={f.source_sheet}")
            else:
                print(f"    费用: 无")

    # Also check what sheets these waybills appeared in from the raw Excel
    print(f"\n{'='*80}")
    print("从原始Excel检查这些运单出现在哪些Sheet中")
    print(f"{'='*80}")

    import openpyxl

    files = [
        (r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单\2025.07.28\（李志）鑫腾跃-中文-对账单20250728.xlsx', '0728'),
        (r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单\2025.08.11\（李志）鑫腾跃-中文-对账单20250811.xlsx', '0811'),
        (r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单\2025.08.05\（李志）鑫腾跃-中文-对账单20250805.xlsx', '0805'),
    ]

    for fpath, period in files:
        if not os.path.exists(fpath):
            continue
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=22, values_only=True):
                for cell in row:
                    if cell and any(wbn in str(cell) for wbn in WAYBILLS):
                        print(f"  [{period}] Sheet '{sname}': 找到 {cell}")
                        break
        wb.close()
