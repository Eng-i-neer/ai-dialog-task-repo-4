"""
报价规则 vs 客户模板公式 — 全科目逐条交叉比对

报价文件规则摘要:
  头程运费: 普货=9.7 特货=11.5 敏感货=12 (西班牙/葡萄牙普货=10.2 特货=12 敏感=13)
           计费: 0.1KG起, 体积/6000 (波兰DHL /4000)
  尾程派送费: 首2KG + 续1KG, 各国不同
  拒收返程费: "派送费*70%" 或 "同派送费"
  COD手续费: 3%, 最低1.5EUR (意大利2EUR, 奥地利5EUR, 德国7EUR)
  VAT: 1.2EUR/票 (代缴)
  清关费: (not explicitly stated in pricing, appears in templates)
  F货附加费: F货/纯电=2EUR/票, F手表=30CNY/票, 电子烟/保健品=1.5EUR/票
  上架费: 退仓上架1.5EUR/票
  转寄操作费: 1.5EUR/票
  偏远: 卡拉布里亚/西西里/撒丁=+0.8EUR/KG, 岛屿与威尼斯=+21EUR
"""
import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
}

PRICING_RULES = {
    '头程运费': {'普货_IC': {'rate': 10, 'desc': '普货 *重量(KG)*10 EUR/KG (≈9.7 报价 markup)'}, 
                 'GS': 7.7, 'SC': 9.5, 'IC': 10},
    '尾程派送费': {
        '波兰': {'first2': 4.0, 'extra1': 0.9, 'return_rule': '70%'},
        '罗马尼亚': {'first2': 4.8, 'extra1': 0.6, 'return_rule': '70%'},
        '匈牙利': {'first2': 4.4, 'extra1': 0.7, 'return_rule': '70%'},
        '捷克': {'first2': 4.1, 'extra1': 0.8, 'return_rule': '70%'},
        '斯洛伐克': {'first2': 4.3, 'extra1': 0.7, 'return_rule': '70%'},
        '保加利亚': {'first2': 4.3, 'extra1': 0.8, 'return_rule': '70%'},
        '克罗地亚': {'first2': 6.4, 'extra1': 1.0, 'return_rule': '70%'},
        '斯洛文尼亚': {'first2': 5.9, 'extra1': 0.9, 'return_rule': '70%'},
        '西班牙': {'first2': 4.0, 'extra1': 1.0, 'return_rule': '同派送费'},
        '葡萄牙': {'first2': 4.0, 'extra1': 1.0, 'return_rule': '同派送费'},
        '希腊': {'first2': 5.7, 'extra1': 0.8, 'return_rule': '同派送费'},
        '意大利': {'first2': 7.3, 'extra1': 1.1, 'return_rule': '同派送费'},
        '奥地利': {'first2': 6.5, 'extra1': 1.0, 'return_rule': '同派送费'},
        '德国': {'first2': 8.3, 'extra1': 1.5, 'return_rule': '同派送费'},
    },
    'COD手续费': {
        'default': {'rate': 0.03, 'min_eur': 1.5},
        '意大利': {'rate': 0.03, 'min_eur': 2.0, 'unsigned': 2.0},
        '奥地利': {'rate': 0.03, 'min_eur': 5.0, 'unsigned': 5.0},
        '德国': {'rate': 0.03, 'min_eur': 7.0, 'unsigned': 7.0},
    },
    'VAT税率': {
        '波兰': 0.23, '罗马尼亚': 0.19, '匈牙利': 0.27, '捷克': 0.21,
        '斯洛伐克': 0.23, '保加利亚': 0.20, '克罗地亚': 0.25,
        '斯洛文尼亚': 0.22, '西班牙': 0.21, '葡萄牙': 0.23,
        '希腊': 0.24, '意大利': 0.22, '奥地利': 0.20, '德国': 0.19,
    },
    'F附加费': {'F货_纯电': 2.0, 'F手表': 30.0, '电子烟_保健品': 1.5},
    '上架费': 1.5,
    '转寄操作费': 1.5,
}


def parse_formula_params(formula):
    """Extract numeric parameters from a formula string."""
    if not formula or not isinstance(formula, str):
        return {}
    params = {}
    # IFS pattern for head freight
    ifs_match = re.search(r'IFS\(.+"GS",\s*\w+\*([\d.]+)\s*,.+"SC",\s*\w+\*([\d.]+)\s*,.+"IC",\s*\w+\*([\d.]+)', formula)
    if ifs_match:
        params['GS_rate'] = float(ifs_match.group(1))
        params['SC_rate'] = float(ifs_match.group(2))
        params['IC_rate'] = float(ifs_match.group(3))

    # Tail freight: IF(K>2, base+(ROUNDUP(K,0)-2)*step, base)
    tail_match = re.search(r'IF\(\w+>2\s*,\s*([\d.]+)\s*\+.*\*([\d.]+)\s*,\s*([\d.]+)\)', formula)
    if tail_match:
        params['tail_first2'] = float(tail_match.group(1))
        params['tail_extra1'] = float(tail_match.group(2))
        params['tail_base'] = float(tail_match.group(3))

    # Tail with +addition: IF(K>2, base+(K-2)*step, base)+extra
    tail_plus = re.search(r'IF\(\w+>2\s*,\s*([\d.]+)\s*\+.*\*([\d.]+)\s*,\s*([\d.]+)\)\s*\+([\d.]+)', formula)
    if tail_plus:
        params['tail_first2'] = float(tail_plus.group(1))
        params['tail_extra1'] = float(tail_plus.group(2))
        params['tail_base'] = float(tail_plus.group(3))
        params['tail_addition'] = float(tail_plus.group(4))

    # COD fee: MAX(K*rate, min_eur*exchange)
    cod_match = re.search(r'MAX\(\w+\*([\d.]+)\*\d+\s*,\s*([\d.]+)\*([\d.]+)\)', formula)
    if cod_match:
        params['cod_rate'] = float(cod_match.group(1))
        params['cod_min_eur'] = float(cod_match.group(2))
        params['cod_exchange'] = float(cod_match.group(3))

    # COD offset: )-offset at end
    cod_offset = re.search(r'\)\s*-\s*([\d.]+)\s*$', formula)
    if cod_offset:
        params['cod_offset'] = float(cod_offset.group(1))

    # Exchange rate embedded: *7.9342
    rate_match = re.findall(r'\*(7\.\d+)', formula)
    if rate_match:
        params['exchange_rate'] = float(rate_match[0])

    # VAT formula: =10*rate or =申报*rate
    vat_match = re.search(r'^=(\d+)\*([\d.]+)$', formula)
    if vat_match:
        params['vat_base'] = int(vat_match.group(1))
        params['vat_rate'] = float(vat_match.group(2))

    # F surcharge: =ROUNDUP(amount*exchange, 2)
    f_match = re.search(r'ROUNDUP\(([\d.]+)\*([\d.]+)', formula)
    if f_match:
        params['f_eur'] = float(f_match.group(1))
        params['f_exchange'] = float(f_match.group(2))

    # Shelf fee: =ROUND(1.5*exchange, 2)
    shelf_match = re.search(r'ROUND\(([\d.]+)\*([\d.]+)', formula)
    if shelf_match:
        params['shelf_eur'] = float(shelf_match.group(1))

    return params


results = {}

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_val = openpyxl.load_workbook(str(fpath), data_only=True)

    print(f"\n{'='*80}")
    print(f"客户: {customer}")
    print(f"{'='*80}")

    cust_rules = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        ws_v = wb_val[sname]

        # ── 1. 运费 Sheet 公式 ──
        if '运费' in sname and '杂费' not in sname:
            print(f"\n  [{sname}] 运费科目公式:")
            for c in range(1, ws.max_column + 1):
                h = ws.cell(1, c).value
                for r in range(2, min(ws.max_row + 1, 10)):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith('='):
                        params = parse_formula_params(v)
                        print(f"    C{c} {h}: {v[:80]}")
                        if params:
                            print(f"       Params: {params}")
                        cust_rules[f'freight_C{c}'] = {'header': h, 'formula': v, 'params': params}
                        break

        # ── 2. COD Sheet 公式 ──
        elif 'COD' in sname:
            print(f"\n  [{sname}] COD科目公式:")
            for c in [11, 12, 13]:
                h = ws.cell(1, c).value
                for r in range(2, min(ws.max_row + 1, 20)):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith('='):
                        params = parse_formula_params(v)
                        print(f"    C{c} {h}: {v[:100]}")
                        if params:
                            print(f"       Params: {params}")
                        cust_rules[f'cod_C{c}'] = {'header': h, 'formula': v, 'params': params}
                        break

        # ── 3. 增值税 Sheet 公式 ──
        elif '增值税' in sname:
            print(f"\n  [{sname}] 增值税科目公式:")
            # Check multiple rows for different countries/rates
            vat_formulas = {}
            for r in range(2, min(ws.max_row + 1, 50)):
                v11 = ws.cell(r, 11).value
                dest = ws_v.cell(r, 7).value
                if isinstance(v11, str) and v11.startswith('='):
                    params = parse_formula_params(v11)
                    key = v11
                    if key not in vat_formulas:
                        vat_formulas[key] = {'formula': v11, 'params': params, 'destinations': []}
                    if dest:
                        vat_formulas[key]['destinations'].append(str(dest))

            for key, info in vat_formulas.items():
                dests = ', '.join(set(info['destinations']))
                print(f"    C11: {info['formula']} -> {dests}")
                if info['params']:
                    print(f"       Params: {info['params']}")

            # Check C12 (清关费)
            for r in range(2, min(ws.max_row + 1, 5)):
                v12 = ws_v.cell(r, 12).value
                h12 = ws.cell(1, 12).value
                if v12 is not None:
                    print(f"    C12 {h12}: value={v12}")
                    cust_rules['vat_clearing_fee'] = v12
                    break

            cust_rules['vat_formulas'] = vat_formulas

        # ── 4. 杂费 Sheet 公式 ──
        elif '杂费' in sname:
            print(f"\n  [{sname}] 杂费科目公式:")
            for c in range(11, ws.max_column + 1):
                h = ws.cell(1, c).value
                if not h:
                    continue
                for r in range(2, min(ws.max_row + 1, 10)):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith('='):
                        params = parse_formula_params(v)
                        print(f"    C{c} {h}: {v[:80]}")
                        if params:
                            print(f"       Params: {params}")
                        cust_rules[f'surcharge_C{c}'] = {'header': h, 'formula': v, 'params': params}
                        break
                    elif v is not None and c > 10 and isinstance(v, (int, float)):
                        print(f"    C{c} {h}: value={v}")
                        break

        # ── 5. F附加费 Sheet ──
        elif 'F' in sname and '附加' in sname:
            print(f"\n  [{sname}] F附加费科目:")
            # Check if formula or fixed value
            fee_col = 9
            h9 = ws.cell(1, fee_col).value
            values_seen = set()
            formula_seen = None
            for r in range(2, min(ws.max_row + 1, 50)):
                v = ws.cell(r, fee_col).value
                v_val = ws_v.cell(r, fee_col).value
                if isinstance(v, str) and v.startswith('='):
                    formula_seen = v
                    params = parse_formula_params(v)
                    if v_val:
                        values_seen.add(round(float(v_val), 2))
                elif isinstance(v, (int, float)):
                    values_seen.add(round(float(v), 2))
                elif isinstance(v_val, (int, float)):
                    values_seen.add(round(float(v_val), 2))

            print(f"    C{fee_col} {h9}: formula={formula_seen}")
            print(f"    Distinct values: {sorted(values_seen)}")
            if formula_seen:
                params = parse_formula_params(formula_seen)
                if params:
                    print(f"    Formula params: {params}")
            cust_rules['f_surcharge'] = {
                'header': h9, 'formula': formula_seen, 'values': sorted(values_seen),
            }

    results[customer] = cust_rules
    wb.close()
    wb_val.close()


# ── Cross comparison ──
print(f"\n\n{'='*80}")
print("CROSS COMPARISON: 报价规则 vs 客户公式参数")
print(f"{'='*80}")

print("\n── 1. 头程运费 单价 (GS/SC/IC) ──")
print(f"  报价标准: GS=7.7 SC=9.5 IC=10")
for customer, rules in results.items():
    for key, info in rules.items():
        if 'freight' in key and 'params' in info:
            p = info['params']
            if 'GS_rate' in p:
                match = p['GS_rate'] == 7.7 and p['SC_rate'] == 9.5 and p['IC_rate'] == 10
                print(f"  {customer}: GS={p['GS_rate']} SC={p['SC_rate']} IC={p['IC_rate']} {'✓' if match else '✗'}")

print("\n── 2. 尾程派送费 (首2KG/续1KG) ──")
print(f"  报价标准: 见报价文件各国")
for customer, rules in results.items():
    for key, info in rules.items():
        if 'freight' in key and 'params' in info:
            p = info['params']
            if 'tail_first2' in p:
                addition = p.get('tail_addition', 0)
                add_str = f" +{addition}" if addition else ""
                print(f"  {customer}: 首2KG={p['tail_first2']} 续1KG={p['tail_extra1']}{add_str} ({info['header']})")

print("\n── 3. COD手续费 (费率/最低/扣减) ──")
print(f"  报价标准: 3%, 最低按国家 1.5/2/5/7 EUR")
for customer, rules in results.items():
    for key, info in rules.items():
        if key == 'cod_C12' and 'params' in info:
            p = info['params']
            print(f"  {customer}: {info['formula'][:80]}")
            if p:
                print(f"    rate={p.get('cod_rate')}, min_eur={p.get('cod_min_eur')}, offset={p.get('cod_offset', 0)}")

print("\n── 4. 增值税率 (各国) ──")
print(f"  报价标准: 波兰=0.23, 意大利=0.22, 德国=0.19 等 (申报价值*税率)")
for customer, rules in results.items():
    vfs = rules.get('vat_formulas', {})
    if vfs:
        for key, info in vfs.items():
            p = info['params']
            dests = ', '.join(set(info['destinations']))[:40]
            if 'vat_rate' in p:
                expected = PRICING_RULES['VAT税率'].get(dests.split(',')[0].strip(), '?')
                match = p['vat_rate'] == expected if isinstance(expected, float) else '?'
                print(f"  {customer}: rate={p['vat_rate']} base={p.get('vat_base')} -> {dests} {'✓' if match == True else ''}")

print("\n── 5. F货/F手表 附加费 ──")
print(f"  报价标准: F货/纯电=2EUR/票, F手表=30CNY/票")
for customer, rules in results.items():
    f_info = rules.get('f_surcharge')
    if f_info:
        vals = f_info.get('values', [])
        formula = f_info.get('formula', 'N/A')
        print(f"  {customer}: values={vals}, formula={formula}")

print("\n── 6. 上架费 ──")
print(f"  报价标准: 1.5 EUR/票")
for customer, rules in results.items():
    for key, info in rules.items():
        if 'surcharge' in key and isinstance(info, dict):
            h = info.get('header', '')
            if '上架' in h:
                p = info.get('params', {})
                print(f"  {customer}: {h} formula={info.get('formula','')[:60]} params={p}")
