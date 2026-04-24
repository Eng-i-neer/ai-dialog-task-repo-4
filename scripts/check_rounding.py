"""分析参考文件中 CNY 的四舍五入规则"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
ref = openpyxl.load_workbook(REF_FILE, data_only=True)
ws = ref[ref.sheetnames[0]]

rate = 7.9342
mismatch_2dec = 0
mismatch_1dec = 0
match_2dec = 0
total = 0

for r in range(2, ws.max_row + 1):
    eur = ws.cell(r, 9).value
    cny = ws.cell(r, 11).value
    if not eur or not cny:
        continue
    total += 1

    calc_2dec = round(eur * rate, 2)
    calc_1dec = round(eur * rate, 1)

    if abs(cny - calc_2dec) < 0.005:
        match_2dec += 1
    elif abs(cny - calc_1dec) < 0.005:
        mismatch_2dec += 1
        mismatch_1dec += 0
    else:
        mismatch_2dec += 1
        mismatch_1dec += 1
        print(f"  Row {r}: EUR={eur}, ref_cny={cny}, calc2={calc_2dec}, calc1={calc_1dec}")

print(f"\nTotal: {total}")
print(f"Match round(,2): {match_2dec}")
print(f"Match round(,1) but not (,2): {mismatch_2dec - mismatch_1dec}")
print(f"No match either: {mismatch_1dec}")
