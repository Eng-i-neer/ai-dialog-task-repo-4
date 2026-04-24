"""Show Li Zhi's tail freight formula for a few countries + the +ROUND(1.5*7.9342,2) addition."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
fpath = BASE / '反馈客户' / '原始模板' / '20260330-汇森李志（东欧）对账单.xlsx'

wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

ws_f = wb_f['20260330期运费']
ws_v = wb_v['20260330期运费']

tail_col = 15
dest_col = 7

# Group by formula pattern
country_formulas = {}
for r in range(2, ws_f.max_row + 1):
    dest = ws_v.cell(r, dest_col).value
    formula = ws_f.cell(r, tail_col).value
    if dest is None:
        break
    if '合计' in str(dest or ''):
        break
    d = str(dest).strip()
    if d not in country_formulas:
        country_formulas[d] = {'formula': formula, 'row': r, 'count': 0}
    country_formulas[d]['count'] += 1

print("李志 尾程公式 按国家:")
print("=" * 120)
for country, info in country_formulas.items():
    formula = info['formula']
    f_str = str(formula or '')

    # Parse first2 and extra1 from the formula
    m = re.search(r'(\d+\.?\d*)\s*\+\s*\(.*?\)\s*\*\s*(\d+\.?\d*)\s*,\s*(\d+\.?\d*)', f_str)
    first2 = float(m.group(1)) if m else None
    extra1 = float(m.group(2)) if m else None

    has_extra = '+ROUND(1.5*7.9342,2)' in f_str or '+ROUND(1.5*' in f_str

    print(f"\n  {country} ({info['count']}条, 首行R{info['row']}):")
    print(f"    首2={first2} EUR, 续1={extra1} EUR")
    print(f"    含+1.5EUR转寄费: {'是' if has_extra else '否'}")
    print(f"    完整公式: {formula}")

wb_f.close()
wb_v.close()
