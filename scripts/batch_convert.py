"""
批量多客户转换器 — 支持李志型和通用型两种模板结构

模板族分类:
  - 李志型 (lizhi): 5 Sheet, Sheet名含账期前缀, COD[0]/运费[1]/尾程杂费[2]/仓储[3]/理赔[4]
  - 通用型 (generic): 6-8 Sheet, 汇总[0]/COD[1]/头程&尾程运费[2]/增值税&清关费[3]/尾程杂费[4]/F货附加费[5]/...
"""
import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from datetime import datetime
import openpyxl

# Import shared functions from convert_bill.py
sys.path.insert(0, str(Path(__file__).parent))
from convert_bill import (
    parse_input, determine_ship_type, is_formula, clear_data_rows,
    ensure_formulas, load_remote_zipcodes, calc_remote_fee_rmb,
    extract_zipcodes_from_template, DELIVERY_PRICING,
    _get_delivery_params, _calc_return_fee_rmb, _build_return_fee_formula,
)

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
INPUT_DIR = BASE / '中介提供'
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'
OUTPUT_BASE = BASE / '反馈客户' / '自动生成'
PRICING_FILE = BASE / '报价规则' / '汇森国际-东欧COD报价20260331生效(5).xlsx'

EXCHANGE_RATE = 7.9342
PRICING_VERSION = 'old'

CUSTOMER_MAP = {
    '中文':   {'customer': '李志',  'template_type': 'lizhi'},
    '中文1':  {'customer': '君悦',  'template_type': 'generic'},
    '中文3':  {'customer': '小美',  'template_type': 'generic'},
    '中文5':  {'customer': 'J',     'template_type': 'generic'},
    '中文6':  {'customer': '涵江',  'template_type': 'generic'},
    '中文7':  {'customer': '阿甘',  'template_type': 'generic'},
    '中文9':  {'customer': '威总',  'template_type': 'generic'},
    '中文12': {'customer': '峰总',  'template_type': 'generic'},
}


def find_template(customer):
    for f in os.listdir(TEMPLATE_DIR):
        if f.endswith('.xlsx') and customer in f:
            return TEMPLATE_DIR / f
    return None


def find_input(code):
    for f in os.listdir(INPUT_DIR):
        if f.endswith('.xlsx') and f'-{code}-' in f:
            return INPUT_DIR / f
    return None


def detect_template_layout(wb):
    """Auto-detect template layout by scanning sheet names."""
    names = wb.sheetnames
    layout = {
        'type': 'unknown',
        'cod_idx': None,
        'freight_idx': None,
        'vat_idx': None,
        'surcharge_idx': None,
        'f_surcharge_idx': None,
        'summary_idx': None,
        'storage_idx': None,
        'claim_idx': None,
    }
    for i, name in enumerate(names):
        nl = name.lower()
        if '汇总' in name:
            layout['summary_idx'] = i
        elif 'cod' in nl or 'COD' in name:
            layout['cod_idx'] = i
        elif '运费' in name and '杂费' not in name:
            layout['freight_idx'] = i
        elif '增值税' in name:
            layout['vat_idx'] = i
        elif '杂费' in name:
            layout['surcharge_idx'] = i
        elif 'f' in nl and '附加' in name:
            layout['f_surcharge_idx'] = i
        elif '仓储' in name:
            layout['storage_idx'] = i
        elif '理赔' in name:
            layout['claim_idx'] = i

    if layout['summary_idx'] is not None:
        layout['type'] = 'generic'
    else:
        layout['type'] = 'lizhi'

    return layout


def detect_col_mapping(ws):
    """Read row-1 headers and build a header->column mapping."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            mapping[str(h).strip()] = c
    return mapping


def find_data_end(ws, check_cols=10):
    """Find the last data row before an all-None or '合计' row."""
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, check_cols + 1))]
        if all(v is None for v in vals):
            return r - 1
        if any('合计' in str(v or '') for v in vals):
            return r - 1
    return ws.max_row


def fill_cod_generic(ws, parcels, exchange_rate, col_map):
    """Fill COD sheet for generic template.
    Generic COD: 14 cols, C5=转单号/C6=订单号 OR C5=客户单号/C6=转单号
    """
    data_end = find_data_end(ws, 14)
    clear_data_rows(ws, 2, data_end)

    cod_parcels = [p for p in parcels.values() if 'cod' in p['fees']]
    cod_parcels.sort(key=lambda p: p.get('date') or datetime.min)

    c5_header = ws.cell(1, 5).value or ''
    c6_header = ws.cell(1, 6).value or ''
    c5_is_transfer = '转单' in c5_header
    c5_is_cust = '客户' in c5_header or '订单' in c5_header

    for i, p in enumerate(cod_parcels):
        row = 2 + i
        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        if c5_is_transfer:
            ws.cell(row, 5, p.get('transfer_no', ''))
            ws.cell(row, 6, p.get('cust_ref', ''))
        else:
            ws.cell(row, 5, p.get('cust_ref', ''))
            ws.cell(row, 6, p.get('transfer_no', ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))
        ws.cell(row, 9, p['fees']['cod']['amount_eur'])
        ws.cell(row, 10, exchange_rate)

    if len(cod_parcels) > 0:
        all_rows = list(range(2, 2 + len(cod_parcels)))
        ensure_formulas(ws, max(2, 2 + len(cod_parcels) - 1), all_rows, [11, 12, 13])

    return len(cod_parcels)


def fill_freight_generic(ws, parcels, exchange_rate):
    """Fill 头程&尾程运费 sheet for generic template.
    Columns vary slightly but pattern is:
    C1=账期 C2=直发/转寄 C3=发货日期 C4=运单号码
    C5=转单号/客户单号 C6=客户单号/转单号 C7=目的地 C8=品名
    C9=重量(KG) C10=尺寸 C11=尾程计费重 C12=普特敏货
    C13+=formulas (头程运费/尾程运费/小计)
    
    Some templates (香隅) have C9=实重 C10=头程计费重 C11=尺寸 C12=尾程计费重
    """
    data_end = find_data_end(ws, 12)
    clear_data_rows(ws, 2, data_end)

    freight_parcels = [p for p in parcels.values()
                       if 'head_freight' in p['fees'] or 'tail_freight' in p['fees']]
    freight_parcels.sort(key=lambda p: p.get('date') or datetime.min)

    c5_header = str(ws.cell(1, 5).value or '')
    c5_is_cust = '客户' in c5_header
    c9_header = str(ws.cell(1, 9).value or '')
    has_actual_weight_c9 = '实重' in c9_header
    c10_header = str(ws.cell(1, 10).value or '')
    has_head_weight_c10 = '头程' in c10_header

    for i, p in enumerate(freight_parcels):
        row = 2 + i
        head_w = p.get('head_charge_weight') or 0
        tail_q = p.get('tail_charge_weight') or p.get('charge_weight') or 0

        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        if c5_is_cust:
            ws.cell(row, 5, p.get('cust_ref', ''))
            ws.cell(row, 6, p.get('transfer_no', ''))
        else:
            ws.cell(row, 5, p.get('transfer_no', ''))
            ws.cell(row, 6, p.get('cust_ref', ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))

        if has_actual_weight_c9 and has_head_weight_c10:
            ws.cell(row, 9, p.get('actual_weight'))
            ws.cell(row, 10, head_w)
            ws.cell(row, 11, p.get('dimensions', ''))
            ws.cell(row, 12, tail_q)
            ws.cell(row, 13, 'IC')
        else:
            ws.cell(row, 9, tail_q)
            ws.cell(row, 10, p.get('dimensions', ''))
            ws.cell(row, 11, tail_q)
            ws.cell(row, 12, 'IC')

    return len(freight_parcels)


def fill_vat_generic(ws, parcels, exchange_rate):
    """Fill 目的地增值税&清关费 sheet.
    C1=账期 C2=直发/转寄 C3=发货日期 C4=运单号码
    C5=转单号/客户单号 C6=客户单号/转单号 C7=目的地 C8=品名
    C9=重量(KG) C10=尺寸 C11=增值税(EUR)[formula] C12=清关费(EUR) C13=小计[formula]
    """
    data_end = find_data_end(ws, 10)
    clear_data_rows(ws, 2, data_end)

    vat_parcels = [p for p in parcels.values() if 'vat' in p['fees']]
    vat_parcels.sort(key=lambda p: p.get('date') or datetime.min)

    c5_header = str(ws.cell(1, 5).value or '')
    c5_is_cust = '客户' in c5_header

    for i, p in enumerate(vat_parcels):
        row = 2 + i
        ws.cell(row, 1, 46111)
        ws.cell(row, 2, determine_ship_type(p))
        ws.cell(row, 3, p['date'])
        ws.cell(row, 4, p['waybill'])
        if c5_is_cust:
            ws.cell(row, 5, p.get('cust_ref', ''))
            ws.cell(row, 6, p.get('transfer_no', ''))
        else:
            ws.cell(row, 5, p.get('transfer_no', ''))
            ws.cell(row, 6, p.get('cust_ref', ''))
        ws.cell(row, 7, p.get('dest', ''))
        ws.cell(row, 8, p.get('category', ''))
        ws.cell(row, 9, p.get('charge_weight') or p.get('actual_weight'))
        ws.cell(row, 10, p.get('dimensions', ''))

    return len(vat_parcels)


def fill_surcharge_generic(ws, parcels, exchange_rate, pricing_version,
                           remote_zips=None, parcel_zipcodes=None):
    """Fill 尾程杂费 sheet for generic template.
    Same logic as lizhi but column names may differ slightly.
    C11=上架费 C12=退件费 C13=二派费 C14=偏远费 C15=增值税(some) C16=小计
    OR
    C11=上架费(RMB) C12=尾程退件入仓费(RMB) C13=二派费 C14=偏远费 ... C16=小计
    """
    data_end = find_data_end(ws, 10)
    clear_data_rows(ws, 2, data_end)

    # Detect column layout from headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[c] = str(h).strip()

    shelf_col = None
    return_col = None
    vat_col = None
    remote_col = None
    subtotal_col = None
    remark_col = None
    zip_col = None
    region_col = None

    for c, h in headers.items():
        hl = h.lower()
        if '上架' in h:
            shelf_col = c
        elif '退件' in h or '返程' in h or '拒收' in h:
            return_col = c
        elif '增值税' in h and c > 10:
            vat_col = c
        elif '偏远' in h:
            remote_col = c
        elif '小计' in h:
            subtotal_col = c
        elif '备注' in h and c > 10:
            remark_col = c
        elif '邮编' in h:
            zip_col = c
        elif '地区' in h or '岛屿' in h:
            region_col = c

    if not remark_col:
        remark_col = (subtotal_col + 1) if subtotal_col else 18
    if not zip_col:
        zip_col = remark_col + 1
    if not region_col:
        region_col = zip_col + 1

    surcharge_fee_keys = {'shelf_fee', 'tail_return_fee', 'tail_return_refund', 'vat'}

    remote_parcels_set = set()
    if remote_zips and parcel_zipcodes:
        for wb_id, zipcode in parcel_zipcodes.items():
            if str(zipcode).strip() in remote_zips:
                remote_parcels_set.add(wb_id)

    merged = {}
    for p in parcels.values():
        wb_id = p['waybill']
        has_surcharge = any(ft in p['fees'] for ft in surcharge_fee_keys)
        has_remote = wb_id in remote_parcels_set

        if not has_surcharge and not has_remote:
            continue

        if wb_id not in merged:
            merged[wb_id] = {
                'date': p['date'],
                'waybill': wb_id,
                'cust_ref': p.get('cust_ref', ''),
                'transfer_no': p.get('transfer_no', ''),
                'category': p.get('category', ''),
                'pieces': p.get('pieces'),
                'charge_weight': p.get('charge_weight'),
                'dimensions': p.get('dimensions', ''),
                'dest': p.get('dest', ''),
                'remark_parts': [],
                'fees': p.get('fees', {}),
            }
        else:
            merged[wb_id]['fees'].update(p.get('fees', {}))

        rec = merged[wb_id]
        if 'shelf_fee' in p['fees'] and '上架费' not in rec['remark_parts']:
            rec['remark_parts'].append('上架费')
        if 'tail_return_fee' in p['fees'] and '退件操作费' not in rec['remark_parts']:
            rec['remark_parts'].append('退件操作费')
        if 'tail_return_refund' in p['fees'] and '签收件，拒收返程费补退' not in rec['remark_parts']:
            rec['remark_parts'].append('签收件，拒收返程费补退')
        if 'vat' in p['fees'] and '增值税' not in rec['remark_parts']:
            rec['remark_parts'].append('增值税')
        if has_remote and '偏远费' not in rec['remark_parts']:
            rec['remark_parts'].append('偏远费')

    sur_list = sorted(merged.values(), key=lambda x: x.get('date') or datetime.min)

    for i, item in enumerate(sur_list):
        row = 2 + i
        ws.cell(row, 1, item['date'])
        ws.cell(row, 2, item['waybill'])
        ws.cell(row, 3, item['cust_ref'])
        ws.cell(row, 4, item['transfer_no'])
        ws.cell(row, 5, item['category'])
        ws.cell(row, 6, item['pieces'])
        ws.cell(row, 7, item['charge_weight'])
        ws.cell(row, 8, item['dimensions'])
        ws.cell(row, 10, item['dest'])

        fees = item.get('fees', {})

        if shelf_col and 'shelf_fee' in fees:
            ws.cell(row, shelf_col, f'=ROUND(1.5*{exchange_rate},2)')

        if return_col and 'tail_return_fee' in fees:
            dest = str(item['dest'] or '')
            formula = _build_return_fee_formula(dest, row, exchange_rate, pricing_version)
            ws.cell(row, return_col, formula)
        if return_col and 'tail_return_refund' in fees:
            dest = str(item['dest'] or '')
            ret_rmb = _calc_return_fee_rmb(dest, exchange_rate, pricing_version)
            ws.cell(row, return_col, -abs(ret_rmb))

        if vat_col and 'vat' in fees:
            vat_eur = fees['vat'].get('amount_eur', 0) or 0
            if 'tail_return_refund' in fees:
                ws.cell(row, vat_col, -abs(round(vat_eur * exchange_rate, 2)))
            else:
                ws.cell(row, vat_col, f'=ROUNDUP({vat_eur}*{exchange_rate},2)')

        if remote_col and remote_zips and parcel_zipcodes:
            zipcode = parcel_zipcodes.get(item['waybill'])
            if zipcode and str(zipcode).strip() in remote_zips:
                zip_info = remote_zips[str(zipcode).strip()]
                weight = item.get('charge_weight') or 1
                remote_rmb = calc_remote_fee_rmb(zip_info['region'], weight, exchange_rate)
                ws.cell(row, remote_col, remote_rmb)
                if zip_col:
                    ws.cell(row, zip_col, str(zipcode).strip())
                if region_col:
                    ws.cell(row, region_col, zip_info['region'])

        if remark_col:
            remark_parts = item['remark_parts']
            if remark_parts:
                if '签收件，拒收返程费补退' in remark_parts:
                    ws.cell(row, remark_col, '签收件，拒收返程费补退')
                elif len(remark_parts) == 1:
                    display_map = {
                        '上架费': '上架费', '退件操作费': '尾程退件操作费',
                        '增值税': '目的地增值税', '偏远费': '偏远费',
                    }
                    ws.cell(row, remark_col, display_map.get(remark_parts[0], remark_parts[0]))
                else:
                    ws.cell(row, remark_col, '+'.join(remark_parts))

    if subtotal_col and len(sur_list) > 0:
        all_rows = list(range(2, 2 + len(sur_list)))
        ensure_formulas(ws, min(data_end, 4), all_rows, [subtotal_col])

    return len(sur_list)


def fill_f_surcharge_generic(ws, parcels, exchange_rate):
    """Fill F货&F手表附加费 sheet.
    This sheet is for F-brand surcharges. The input file doesn't have a direct
    'F货附加费' sheet, so we leave formulas intact and only clear/refill data cols.
    For now we preserve existing data since we have no mapping for this fee type.
    """
    # We don't have F-brand fee data in input parsing, so skip filling
    return 0


def fill_template_generic(template_path, input_data, output_path, exchange_rate,
                          remote_zips=None, parcel_zipcodes=None, pricing_version='old'):
    """Fill a generic (non-李志) template."""
    wb = openpyxl.load_workbook(template_path)
    layout = detect_template_layout(wb)

    results = {}

    if layout['cod_idx'] is not None:
        ws = wb.worksheets[layout['cod_idx']]
        col_map = detect_col_mapping(ws)
        n = fill_cod_generic(ws, input_data, exchange_rate, col_map)
        results['COD'] = n
        print(f"  COD: {n} rows")

    if layout['freight_idx'] is not None:
        ws = wb.worksheets[layout['freight_idx']]
        n = fill_freight_generic(ws, input_data, exchange_rate)
        results['Freight'] = n
        print(f"  Freight: {n} rows")

    if layout['vat_idx'] is not None:
        ws = wb.worksheets[layout['vat_idx']]
        n = fill_vat_generic(ws, input_data, exchange_rate)
        results['VAT'] = n
        print(f"  VAT: {n} rows")

    if layout['surcharge_idx'] is not None:
        ws = wb.worksheets[layout['surcharge_idx']]
        n = fill_surcharge_generic(ws, input_data, exchange_rate, pricing_version,
                                   remote_zips, parcel_zipcodes)
        results['Surcharge'] = n
        print(f"  Surcharge: {n} rows")

    if layout['f_surcharge_idx'] is not None:
        results['F-Surcharge'] = 0

    wb.save(output_path)
    wb.close()
    print(f"  Saved: {output_path}")
    return results


def fill_template_lizhi(template_path, input_data, output_path, exchange_rate,
                        remote_zips=None, parcel_zipcodes=None, pricing_version='old'):
    """Fill 李志 template using the original logic from convert_bill.py."""
    from convert_bill import fill_template
    fill_template(template_path, input_data, output_path, 46111, exchange_rate,
                  remote_zips=remote_zips, parcel_zipcodes=parcel_zipcodes,
                  pricing_version=pricing_version)


def main():
    print("Loading remote zipcodes...")
    remote_zips = load_remote_zipcodes(str(PRICING_FILE))
    print(f"  Loaded {len(remote_zips)} remote zipcodes")

    results_all = {}

    for code, info in CUSTOMER_MAP.items():
        customer = info['customer']
        ttype = info['template_type']

        input_path = find_input(code)
        template_path = find_template(customer)

        if not input_path or not template_path:
            print(f"\nSKIP {code} -> {customer}: input={input_path is not None}, template={template_path is not None}")
            continue

        print(f"\n{'='*60}")
        print(f"Processing: {code} -> {customer} (type={ttype})")
        print(f"  Input: {input_path.name}")
        print(f"  Template: {template_path.name}")

        # Parse input
        parcels = parse_input(str(input_path))
        print(f"  Parsed {len(parcels)} waybills")

        # Extract zipcodes from template surcharge sheet
        parcel_zipcodes = {}
        try:
            parcel_zipcodes = extract_zipcodes_from_template(str(template_path))
        except Exception:
            pass

        # Output
        out_dir = OUTPUT_BASE / customer
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f'20260330-{customer}-自动生成.xlsx'

        if ttype == 'lizhi':
            fill_template_lizhi(str(template_path), parcels, str(out_file),
                                EXCHANGE_RATE, remote_zips, parcel_zipcodes, PRICING_VERSION)
        else:
            fill_template_generic(str(template_path), parcels, str(out_file),
                                  EXCHANGE_RATE, remote_zips, parcel_zipcodes, PRICING_VERSION)

        results_all[customer] = {'code': code, 'parcels': len(parcels), 'output': str(out_file)}

    print(f"\n\n{'='*60}")
    print("BATCH CONVERSION COMPLETE")
    print(f"{'='*60}")
    for customer, r in results_all.items():
        print(f"  {customer}: {r['parcels']} waybills -> {Path(r['output']).name}")


if __name__ == '__main__':
    main()
