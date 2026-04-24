"""
Print the exact tail-freight formulas for each customer, for each distinct destination country.
Also print head-freight formulas in full for understanding.
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

PRICING = {
    '波兰':     {'first2': 4.0, 'extra1': 0.9},
    '罗马尼亚':  {'first2': 4.8, 'extra1': 0.6},
    '匈牙利':   {'first2': 4.4, 'extra1': 0.7},
    '捷克':     {'first2': 4.1, 'extra1': 0.8},
    '斯洛伐克':  {'first2': 4.3, 'extra1': 0.7},
    '保加利亚':  {'first2': 4.3, 'extra1': 0.8},
    '克罗地亚':  {'first2': 6.4, 'extra1': 1.0},
    '斯洛文尼亚': {'first2': 5.9, 'extra1': 0.9},
    '西班牙':   {'first2': 4.0, 'extra1': 1.0},
    '葡萄牙':   {'first2': 4.0, 'extra1': 1.0},
    '希腊':     {'first2': 5.7, 'extra1': 0.8},
    '意大利':   {'first2': 7.3, 'extra1': 1.1},
    '奥地利':   {'first2': 6.5, 'extra1': 1.0},
    '德国':     {'first2': 8.3, 'extra1': 1.5},
}

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    ws_f = ws_v = None
    for sname in wb_f.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws_f = wb_f[sname]
            ws_v = wb_v[sname]
            break
    if not ws_f:
        wb_f.close()
        wb_v.close()
        continue

    tail_col = head_col = dest_col = None
    for c in range(1, ws_f.max_column + 1):
        h = str(ws_f.cell(1, c).value or '')
        if '尾程' in h and '运费' in h and '(EUR)' in h:
            tail_col = c
        elif '尾程运费' in h:
            if tail_col is None:
                tail_col = c
        if '头程运费' in h:
            head_col = c
        if '目的' in h:
            dest_col = c

    print(f"\n{'='*80}")
    print(f"{customer} — {ws_f.title}")
    print(f"  头程运费列=C{head_col}, 尾程运费列=C{tail_col}, 目的地列=C{dest_col}")
    print(f"{'='*80}")

    # Print head formula row 2
    hf = ws_f.cell(2, head_col).value if head_col else None
    print(f"\n  头程公式(R2): {hf}")

    # Collect tail formulas per distinct country
    seen = {}
    for r in range(2, ws_f.max_row + 1):
        dest = ws_v.cell(r, dest_col).value if dest_col else None
        if dest is None:
            break
        d = str(dest).strip()
        if '合计' in d:
            break
        if d not in seen:
            tf = ws_f.cell(r, tail_col).value
            seen[d] = (r, tf)

    print(f"\n  尾程公式 (按国家第一次出现):")
    for country, (row, formula) in seen.items():
        print(f"\n    [{country}] (row {row}):")
        print(f"      {formula}")

        # Try to match pricing
        pricing = PRICING.get(country)
        if pricing and isinstance(formula, str):
            # Extract all numbers from the formula
            nums = [float(x) for x in re.findall(r'[\d]+\.[\d]+|[\d]+', formula) if float(x) > 0.1]
            f2 = pricing['first2']
            e1 = pricing['extra1']
            if f2 in nums and e1 in nums:
                print(f"      → 报价匹配 ✓ (首2={f2}, 续1={e1})")
            else:
                print(f"      → 报价不匹配 ✗ 报价=(首2={f2}, 续1={e1}), 公式数值={nums}")

    wb_f.close()
    wb_v.close()


# ============================================================
# PART 2: Head freight formula deep dive - show an example with
# the input file side
# ============================================================
print(f"\n\n{'='*80}")
print("头程运费 — 详细拆解")
print("="*80)

# Load pricing file for reference
pricing_path = BASE / '汇森国际-东欧COD报价20260331生效(5).xlsx'
wb_p = openpyxl.load_workbook(str(pricing_path), data_only=True)
ws_cod = wb_p['欧洲COD']

print(f"\n报价文件 [欧洲COD] 头程运费标准:")
for r in range(1, ws_cod.max_row + 1):
    for c in range(1, ws_cod.max_column + 1):
        v = ws_cod.cell(r, c).value
        if v and '头程' in str(v):
            print(f"  R{r}: ", end='')
            for cc in range(1, 10):
                print(f"C{cc}={ws_cod.cell(r, cc).value}  ", end='')
            print()
            # also print the next row for rates
            r2 = r + 1
            print(f"  R{r2}: ", end='')
            for cc in range(1, 10):
                print(f"C{cc}={ws_cod.cell(r2, cc).value}  ", end='')
            print()
            break
    else:
        continue
    break

wb_p.close()

# Now trace with input file
inp_path = BASE / '中介提供'
input_files = list(inp_path.glob('*.xlsx'))
if input_files:
    inp = input_files[0]
    wb_in = openpyxl.load_workbook(str(inp), data_only=True)
    head_ws = None
    for sname in wb_in.sheetnames:
        if '头程运费' in sname:
            head_ws = wb_in[sname]
            break
    if head_ws:
        print(f"\n输入文件 [{head_ws.title}] 头几行结构:")
        for r in range(1, 8):
            vals = []
            for c in range(1, 16):
                v = head_ws.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}={v}")
            if vals:
                print(f"  R{r}: {', '.join(vals)}")

        print(f"\n输入文件 头程运费 数据行示例:")
        hr = None
        for r in range(1, 15):
            v = head_ws.cell(r, 3).value
            if v and '运单' in str(v):
                hr = r
                break
        if hr:
            print(f"  Header(R{hr}): ", end='')
            for c in range(1, 16):
                v = head_ws.cell(hr, c).value
                if v:
                    print(f"C{c}={v}  ", end='')
            print()
            for r in range(hr + 1, hr + 6):
                print(f"  R{r}: ", end='')
                for c in range(1, 16):
                    v = head_ws.cell(r, c).value
                    if v is not None:
                        print(f"C{c}={v}  ", end='')
                print()
    wb_in.close()
