"""找出所有"不在COD Sheet中"的运单，看它们出现在哪些Sheet"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import defaultdict

INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# 1) 收集COD Sheet中的所有运单号
cod_waybills = set()
ws_cod = inp['COD']
for r in range(10, ws_cod.max_row + 1):
    v = ws_cod.cell(r, 3).value
    if v and isinstance(v, str) and len(v) > 5:
        cod_waybills.add(v)
print(f"COD Sheet 运单数: {len(cod_waybills)}")

# 2) 收集所有Sheet中的运单号及其出现的Sheet
waybill_sheets = defaultdict(set)
for sname in inp.sheetnames:
    if sname == '汇总':
        continue
    ws = inp[sname]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if v and isinstance(v, str) and (v.startswith('DE') or v.startswith('IT') or v.startswith('HR') or v.startswith('GR')):
            waybill_sheets[v].add(sname)

print(f"所有Sheet中出现的总运单数: {len(waybill_sheets)}")

# 3) 找出不在COD中的运单
non_cod = {wb: sheets for wb, sheets in waybill_sheets.items() if wb not in cod_waybills}
print(f"不在COD Sheet中的运单数: {len(non_cod)}")

# 4) 按出现的Sheet组合分组
combo_groups = defaultdict(list)
for wb, sheets in non_cod.items():
    key = tuple(sorted(sheets))
    combo_groups[key].append(wb)

print(f"\n不在COD中的运单，按Sheet组合分组:")
for combo, wbs in sorted(combo_groups.items(), key=lambda x: -len(x[1])):
    print(f"\n  {' + '.join(combo)}: {len(wbs)}条")
    for wb in sorted(wbs)[:5]:
        print(f"    {wb}")
    if len(wbs) > 5:
        print(f"    ...还有{len(wbs)-5}条")

# 5) 特别看"只在 上架费+尾程退件操作费"的运单
target_combo = tuple(sorted(['上架费', '尾程退件操作费']))
if target_combo in combo_groups:
    only_shelf_return = combo_groups[target_combo]
    print(f"\n\n=== 只在 上架费+尾程退件操作费 中出现的运单: {len(only_shelf_return)}条 ===")
    specials = {'DE12510231810002', 'DE12511241510023'}
    for wb in sorted(only_shelf_return):
        flag = ' <<<< 模板COD特殊行' if wb in specials else ''
        print(f"  {wb}{flag}")
else:
    print(f"\n  没有运单只在 上架费+尾程退件操作费 这个组合中")

# 6) 看看这两条特殊的和其他"只在上架费+退件操作费"的运单有什么不同
specials = {'DE12510231810002', 'DE12511241510023'}
print(f"\n\n=== 对比：特殊行 vs 其他同组合运单 ===")
if target_combo in combo_groups:
    for wb in sorted(combo_groups[target_combo]):
        is_special = wb in specials
        # 看日期
        for sname in ['上架费']:
            ws = inp[sname]
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 3).value == wb:
                    date = ws.cell(r, 2).value
                    amount = ws.cell(r, 13).value
                    print(f"  {wb}: 日期={date} 上架费={amount} {'<<<< 特殊' if is_special else ''}")
                    break

inp.close()
