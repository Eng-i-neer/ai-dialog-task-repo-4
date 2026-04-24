"""检查模板文件的sheet结构和数据行范围"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
wb = openpyxl.load_workbook(TEMPLATE)

print(f"Sheets: {wb.sheetnames}")
for idx, sname in enumerate(wb.sheetnames):
    ws = wb[sname]
    print(f"\n--- Sheet[{idx}]: {sname} ---")
    print(f"  max_row={ws.max_row}, max_col={ws.max_column}")
    
    # 打印header (row 1)
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(f"C{c}={v}")
    print(f"  Headers: {headers}")
    
    # 检查公式列 - 扫描前10个数据行
    print(f"  数据行公式检查 (R2-R10):")
    for r in range(2, min(12, ws.max_row + 1)):
        formulas = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith('='):
                formulas.append(f"C{c}={v[:60]}")
        if formulas:
            print(f"    R{r}: {formulas}")
        else:
            # 打印数据值
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"C{c}={repr(v)[:40]}")
            if vals:
                print(f"    R{r} (data): {vals}")
            else:
                print(f"    R{r} (empty)")

wb.close()
