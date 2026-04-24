"""扫描所有原始模板的Sheet结构、汇率、客户编号等关键参数"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

TEMPLATE_DIR = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\反馈客户\原始模板'

for fname in sorted(os.listdir(TEMPLATE_DIR)):
    if not fname.endswith('.xlsx'):
        continue
    fpath = os.path.join(TEMPLATE_DIR, fname)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    
    print(f"\n{'='*80}")
    print(f"模板: {fname}")
    print(f"Sheets: {wb.sheetnames}")
    
    for i, sname in enumerate(wb.sheetnames):
        ws = wb[sname]
        # 表头
        headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column + 1, 22))]
        # 数据行数（找到第一个全空行或合计行）
        data_rows = 0
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 10))]
            if all(v is None for v in vals):
                break
            # 检查是否是合计行
            any_total = any('合计' in str(v or '') for v in vals)
            if any_total:
                break
            data_rows += 1
        
        print(f"  [{i}] {sname}: {data_rows} data rows, {ws.max_column} cols")
        print(f"      Headers: {[h for h in headers if h][:12]}")
        
        # 如果是COD回款Sheet(index 0)，提取客户编号和汇率
        if i == 0 and data_rows > 0:
            # 客户编号(C1=账期)
            acct = ws.cell(2, 1).value
            # 汇率
            rate = ws.cell(2, 10).value
            if rate is None:
                for r in range(2, min(data_rows + 2, 20)):
                    rate = ws.cell(r, 10).value
                    if rate and isinstance(rate, (int, float)):
                        break
            print(f"      客户编号(C1): {acct}, 汇率(C10): {rate}")
            # 第一条运单日期
            first_date = ws.cell(2, 3).value
            print(f"      首行日期: {first_date}")
    
    wb.close()
