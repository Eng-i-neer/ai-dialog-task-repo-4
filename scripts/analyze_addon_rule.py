"""精确分析 addon=30 vs 11.91 的判断条件"""
import openpyxl

REF_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

ref = openpyxl.load_workbook(REF_FILE, data_only=True)
inp = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# Build sign_fee lookup from input (签收费)
sign_fee_wbs = set()
for sname in inp.sheetnames:
    if '签收' in sname:
        ws_tmp = inp[sname]
        for r in range(10, ws_tmp.max_row + 1):
            wb = ws_tmp.cell(r, 3).value
            if wb:
                sign_fee_wbs.add(wb)
        print(f"签收费 sheet '{sname}': {ws_tmp.max_row - 9} rows")

# Build 转仓 lookup
warehouse_wbs = set()
for sname in inp.sheetnames:
    if '转仓' in sname:
        ws_tmp = inp[sname]
        for r in range(10, ws_tmp.max_row + 1):
            wb = ws_tmp.cell(r, 3).value
            if wb:
                warehouse_wbs.add(wb)
        print(f"转仓 sheet '{sname}': {ws_tmp.max_row - 9} rows")

ws_freight = ref[ref.sheetnames[1]]
addon_30_has_sign = 0
addon_30_no_sign = 0
addon_12_has_sign = 0
addon_12_no_sign = 0

addon_30_has_wh = 0
addon_12_has_wh = 0

# Check if addon=30 -> has 签收费 or 转仓
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    if not wb or not addon:
        continue
    has_sign = wb in sign_fee_wbs
    has_wh = wb in warehouse_wbs
    if abs(addon - 30) < 0.1:
        if has_sign:
            addon_30_has_sign += 1
        else:
            addon_30_no_sign += 1
        if has_wh:
            addon_30_has_wh += 1
    elif abs(addon - 11.91) < 0.1:
        if has_sign:
            addon_12_has_sign += 1
        else:
            addon_12_no_sign += 1
        if has_wh:
            addon_12_has_wh += 1

print(f"\naddon=30:  has_sign={addon_30_has_sign}, no_sign={addon_30_no_sign}, has_warehouse={addon_30_has_wh}")
print(f"addon=11.91: has_sign={addon_12_has_sign}, no_sign={addon_12_no_sign}, has_warehouse={addon_12_has_wh}")

# Check: is addon=30 when charge_weight<=1kg (head)?
print("\n=== addon vs charge_weight (德国only) ===")
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    dest = ws_freight.cell(r, 7).value
    head_w = ws_freight.cell(r, 10).value
    if not wb or not addon or not dest:
        continue
    if '德' not in dest:
        continue
    if abs(addon - 30) < 0.1 and head_w and head_w > 1:
        print(f"  EXCEPTION: addon=30 but head_w={head_w} @ {wb}")

print("\n=== 加费=30时head_weight分布(德国) ===")
hw_vals_30 = []
hw_vals_12 = []
for r in range(2, ws_freight.max_row + 1):
    wb = ws_freight.cell(r, 4).value
    addon = ws_freight.cell(r, 17).value
    dest = ws_freight.cell(r, 7).value
    head_w = ws_freight.cell(r, 10).value or 0
    if not wb or not addon:
        continue
    if abs(addon - 30) < 0.1:
        hw_vals_30.append(head_w)
    elif abs(addon - 11.91) < 0.1:
        hw_vals_12.append(head_w)

print(f"addon=30 head_weight: {sorted(set(hw_vals_30))}")
print(f"addon=11.91 head_weight: {sorted(set(hw_vals_12))}")
