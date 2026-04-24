"""
Trace how input fee categories aggregate into template output columns.
Focus on the "hidden" mappings - fees that don't have obvious sheet targets.

Key question: 服务费, XX地派服务费, 短信费, 转寄操作费, 代理送货费, 账号管理费, 海外仓操作费
  → Where do these go in the output?

Hypothesis from 李志 template analysis:
  - 运费 Sheet 的 C14=头程运费 C15=尾程运费 are formula cols
  - Let's check what those formulas compute
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

customers = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
}

print("=" * 80)
print("PART A: 运费 Sheet 公式分析 — 头程/尾程运费如何计算")
print("=" * 80)

for customer, fname in customers.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=False)

    # Find freight sheet
    for sname in wb.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws = wb[sname]
            print(f"\n--- {customer} / {sname} ---")
            for c in range(1, ws.max_column + 1):
                h = ws.cell(1, c).value
                if h:
                    print(f"  C{c} = {h}", end='')
                    # Check formula in row 2
                    v2 = ws.cell(2, c).value
                    if isinstance(v2, str) and v2.startswith('='):
                        print(f"  [F: {v2}]")
                    else:
                        print()
            break
    wb.close()


print(f"\n\n{'='*80}")
print("PART B: COD Sheet 公式 — COD手续费如何计算")
print("=" * 80)

for customer, fname in customers.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=False)

    for sname in wb.sheetnames:
        if 'COD' in sname:
            ws = wb[sname]
            print(f"\n--- {customer} / {sname} ---")
            for c in [11, 12, 13]:
                h = ws.cell(1, c).value
                # Find first formula row
                for r in range(2, min(ws.max_row + 1, 20)):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith('='):
                        print(f"  C{c} ({h}): {v}")
                        break
            break
    wb.close()


print(f"\n\n{'='*80}")
print("PART C: F货附加费 Sheet 公式 — 附加费如何计算")
print("=" * 80)

for customer, fname in customers.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=False)

    for sname in wb.sheetnames:
        if 'F' in sname and '附加' in sname:
            ws = wb[sname]
            print(f"\n--- {customer} / {sname} ---")
            for c in range(1, ws.max_column + 1):
                h = ws.cell(1, c).value
                if h:
                    print(f"  C{c} = {h}", end='')
                    for r in range(2, min(ws.max_row + 1, 6)):
                        v = ws.cell(r, c).value
                        if isinstance(v, str) and v.startswith('='):
                            print(f"  [F: {v}]")
                            break
                        elif v is not None:
                            print(f"  [V: {v}]")
                            break
                    else:
                        print()
            break
    wb.close()


print(f"\n\n{'='*80}")
print("PART D: 增值税 Sheet 公式")
print("=" * 80)

for customer, fname in customers.items():
    fpath = TEMPLATE_DIR / fname
    wb = openpyxl.load_workbook(str(fpath), data_only=False)

    for sname in wb.sheetnames:
        if '增值税' in sname:
            ws = wb[sname]
            print(f"\n--- {customer} / {sname} ---")
            for c in range(1, ws.max_column + 1):
                h = ws.cell(1, c).value
                if h:
                    print(f"  C{c} = {h}", end='')
                    for r in range(2, min(ws.max_row + 1, 6)):
                        v = ws.cell(r, c).value
                        if isinstance(v, str) and v.startswith('='):
                            print(f"  [F: {v}]")
                            break
                        elif v is not None:
                            print(f"  [V: {v}]")
                            break
                    else:
                        print()
            break
    wb.close()


print(f"\n\n{'='*80}")
print("PART E: 逐运单追踪 — 一个运单在输入中的所有科目 vs 模板中的所有出现")
print("=" * 80)

# Pick 君悦 as example (has diverse fee types)
input_path = None
for f in os.listdir(BASE / '中介提供'):
    if '-中文1-' in f and f.endswith('.xlsx'):
        input_path = BASE / '中介提供' / f
        break

template_path = TEMPLATE_DIR / '汇森-君悦（东欧）对账单-20260330.xlsx'

# Parse input: build waybill -> list of (sheet, amount)
wb_in = openpyxl.load_workbook(str(input_path), data_only=True)
waybill_fees = {}
for sname in wb_in.sheetnames:
    if sname in ('汇总', '总表'):
        continue
    ws = wb_in[sname]
    header_row = None
    for r in range(1, 15):
        v = ws.cell(r, 3).value
        if v and '运单' in str(v):
            header_row = r
            break
    if not header_row:
        continue
    for r in range(header_row + 1, ws.max_row + 1):
        wb_id = ws.cell(r, 3).value
        if not wb_id or not isinstance(wb_id, str):
            continue
        wb_id = wb_id.strip()
        amt = ws.cell(r, 13).value
        if wb_id not in waybill_fees:
            waybill_fees[wb_id] = []
        waybill_fees[wb_id].append((sname, amt))
wb_in.close()

# Parse template: build waybill -> list of (sheet, row, values)
wb_tmpl = openpyxl.load_workbook(str(template_path), data_only=True)
waybill_template = {}
for sname in wb_tmpl.sheetnames:
    ws = wb_tmpl[sname]
    if '汇总' in sname:
        continue
    # Find waybill col
    wb_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h and '运单' in str(h):
            wb_col = c
            break
    if not wb_col:
        continue
    for r in range(2, ws.max_row + 1):
        wb_id = ws.cell(r, wb_col).value
        if not wb_id:
            continue
        wb_id = str(wb_id).strip()
        vals = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            v = ws.cell(r, c).value
            if v is not None and h:
                vals[str(h).strip()] = v
        if wb_id not in waybill_template:
            waybill_template[wb_id] = []
        waybill_template[wb_id].append((sname, r, vals))
wb_tmpl.close()

# Print 3 sample waybills that have the most fee types
samples = sorted(waybill_fees.items(), key=lambda x: -len(x[1]))[:3]
for wb_id, fees in samples:
    print(f"\n  运单: {wb_id}")
    print(f"  输入科目({len(fees)}):")
    total_eur = 0
    for sname, amt in fees:
        print(f"    [{sname}] = {amt} EUR")
        total_eur += (amt or 0)
    print(f"    合计: {total_eur:.2f} EUR")

    if wb_id in waybill_template:
        print(f"  模板出现({len(waybill_template[wb_id])}):")
        for sname, row, vals in waybill_template[wb_id]:
            fee_vals = {k: v for k, v in vals.items()
                        if any(kw in k for kw in ['运费', '手续费', '附加', '金额', '小计', '增值',
                                                    '清关', '上架', '退件', '偏远', '返程', '费'])}
            print(f"    [{sname}] R{row}: {fee_vals}")
