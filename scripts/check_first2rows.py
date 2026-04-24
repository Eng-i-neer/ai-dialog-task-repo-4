"""检查：前两行数据到底是模板残留还是输入文件解析的"""
import openpyxl

# 模板文件
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
# 输入文件
INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'

tmpl = openpyxl.load_workbook(TEMPLATE, data_only=True)
inp = openpyxl.load_workbook(INPUT, data_only=True)

ws_tmpl = tmpl[tmpl.sheetnames[0]]

# 模板前几行
print("=== 模板 COD Sheet Row 2-3 ===")
for r in [2, 3]:
    print(f"  Row {r}:")
    for c in range(1, 15):
        v = ws_tmpl.cell(r, c).value
        if v is not None:
            print(f"    Col {c}: {repr(v)}")

# 检查这两个运单号是否在输入文件中
wb1 = 'DE12510231810002'  # Row 2
wb2 = 'DE12511241510023'  # Row 3

print(f"\n=== 在输入文件中搜索 {wb1} ===")
ws_cod = inp['COD']
found1 = False
for r in range(10, ws_cod.max_row + 1):
    if ws_cod.cell(r, 3).value == wb1:
        found1 = True
        print(f"  Found in COD row {r}")
        break
if not found1:
    print(f"  NOT FOUND in COD sheet")
    for sname in inp.sheetnames:
        ws = inp[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(19, ws.max_column + 1)):
                if ws.cell(r, c).value == wb1:
                    print(f"  Found in sheet '{sname}' row {r} col {c}")

print(f"\n=== 在输入文件中搜索 {wb2} ===")
found2 = False
for r in range(10, ws_cod.max_row + 1):
    if ws_cod.cell(r, 3).value == wb2:
        found2 = True
        print(f"  Found in COD row {r}")
        break
if not found2:
    print(f"  NOT FOUND in COD sheet")
    for sname in inp.sheetnames:
        ws = inp[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(19, ws.max_column + 1)):
                if ws.cell(r, c).value == wb2:
                    print(f"  Found in sheet '{sname}' row {r} col {c}")

# 模板备注列说了什么
print(f"\n=== 模板 Row 2-3 备注 (Col N) ===")
print(f"  Row 2: {ws_tmpl.cell(2, 14).value}")
print(f"  Row 3: {ws_tmpl.cell(3, 14).value}")
