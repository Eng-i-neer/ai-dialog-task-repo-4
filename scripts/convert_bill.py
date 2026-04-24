"""
对账单转换器 v5 — 复制模板 + 清空 + 填入数据

模板结构（从实际文件确认）：

Sheet[0] 20260330期COD回款:
  Col1=账期 Col2=直发/转寄 Col3=发货日期 Col4=运单号码 Col5=订单号
  Col6=转单号 Col7=目的地 Col8=品名 Col9=代收金额 Col10=汇率
  Col11=金额(CNY) [公式] Col12=COD手续费(CNY) [公式] Col13=小计 [公式] Col14=备注
  R2-R3: 手工历史数据（无公式），需要补公式
  R4+:   有公式

Sheet[1] 20260330期运费:
  Col1=账期 Col2=直发/转寄 Col3=发货日期 Col4=运单号码 Col5=订单号
  Col6=转单号 Col7=目的地 Col8=品名 Col9=收件实重 Col10=头程计费重量(KG)
  Col11=尺寸 Col12=尾程计费重 Col13=普特敏货
  Col14=头程运费[公式] Col15=尾程运费[公式] Col16=代收手续费[公式]
  Col17=附加费[公式] Col18=德国旺季附加费 Col19=小计[公式]
  R2+: 全部有公式

Sheet[2] 20260330期尾程杂费:
  Col1=寄件日期 Col2=运单号码 Col3=客户单号 Col4=转单号 Col5=品名
  Col6=件数 Col7=重量(KG) Col8=尺寸 Col9=类型 Col10=目的地
  Col11=科目 Col12=上架费(RMB)[公式] Col13=尾程退件入仓费(RMB) Col14=二派费
  Col15=增值税 Col16=偏远费 Col17=小计[公式]
  R2-R3: 手工数据（补退行），R4+: 有公式
"""

import openpyxl
from datetime import datetime
from pathlib import Path
import re


def load_remote_zipcodes(pricing_path):
    """从报价文件加载偏远邮编表 -> {邮编: {country, region}}"""
    wb = openpyxl.load_workbook(pricing_path, data_only=True)
    ws = wb['偏远邮编']
    zips = {}
    for r in range(2, ws.max_row + 1):
        for col_zip, col_region, default_region in [
            (1, 3, '克罗地亚'), (4, 5, '西西里岛'),
            (6, 7, '撒丁岛'), (8, 9, '其他岛屿'),
        ]:
            z = ws.cell(r, col_zip).value
            if z:
                region = str(ws.cell(r, col_region).value or default_region).strip()
                country_val = ws.cell(r, 2).value if col_zip == 1 else '意大利'
                zips[str(z).strip()] = {
                    'country': str(country_val or default_region).strip(),
                    'region': region,
                }
    wb.close()
    return zips


def calc_remote_fee_rmb(region, weight_kg, exchange_rate):
    """根据地区和重量计算偏远费(RMB)
    报价: 卡拉布里亚/西西里岛/撒丁岛 = +0.8 EUR/KG
          其他岛屿(威尼斯等) = +21 EUR/票
          克罗地亚偏远 = 报价表未明确(暂按0.8EUR/KG)
    """
    if region in ('卡拉布里亚', '西西里岛', '撒丁岛', '克罗地亚'):
        fee_eur = 0.8 * (weight_kg or 1)
    elif region == '其他岛屿':
        fee_eur = 21.0
    else:
        fee_eur = 0.8 * (weight_kg or 1)
    return round(fee_eur * exchange_rate, 2)


def parse_input(filepath):
    """解析上游对账单，按运单号建立全量索引"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    parcels = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in ('汇总', '总表'):
            continue
        ws = wb[sheet_name]

        # 找到header行 (通常在第9行)
        header_row = None
        for r in range(1, min(15, ws.max_row + 1)):
            v = ws.cell(r, 3).value
            if v and '运单' in str(v):
                header_row = r
                break
        if not header_row:
            continue

        # 读取header
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                headers[c] = str(v).strip()

        for r in range(header_row + 1, ws.max_row + 1):
            waybill = ws.cell(r, 3).value
            if not waybill or not isinstance(waybill, str):
                continue
            waybill = waybill.strip()
            if not waybill:
                continue

            amount = ws.cell(r, 13).value

            if waybill not in parcels:
                parcels[waybill] = {
                    'date': ws.cell(r, 2).value,
                    'waybill': waybill,
                    'transfer_no': ws.cell(r, 4).value,
                    'route': ws.cell(r, 5).value,
                    'pieces': ws.cell(r, 6).value,
                    'actual_weight': ws.cell(r, 7).value,
                    'charge_weight': ws.cell(r, 8).value,
                    'dimensions': ws.cell(r, 9).value,
                    'dest': ws.cell(r, 11).value,
                    'cust_ref': ws.cell(r, 16).value,
                    'category': ws.cell(r, 17).value,
                    'fees': {},
                    'fee_details': [],
                }

            fee_detail = {
                'sheet': sheet_name,
                'amount_eur': amount,
                'charge_weight': ws.cell(r, 8).value,
            }
            parcels[waybill]['fee_details'].append(fee_detail)

            fee_key = _classify_fee(sheet_name)
            if fee_key:
                parcels[waybill]['fees'][fee_key] = fee_detail
                if fee_key == 'head_freight':
                    parcels[waybill]['head_charge_weight'] = ws.cell(r, 8).value
                if fee_key == 'tail_freight':
                    parcels[waybill]['tail_charge_weight'] = ws.cell(r, 8).value

    wb.close()
    return parcels


def _classify_fee(sheet_name):
    mapping = {
        'COD': 'cod',
        '上架费': 'shelf_fee',
        '尾程退件操作费(补退)': 'tail_return_refund',
        '尾程退件操作费': 'tail_return_fee',
        '尾程退件扣费': 'tail_return_fee',
        '服务费': 'service_fee',
        'DE地派服务费': 'de_delivery_fee',
        '尾程运费': 'tail_freight',
        '目的地增值税': 'vat',
        'GR地派服务费': 'gr_delivery_fee',
        '转寄操作费': 'forward_fee',
        'HR地派服务费': 'hr_delivery_fee',
        'IT地派服务费': 'it_delivery_fee',
        '短信费': 'sms_fee',
        '头程运费': 'head_freight',
    }
    for k, v in mapping.items():
        if k in sheet_name:
            return v
    return None


def determine_ship_type(parcel):
    cust_ref = str(parcel.get('cust_ref', '') or '')
    if re.match(r'^JJD\d+', cust_ref):
        return '转寄'
    if re.match(r'^[A-Z]{1,4}-\d{2}-\d{1,2}-.+$', cust_ref):
        return '直发'
    if re.match(r'^BB-\d{2}-.+$', cust_ref):
        return '直发'
    transfer = str(parcel.get('transfer_no', '') or '')
    if transfer.startswith('JJD'):
        return '转寄'
    return '直发'


DELIVERY_PRICING = {
    'new': {
        '波兰':     {'first2kg': 4.0, 'extra1kg': 0.9, 'rule': '70%'},
        '罗马尼亚':  {'first2kg': 4.8, 'extra1kg': 0.6, 'rule': '70%'},
        '匈牙利':   {'first2kg': 4.4, 'extra1kg': 0.7, 'rule': '70%'},
        '捷克':     {'first2kg': 4.1, 'extra1kg': 0.8, 'rule': '70%'},
        '斯洛伐克':  {'first2kg': 4.3, 'extra1kg': 0.7, 'rule': '70%'},
        '保加利亚':  {'first2kg': 4.3, 'extra1kg': 0.8, 'rule': '70%'},
        '克罗地亚':  {'first2kg': 6.4, 'extra1kg': 1.0, 'rule': '70%'},
        '斯洛文尼亚': {'first2kg': 5.9, 'extra1kg': 0.9, 'rule': '70%'},
        '西班牙':   {'first2kg': 4.0, 'extra1kg': 1.0, 'rule': '同派送费'},
        '葡萄牙':   {'first2kg': 4.0, 'extra1kg': 1.0, 'rule': '同派送费'},
        '希腊':     {'first2kg': 5.7, 'extra1kg': 0.8, 'rule': '同派送费'},
        '意大利':   {'first2kg': 7.3, 'extra1kg': 1.1, 'rule': '同派送费'},
        '奥地利':   {'first2kg': 6.5, 'extra1kg': 1.0, 'rule': '同派送费'},
        '德国':     {'first2kg': 8.3, 'extra1kg': 1.5, 'rule': '同派送费'},
    },
    'old': {
        '德国':     {'first2kg': 8.0, 'extra1kg': 1.5, 'rule': '同派送费'},
        '意大利':   {'first2kg': 6.7, 'extra1kg': 1.0, 'rule': '同派送费'},
        '克罗地亚':  {'first2kg': 5.6, 'extra1kg': 0.9, 'rule': '70%'},
    },
}


def _get_delivery_params(dest, pricing_version='new'):
    """获取指定国家的派送费参数 (first2kg, extra1kg, is_70pct)"""
    tables = DELIVERY_PRICING.get(pricing_version, DELIVERY_PRICING['new'])
    fallback = DELIVERY_PRICING['new']
    info = tables.get(dest) or fallback.get(dest)
    if not info:
        info = fallback.get('德国')
    return info['first2kg'], info['extra1kg'], info['rule'] == '70%'


def _calc_return_fee_rmb(dest, exchange_rate, pricing_version='new'):
    """根据目的国计算尾程退件入仓费(RMB) — 仅首2KG基础费，用于退回补退等固定值场景"""
    base, _, is_70 = _get_delivery_params(dest, pricing_version)
    fee = base * 0.7 if is_70 else base
    return round(fee * exchange_rate, 2)


def _build_return_fee_formula(dest, row, exchange_rate, pricing_version='new'):
    """根据目的地生成 C13 退件费公式，含重量加价逻辑。
    "同派送费"国家: =ROUND(IF(G>2, base+(ROUNDUP(G,0)-2)*step, base)*rate, 2)
    "70%"国家:     =ROUND(IF(G>2, base+(ROUNDUP(G,0)-2)*step, base)*rate*70%, 2)
    """
    base, step, is_70 = _get_delivery_params(dest, pricing_version)
    g = f'G{row}'
    pct = '*70%' if is_70 else ''
    return (
        f'=ROUND(IF({g}>2,{base}+(ROUNDUP({g},0)-2)*{step},{base})'
        f'*{exchange_rate}{pct},2)'
    )


def is_formula(v):
    return isinstance(v, str) and v.startswith('=')


def clear_data_rows(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and not is_formula(v):
                ws.cell(r, c).value = None


def ensure_formulas(ws, formula_source_row, target_rows, formula_cols):
    """把 formula_source_row 的公式复制到 target_rows"""
    for col in formula_cols:
        src_val = ws.cell(formula_source_row, col).value
        if not is_formula(src_val):
            continue
        for tr in target_rows:
            existing = ws.cell(tr, col).value
            if not is_formula(existing):
                new_formula = re.sub(
                    r'([A-Z]+)' + str(formula_source_row),
                    lambda m: m.group(1) + str(tr),
                    src_val
                )
                ws.cell(tr, col, new_formula)


def fill_template(template_path, input_data, output_path, customer_code, exchange_rate,
                  remote_zips=None, parcel_zipcodes=None, pricing_version='new'):
    wb = openpyxl.load_workbook(template_path)

    # ── Sheet 0: COD回款 ──
    ws_cod = wb.worksheets[0]
    cod_data_end = ws_cod.max_row
    for r in range(2, ws_cod.max_row + 1):
        all_none = all(ws_cod.cell(r, c).value is None for c in range(1, 14))
        if all_none:
            cod_data_end = r - 1
            break

    clear_data_rows(ws_cod, 2, cod_data_end)

    cod_parcels = [p for p in input_data.values() if 'cod' in p['fees']]
    cod_parcels.sort(key=lambda p: p.get('date') or datetime.min)

    for i, p in enumerate(cod_parcels):
        row = 2 + i
        ws_cod.cell(row, 1, customer_code)
        ws_cod.cell(row, 2, determine_ship_type(p))
        ws_cod.cell(row, 3, p['date'])
        ws_cod.cell(row, 4, p['waybill'])
        ws_cod.cell(row, 5, p.get('cust_ref', ''))
        ws_cod.cell(row, 6, p.get('transfer_no', ''))
        ws_cod.cell(row, 7, p.get('dest', ''))
        ws_cod.cell(row, 8, p.get('category', ''))
        ws_cod.cell(row, 9, p['fees']['cod']['amount_eur'])
        ws_cod.cell(row, 10, exchange_rate)

    ensure_formulas(ws_cod, 4, [2, 3], [11, 12, 13])
    print(f"  COD: {len(cod_parcels)} rows filled (row 2-{1+len(cod_parcels)})")

    # ── Sheet 1: 运费 ──
    ws_freight = wb.worksheets[1]
    freight_data_end = ws_freight.max_row
    for r in range(2, ws_freight.max_row + 1):
        all_none = all(ws_freight.cell(r, c).value is None for c in range(1, 14))
        if all_none:
            freight_data_end = r - 1
            break

    clear_data_rows(ws_freight, 2, freight_data_end)

    freight_parcels = [p for p in input_data.values()
                       if 'head_freight' in p['fees'] or 'tail_freight' in p['fees']]
    freight_parcels.sort(key=lambda p: p.get('date') or datetime.min)

    for i, p in enumerate(freight_parcels):
        row = 2 + i
        head_w = p.get('head_charge_weight') or 0
        tail_q = p.get('tail_charge_weight') or p.get('charge_weight') or 0

        ws_freight.cell(row, 1, customer_code)
        ws_freight.cell(row, 2, determine_ship_type(p))
        ws_freight.cell(row, 3, p['date'])
        ws_freight.cell(row, 4, p['waybill'])
        ws_freight.cell(row, 5, p.get('cust_ref', ''))
        ws_freight.cell(row, 6, p.get('transfer_no', ''))
        ws_freight.cell(row, 7, p.get('dest', ''))
        ws_freight.cell(row, 8, p.get('category', ''))
        ws_freight.cell(row, 9, p.get('actual_weight'))
        ws_freight.cell(row, 10, head_w)
        ws_freight.cell(row, 11, p.get('dimensions', ''))
        ws_freight.cell(row, 12, tail_q)
        ws_freight.cell(row, 13, 'IC')

    print(f"  Freight: {len(freight_parcels)} rows filled (row 2-{1+len(freight_parcels)})")

    # ── Sheet 2: 尾程杂费 ──
    # 列结构: C1=寄件日期 C2=运单号码 C3=客户单号 C4=转单号 C5=品名
    #         C6=件数 C7=重量(KG) C8=尺寸 C9=类型 C10=目的地 C11=科目
    #         C12=上架费(RMB) C13=尾程退件入仓费(RMB) C14=二派费
    #         C15=增值税 C16=偏远费 C17=小计[公式=SUM(L:P)] C18=备注
    #         C19=邮编 C20=岛屿/地区名
    # 规则: 同一运单合并到一行，不同费用填到不同列
    ws_sur = wb.worksheets[2]
    sur_data_end = ws_sur.max_row
    for r in range(2, ws_sur.max_row + 1):
        all_none = all(ws_sur.cell(r, c).value is None for c in range(1, 11))
        if all_none:
            sur_data_end = r - 1
            break

    clear_data_rows(ws_sur, 2, sur_data_end)

    surcharge_fee_keys = {
        'shelf_fee', 'tail_return_fee', 'tail_return_refund', 'vat',
    }

    # 先判断哪些运单有偏远费（通过 parcel_zipcodes + remote_zips）
    remote_parcels = set()
    if remote_zips and parcel_zipcodes:
        for wb_id, zipcode in parcel_zipcodes.items():
            if str(zipcode).strip() in remote_zips:
                remote_parcels.add(wb_id)

    merged = {}
    for p in input_data.values():
        wb_id = p['waybill']
        has_surcharge = any(ft in p['fees'] for ft in surcharge_fee_keys)
        has_remote = wb_id in remote_parcels

        if not has_surcharge and not has_remote:
            continue

        if wb_id not in merged:
            merged[wb_id] = {
                'date': p['date'],
                'waybill': wb_id,
                'cust_ref': p.get('cust_ref', ''),
                'transfer_no': p.get('transfer_no', ''),
                'category': p.get('category', ''),
                'pieces': p.get('pieces'),
                'charge_weight': p.get('charge_weight'),
                'dimensions': p.get('dimensions', ''),
                'dest': p.get('dest', ''),
                'remark_parts': [],
            }

        rec = merged[wb_id]
        if 'shelf_fee' in p['fees'] and '上架费' not in rec['remark_parts']:
            rec['remark_parts'].append('上架费')
        if 'tail_return_fee' in p['fees'] and '退件操作费' not in rec['remark_parts']:
            rec['remark_parts'].append('退件操作费')
        if 'tail_return_refund' in p['fees'] and '签收件，拒收返程费补退' not in rec['remark_parts']:
            rec['remark_parts'].append('签收件，拒收返程费补退')
        if 'vat' in p['fees'] and '增值税' not in rec['remark_parts']:
            rec['remark_parts'].append('增值税')
        if has_remote and '偏远费' not in rec['remark_parts']:
            rec['remark_parts'].append('偏远费')

    # C13 退件费: 不是简单 EUR*汇率, 而是基于报价表的尾程派送费*70% 等
    # 模板中的值: DE=63.47, IT=53.16, HR=31.10, 补退=-65.23
    # 暂时保留公式列(C12)自动计算上架费，手动填入退件费和增值税的实际RMB值
    sur_list = sorted(merged.values(), key=lambda x: x.get('date') or datetime.min)

    for i, item in enumerate(sur_list):
        row = 2 + i
        ws_sur.cell(row, 1, item['date'])
        ws_sur.cell(row, 2, item['waybill'])
        ws_sur.cell(row, 3, item['cust_ref'])
        ws_sur.cell(row, 4, item['transfer_no'])
        ws_sur.cell(row, 5, item['category'])
        ws_sur.cell(row, 6, item['pieces'])
        ws_sur.cell(row, 7, item['charge_weight'])
        ws_sur.cell(row, 8, item['dimensions'])
        ws_sur.cell(row, 10, item['dest'])

        wb_id = item['waybill']
        p = input_data.get(wb_id, {})
        fees = p.get('fees', {})

        # C12 上架费: 只对输入文件有上架费记录的运单写公式
        if 'shelf_fee' in fees:
            ws_sur.cell(row, 12, f'=ROUND(1.5*{exchange_rate},2)')

        # C13 尾程退件入仓费: 根据目的地生成对应公式
        if 'tail_return_fee' in fees:
            dest = str(item['dest'] or '')
            c13_formula = _build_return_fee_formula(dest, row, exchange_rate, pricing_version)
            ws_sur.cell(row, 13, c13_formula)
        if 'tail_return_refund' in fees:
            dest = str(item['dest'] or '')
            ret_rmb = _calc_return_fee_rmb(dest, exchange_rate, pricing_version)
            ws_sur.cell(row, 13, -abs(ret_rmb))

        # C15 增值税: 用公式而非固定值
        if 'vat' in fees:
            vat_eur = fees['vat'].get('amount_eur', 0) or 0
            if 'tail_return_refund' in fees:
                vat_rmb = round(vat_eur * exchange_rate, 2)
                ws_sur.cell(row, 15, -abs(vat_rmb))
            else:
                ws_sur.cell(row, 15, f'=ROUNDUP({vat_eur}*{exchange_rate},2)')

        # C16 偏远费 + C19 邮编 + C20 地区
        if remote_zips and parcel_zipcodes:
            wb_id = item['waybill']
            zipcode = parcel_zipcodes.get(wb_id)
            if zipcode and str(zipcode).strip() in remote_zips:
                zip_info = remote_zips[str(zipcode).strip()]
                weight = item.get('charge_weight') or 1
                remote_rmb = calc_remote_fee_rmb(zip_info['region'], weight, exchange_rate)
                ws_sur.cell(row, 16, remote_rmb)
                ws_sur.cell(row, 19, str(zipcode).strip())
                ws_sur.cell(row, 20, zip_info['region'])

        # C18 备注
        remark_parts = item['remark_parts']
        if remark_parts:
            if '签收件，拒收返程费补退' in remark_parts:
                ws_sur.cell(row, 18, '签收件，拒收返程费补退')
            else:
                label_map = {
                    '上架费': '上架费',
                    '退件操作费': '退件操作费',
                    '增值税': '增值税',
                    '偏远费': '偏远费',
                }
                if len(remark_parts) == 1:
                    single = remark_parts[0]
                    display_map = {
                        '上架费': '上架费',
                        '退件操作费': '尾程退件操作费',
                        '增值税': '目的地增值税',
                        '偏远费': '偏远费',
                    }
                    ws_sur.cell(row, 18, display_map.get(single, single))
                else:
                    ws_sur.cell(row, 18, '+'.join(
                        label_map.get(p, p) for p in remark_parts))

    # C17(小计=SUM) 需要扩展到所有数据行
    all_data_rows = list(range(2, 2 + len(sur_list)))
    ensure_formulas(ws_sur, 4, all_data_rows, [17])
    print(f"  Surcharge: {len(sur_list)} rows filled (row 2-{1+len(sur_list)})")

    wb.save(output_path)
    wb.close()
    print(f"\n  Saved: {output_path}")


def extract_zipcodes_from_template(template_path):
    """从模板文件的尾程杂费表提取已知的运单->邮编映射（作为参考数据）"""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.worksheets[2]
    zipcodes = {}
    for r in range(2, ws.max_row + 1):
        waybill = ws.cell(r, 2).value
        zipcode = ws.cell(r, 19).value
        if waybill and zipcode:
            zipcodes[str(waybill)] = str(zipcode)
    wb.close()
    return zipcodes


if __name__ == '__main__':
    TEMPLATE_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
    INPUT_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
    PRICING_FILE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\汇森国际-东欧COD报价20260331生效(5).xlsx'
    EXCHANGE_RATE = 7.9342
    CUSTOMER_CODE = 46111

    OUTPUT_DIR = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站'
    OUTPUT_FILE = str(Path(OUTPUT_DIR) / '20260330-汇森李志（东欧）对账单-自动生成v5.xlsx')
    # 3月30日的账单在新报价(3月31日生效)之前，使用旧报价参数
    # 旧报价仅 DE/IT/HR 与新报价不同(首2KG/续重)，其余国家参数不变
    PRICING_VERSION = 'old'

    print("Loading remote zipcodes from pricing...")
    remote_zips = load_remote_zipcodes(PRICING_FILE)
    print(f"  Loaded {len(remote_zips)} remote zipcodes")

    print("Extracting known zipcodes from template...")
    parcel_zipcodes = extract_zipcodes_from_template(TEMPLATE_FILE)
    print(f"  Found {len(parcel_zipcodes)} parcel-zipcode mappings")

    print("\nParsing input...")
    parcels = parse_input(INPUT_FILE)
    print(f"Found {len(parcels)} waybills\n")

    print(f"Filling template (pricing: {PRICING_VERSION})...")
    fill_template(TEMPLATE_FILE, parcels, OUTPUT_FILE, CUSTOMER_CODE, EXCHANGE_RATE,
                  remote_zips=remote_zips, parcel_zipcodes=parcel_zipcodes,
                  pricing_version=PRICING_VERSION)
    print("\nDone!")
