"""
Compare tail freight formulas in all customer templates against the OLD pricing file
(汇森国际-东欧COD报价20251108生效(13).xlsx), which was effective before 2026-03-31.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMERS = {
    '李志': '20260330-汇森李志（东欧）对账单.xlsx',
    '君悦': '汇森-君悦（东欧）对账单-20260330.xlsx',
    'J':    '汇森-J（东欧）对账单-20260330.xlsx',
    '涵江': '汇森-涵江（东欧）对账单-20260330.xlsx',
    '阿甘': '汇森-阿甘（东欧）对账单-20260330.xlsx',
    '威总': '汇森-威总（东欧）对账单-20260330.xlsx',
    '峰总': '汇森-峰总（东欧）对账单-20260330.xlsx',
    '小美': '汇森-小美（东欧）对账单-20260330.xlsx',
}

OLD_PRICING = {
    '波兰':      {'first2': 3.8, 'extra1': 0.8, 'service': 'INPOST'},
    '罗马尼亚':   {'first2': 4.8, 'extra1': 0.6, 'service': 'SAMEDAY'},
    '匈牙利':    {'first2': 4.1, 'extra1': 0.6, 'service': 'Express One'},
    '捷克':      {'first2': 3.8, 'extra1': 0.6, 'service': 'WEDO'},
    '斯洛伐克':   {'first2': 4.0, 'extra1': 0.6, 'service': 'POST'},
    '保加利亚':   {'first2': 4.3, 'extra1': 0.8, 'service': 'Sameday'},
    '克罗地亚':   {'first2': 5.6, 'extra1': 0.9, 'service': 'GLS'},
    '斯洛文尼亚':  {'first2': 5.2, 'extra1': 0.8, 'service': 'GLS'},
    '西班牙':    {'first2': 4.0, 'extra1': 1.0, 'service': 'CTT'},
    '葡萄牙':    {'first2': 4.0, 'extra1': 1.0, 'service': 'CTT'},
    '希腊':      {'first2': 5.7, 'extra1': 0.8, 'service': 'Geniki'},
    '意大利':    {'first2': 6.7, 'extra1': 1.0, 'service': 'GLS'},
    '奥地利':    {'first2': 6.5, 'extra1': 1.0, 'service': 'DPD'},
    '德国':      {'first2': 8.0, 'extra1': 1.5, 'service': 'DHL'},
}

NEW_PRICING = {
    '波兰':      {'first2': 4.0, 'extra1': 0.9},
    '罗马尼亚':   {'first2': 4.8, 'extra1': 0.6},
    '匈牙利':    {'first2': 4.4, 'extra1': 0.7},
    '捷克':      {'first2': 4.1, 'extra1': 0.8},
    '斯洛伐克':   {'first2': 4.3, 'extra1': 0.7},
    '保加利亚':   {'first2': 4.3, 'extra1': 0.8},
    '克罗地亚':   {'first2': 6.4, 'extra1': 1.0},
    '斯洛文尼亚':  {'first2': 5.9, 'extra1': 0.9},
    '西班牙':    {'first2': 4.0, 'extra1': 1.0},
    '葡萄牙':    {'first2': 4.0, 'extra1': 1.0},
    '希腊':      {'first2': 5.7, 'extra1': 0.8},
    '意大利':    {'first2': 7.3, 'extra1': 1.1},
    '奥地利':    {'first2': 6.5, 'extra1': 1.0},
    '德国':      {'first2': 8.3, 'extra1': 1.5},
}


def extract_tail_params(formula):
    if not formula or not isinstance(formula, str):
        return None, None, None
    m = re.search(r'(\d+\.?\d*)\s*\+\s*\(.*?\)\s*\*\s*(\d+\.?\d*)\s*,\s*(\d+\.?\d*)', formula)
    if m:
        return float(m.group(1)), float(m.group(2)), float(m.group(3))
    return None, None, None


print("=" * 120)
print("尾程运费公式 vs 旧报价(20251108生效) vs 新报价(20260331生效) 逐国家对比")
print("=" * 120)

all_results = {}

for customer, fname in CUSTOMERS.items():
    fpath = TEMPLATE_DIR / fname
    wb_f = openpyxl.load_workbook(str(fpath), data_only=False)
    wb_v = openpyxl.load_workbook(str(fpath), data_only=True)

    ws_f = ws_v = None
    for sname in wb_f.sheetnames:
        if '运费' in sname and '杂费' not in sname:
            ws_f = wb_f[sname]
            ws_v = wb_v[sname]
            break
    if not ws_f:
        wb_f.close()
        wb_v.close()
        continue

    tail_col = dest_col = None
    for c in range(1, ws_f.max_column + 1):
        h = str(ws_f.cell(1, c).value or '')
        if '尾程' in h and '运费' in h:
            tail_col = c
        if '目的' in h:
            dest_col = c

    country_formulas = {}
    for r in range(2, ws_f.max_row + 1):
        dest = ws_v.cell(r, dest_col).value if dest_col else None
        formula = ws_f.cell(r, tail_col).value
        if dest is None:
            break
        d = str(dest).strip()
        if '合计' in d:
            break
        if d not in country_formulas:
            country_formulas[d] = formula

    for country, formula in country_formulas.items():
        f1, e1, _ = extract_tail_params(str(formula))
        if f1 is None:
            continue

        old_p = OLD_PRICING.get(country)
        new_p = NEW_PRICING.get(country)

        old_match = old_p and f1 == old_p['first2'] and e1 == old_p['extra1']
        new_match = new_p and f1 == new_p['first2'] and e1 == new_p['extra1']

        has_extra = isinstance(formula, str) and formula.count('+') > 1

        key = (country, f1, e1)
        if key not in all_results:
            all_results[key] = {
                'customers': [],
                'old_match': old_match,
                'new_match': new_match,
                'old_p': old_p,
                'new_p': new_p,
                'has_extra': has_extra,
                'formula_sample': formula,
            }
        all_results[key]['customers'].append(customer)

    wb_f.close()
    wb_v.close()


print(f"\n{'国家':<10} {'模板首2':>6} {'模板续1':>6} │ {'旧报价首2':>8} {'旧报价续1':>8} {'旧匹配':>6} │ {'新报价首2':>8} {'新报价续1':>8} {'新匹配':>6} │ 客户列表")
print("─" * 120)

for (country, f1, e1), info in sorted(all_results.items()):
    old_p = info['old_p']
    new_p = info['new_p']
    old_str = f"{old_p['first2']:>8} {old_p['extra1']:>8}" if old_p else f"{'N/A':>8} {'N/A':>8}"
    new_str = f"{new_p['first2']:>8} {new_p['extra1']:>8}" if new_p else f"{'N/A':>8} {'N/A':>8}"

    old_mark = '  ✓' if info['old_match'] else '  ✗'
    new_mark = '  ✓' if info['new_match'] else '  ✗'

    extra = ' [+1.5EUR]' if info['has_extra'] else ''
    customers_str = ', '.join(info['customers'])
    print(f"{country:<10} {f1:>6} {e1:>6} │ {old_str} {old_mark:>6} │ {new_str} {new_mark:>6} │ {customers_str}{extra}")


print(f"\n\n{'='*120}")
print("汇总统计")
print("=" * 120)

old_match_count = sum(1 for v in all_results.values() if v['old_match'])
new_match_count = sum(1 for v in all_results.values() if v['new_match'])
total = len(all_results)

print(f"  共 {total} 个 (国家, 费率) 组合")
print(f"  匹配旧报价(20251108): {old_match_count}/{total} = {old_match_count/total*100:.0f}%")
print(f"  匹配新报价(20260331): {new_match_count}/{total} = {new_match_count/total*100:.0f}%")


print(f"\n\n{'='*120}")
print("新旧报价差异 — 哪些国家调价了")
print("=" * 120)
print(f"{'国家':<12} {'旧首2':>6} {'旧续1':>6} → {'新首2':>6} {'新续1':>6}  变化")
print("─" * 80)
for country in sorted(OLD_PRICING.keys()):
    old = OLD_PRICING[country]
    new = NEW_PRICING.get(country)
    if not new:
        continue
    changed = old['first2'] != new['first2'] or old['extra1'] != new['extra1']
    mark = '  ⬆ 涨价' if changed else '  = 不变'
    d1 = f"+{new['first2']-old['first2']:.1f}" if new['first2'] != old['first2'] else '  -'
    d2 = f"+{new['extra1']-old['extra1']:.1f}" if new['extra1'] != old['extra1'] else '  -'
    print(f"{country:<12} {old['first2']:>6} {old['extra1']:>6} → {new['first2']:>6} {new['extra1']:>6}  {mark}  (首2{d1}, 续1{d2})")
