"""检查模板中"账期"列的值和日期格式"""
import openpyxl

TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'
wb = openpyxl.load_workbook(TEMPLATE)

# COD sheet - check col A (账期)
ws = wb[wb.sheetnames[0]]
print("=== COD Sheet Col A (账期) ===")
vals = set()
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 1).value
    nf = ws.cell(r, 1).number_format
    if v is not None:
        vals.add((repr(v), nf))
for v, nf in vals:
    print(f"  value={v}, format={nf}")

# Check col C date format
print("\n=== COD Sheet Col C (发货日期) format ===")
for r in [2, 3, 4, 5]:
    cell = ws.cell(r, 3)
    print(f"  Row {r}: value={repr(cell.value)}, format='{cell.number_format}'")

# Freight sheet col A
ws2 = wb[wb.sheetnames[1]]
print("\n=== Freight Sheet Col A (账期) ===")
vals2 = set()
for r in range(2, 10):
    v = ws2.cell(r, 1).value
    nf = ws2.cell(r, 1).number_format
    print(f"  Row {r}: value={repr(v)}, format='{nf}'")

# Check what the actual "账期" value looks like
# From screenshot: "3月30日" - is it a date or a string?
print("\n=== COD first 5 rows full dump ===")
for r in range(2, 7):
    print(f"  Row {r}:")
    for c in range(1, 15):
        v = ws.cell(r, c).value
        nf = ws.cell(r, c).number_format
        if v is not None:
            print(f"    Col {c}: value={repr(v)}, type={type(v).__name__}, fmt='{nf}'")
