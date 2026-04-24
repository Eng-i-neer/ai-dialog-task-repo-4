"""
分析科目映射：
1. 输入文件（代理）的所有科目（Sheet名）
2. 每个客户模板的所有科目（Sheet + 列）
3. 建立科目级别的映射关系
"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(r'c:\Users\59571\Desktop\deutsch-app\舅妈网站')
INPUT_DIR = BASE / '中介提供'
TEMPLATE_DIR = BASE / '反馈客户' / '原始模板'

CUSTOMER_MAP = {
    '中文':   '李志',
    '中文1':  '君悦',
    '中文3':  '小美',
    '中文5':  'J',
    '中文6':  '涵江',
    '中文7':  '阿甘',
    '中文9':  '威总',
    '中文12': '峰总',
}


def find_file(directory, keyword):
    for f in os.listdir(directory):
        if f.endswith('.xlsx') and keyword in f:
            return directory / f
    return None


print("=" * 80)
print("PART 1: 代理输入文件的科目（Sheet）清单")
print("=" * 80)

# Analyze each input file's sheets and their data
input_categories = {}
for code, customer in CUSTOMER_MAP.items():
    input_path = find_file(INPUT_DIR, f'-{code}-')
    if not input_path:
        continue

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    print(f"\n--- {code} -> {customer} ({input_path.name}) ---")

    sheets_info = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if sname in ('汇总', '总表'):
            print(f"  [{sname}] (skip - summary)")
            continue

        # Find header row
        header_row = None
        for r in range(1, min(15, ws.max_row + 1)):
            v = ws.cell(r, 3).value
            if v and '运单' in str(v):
                header_row = r
                break

        if not header_row:
            print(f"  [{sname}] (no header found)")
            continue

        # Count data rows
        data_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            waybill = ws.cell(r, 3).value
            if waybill and isinstance(waybill, str) and waybill.strip():
                data_count += 1

        # Sample amount column (C13)
        sample_amt = None
        for r in range(header_row + 1, min(header_row + 3, ws.max_row + 1)):
            v = ws.cell(r, 13).value
            if v is not None:
                sample_amt = v
                break

        print(f"  [{sname}] {data_count} rows, sample_amt={sample_amt}")
        sheets_info.append({
            'name': sname,
            'count': data_count,
        })

        if code not in input_categories:
            input_categories[code] = {}
        input_categories[code][sname] = data_count

    wb.close()

# Aggregate all unique input categories
all_input_cats = set()
for code, cats in input_categories.items():
    all_input_cats.update(cats.keys())

print(f"\n\n{'='*80}")
print("PART 2: 代理侧全部科目汇总")
print("=" * 80)
for cat in sorted(all_input_cats):
    clients_with = []
    for code, cats in input_categories.items():
        if cat in cats:
            clients_with.append(f"{code}({cats[cat]})")
    print(f"  {cat}: {', '.join(clients_with)}")


print(f"\n\n{'='*80}")
print("PART 3: 客户模板的科目（Sheet + 费用列）")
print("=" * 80)

template_categories = {}
for code, customer in CUSTOMER_MAP.items():
    template_path = find_file(TEMPLATE_DIR, customer)
    if not template_path:
        continue

    wb = openpyxl.load_workbook(str(template_path), data_only=False)
    print(f"\n--- {customer} ({template_path.name}) ---")

    cust_cats = []
    for si, sname in enumerate(wb.sheetnames):
        ws = wb[sname]
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h:
                headers[c] = str(h).strip()

        # Find fee-related columns (typically numeric/formula columns after position data)
        fee_cols = []
        for c, h in headers.items():
            if any(kw in h for kw in ['运费', '手续费', '附加费', '上架', '退件', '返程',
                                       '增值税', '清关', '偏远', '二派', '小计', '金额',
                                       '赔付', '理赔', '退运费', '仓储', '科目']):
                is_formula_col = False
                for r in range(2, min(ws.max_row + 1, 6)):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith('='):
                        is_formula_col = True
                        break
                fee_cols.append({
                    'col': c,
                    'header': h,
                    'is_formula': is_formula_col,
                })

        # Count data rows
        data_count = 0
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 8))]
            if all(v is None for v in vals):
                break
            if any('合计' in str(v or '') for v in vals):
                break
            data_count += 1

        fee_desc = ', '.join(f"C{fc['col']}={fc['header']}{'[F]' if fc['is_formula'] else ''}"
                             for fc in fee_cols)
        print(f"  [{si}] {sname}: {data_count} rows | {fee_desc}")

        cust_cats.append({
            'sheet_name': sname,
            'sheet_idx': si,
            'data_rows': data_count,
            'fee_cols': fee_cols,
        })

    template_categories[customer] = cust_cats
    wb.close()


print(f"\n\n{'='*80}")
print("PART 4: 科目映射矩阵 — 代理科目 -> 客户模板目标位置")
print("=" * 80)

# Build the mapping
input_cat_list = sorted(all_input_cats)
print(f"\n代理侧科目({len(input_cat_list)}个):")
for i, cat in enumerate(input_cat_list):
    print(f"  [{i+1}] {cat}")

print(f"\n逐客户映射:")
for code, customer in CUSTOMER_MAP.items():
    if customer not in template_categories:
        continue
    print(f"\n  === {customer} ===")
    tcats = template_categories[customer]

    for icat in input_cat_list:
        targets = []
        icat_lower = icat.lower()

        for tcat in tcats:
            sname = tcat['sheet_name']
            # Direct sheet-level matches
            if 'COD' in icat and 'COD' in sname:
                targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> 数据行")
            elif '头程运费' in icat and '运费' in sname:
                targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> 头程运费列")
            elif '尾程运费' in icat and '运费' in sname:
                targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> 尾程运费列")
            elif '上架费' in icat:
                for fc in tcat['fee_cols']:
                    if '上架' in fc['header']:
                        targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> C{fc['col']} {fc['header']}")
            elif '退件' in icat and '操作费' in icat:
                for fc in tcat['fee_cols']:
                    if '退件' in fc['header'] or '返程' in fc['header'] or '拒收' in fc['header']:
                        targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> C{fc['col']} {fc['header']}")
            elif '增值税' in icat and '目的地增值税' in icat:
                # Check VAT sheet
                if '增值税' in sname:
                    targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> 增值税列")
                # Also check surcharge sheet VAT column
                for fc in tcat['fee_cols']:
                    if '增值税' in fc['header'] and '杂费' in sname:
                        targets.append(f"Sheet[{tcat['sheet_idx']}] {sname} -> C{fc['col']} {fc['header']}")
            elif '服务费' in icat or '地派' in icat:
                # Service fees / delivery fees → may map to freight surcharge or separate
                pass
            elif '转寄' in icat:
                pass
            elif '短信' in icat:
                pass

        if targets:
            print(f"    {icat} -> {'; '.join(targets)}")
        else:
            print(f"    {icat} -> ??? (无直接映射)")
