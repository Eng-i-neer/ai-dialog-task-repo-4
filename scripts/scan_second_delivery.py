# -*- coding: utf-8 -*-
"""Scan all 李志 agent bills for 二派费 related sheets and column structures."""
import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

base = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单'

keywords_2nd = ['二派', '二次派送', '2nd', 'second']

for root, dirs, files in os.walk(base):
    for f in files:
        if '李志' not in f or not f.endswith('.xlsx'):
            continue
        fp = os.path.join(root, f)
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        except Exception as e:
            continue

        for sname in wb.sheetnames:
            sname_lower = sname.lower()
            if any(k in sname_lower or k in sname for k in keywords_2nd):
                ws = wb[sname]
                print(f"\n{'='*60}")
                print(f"文件: {os.path.basename(fp)}")
                print(f"Sheet: {sname}")

                for row_idx in range(1, min(15, (ws.max_row or 15) + 1)):
                    vals = []
                    for col_idx in range(1, min(20, (ws.max_column or 20) + 1)):
                        v = ws.cell(row_idx, col_idx).value
                        vals.append(str(v)[:30] if v else '')
                    line = ' | '.join(vals)
                    if any(v.strip() for v in vals):
                        print(f"  R{row_idx}: {line}")

                print(f"  总行数: {ws.max_row}")
        wb.close()
