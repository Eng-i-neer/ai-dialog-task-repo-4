# -*- coding: utf-8 -*-
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'c:\Users\59571\Desktop\deutsch-app\舅妈网站\中介提供\中邮原账单'

files = [
    os.path.join(BASE, '2025.08.05', '（李志）鑫腾跃-中文-对账单20250805.xlsx'),
    os.path.join(BASE, '2025.07.28', '（李志）鑫腾跃-中文-对账单20250728.xlsx'),
    os.path.join(BASE, '2025.08.11', '（李志）鑫腾跃-中文-对账单20250811.xlsx'),
    os.path.join(BASE, '2025.08.18', '（李志）鑫腾跃-中文-对账单20250818.xlsx'),
]

for f in files:
    if not os.path.exists(f):
        print(f"NOT FOUND: {f}")
        continue
    period = os.path.basename(f).split('对账单')[1].replace('.xlsx','')
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    print(f"\n{'='*80}")
    print(f"文件: {os.path.basename(f)} (期次: {period})")
    print(f"{'='*80}")
    
    for sname in wb.sheetnames[:5]:
        ws = wb[sname]
        # Find header row
        header_row = None
        for r in range(1, 12):
            for c in range(1, 25):
                v = ws.cell(r, c).value
                if v and ('单号' in str(v) or '运单' in str(v)):
                    header_row = r
                    break
            if header_row:
                break
        
        if not header_row:
            print(f"\n  Sheet [{sname}]: 未找到表头行")
            continue
            
        print(f"\n  Sheet [{sname}] 表头行={header_row}:")
        headers = []
        for c in range(1, 25):
            v = ws.cell(header_row, c).value
            if v:
                headers.append((c, str(v).strip()))
        for col, h in headers:
            print(f"    Col {col:2d}: {h}")
        
        # Show first data row
        print(f"  首行数据 (row {header_row+1}):")
        for col, h in headers:
            v = ws.cell(header_row+1, col).value
            print(f"    Col {col:2d} ({h}): {v}")
    
    wb.close()
