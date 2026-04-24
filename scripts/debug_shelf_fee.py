"""调查3条上架费缺失的具体原因"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

INPUT = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\鑫腾跃 -中文-对账单20260330.xlsx'
TEMPLATE = r'C:\Users\59571\Documents\WeChat Files\wxid_eki7t6sbsp8o12\FileStorage\File\2026-04\20260330-汇森李志（东欧）对账单.xlsx'

inp_wb = openpyxl.load_workbook(INPUT, data_only=True)

# 3条问题运单
targets = ['DE12603061510015', 'HR12603041510002', 'DE12602061410029']

# 1. 在输入文件每个sheet中搜索这3个运单
print("="*80)
print("在输入文件中搜索问题运单")
print("="*80)
for sname in inp_wb.sheetnames:
    ws = inp_wb[sname]
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v and str(v).strip() in targets:
                wb_val = str(v).strip()
                all_vals = []
                for cc in range(1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if vv is not None:
                        all_vals.append(f"C{cc}={vv}")
                print(f"\n  Sheet '{sname}' R{r}: {wb_val}")
                print(f"    {all_vals}")

# 2. 看模板中这3个运单的数据
print("\n" + "="*80)
print("在模板尾程杂费表中的数据")
print("="*80)
tmpl_wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
ws_sur = tmpl_wb['20260330期尾程杂费']
for r in range(2, ws_sur.max_row + 1):
    wb = str(ws_sur.cell(r, 2).value or '')
    if wb in targets:
        vals = {}
        for c in range(1, 21):
            v = ws_sur.cell(r, c).value
            if v is not None:
                hdr = ws_sur.cell(1, c).value or f'C{c}'
                vals[hdr] = v
        print(f"\n  模板 R{r} {wb}: {vals}")

# 3. 用 parse_input 解析看结果
print("\n" + "="*80)
print("parse_input 解析结果")
print("="*80)
sys.path.insert(0, 'scripts')
from convert_bill import parse_input
parcels = parse_input(INPUT)
for wb in targets:
    p = parcels.get(wb)
    if p:
        print(f"\n  {wb}:")
        print(f"    fees keys: {list(p['fees'].keys())}")
        for fk, fv in p['fees'].items():
            print(f"    fee '{fk}': {fv}")
        print(f"    fee_details:")
        for fd in p.get('fee_details', []):
            print(f"      sheet={fd['sheet']}, amount={fd['amount_eur']}")
    else:
        print(f"\n  {wb}: NOT FOUND in parcels!")

inp_wb.close()
tmpl_wb.close()
