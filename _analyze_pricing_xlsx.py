# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\59571\Desktop\deutsch-app\舅妈网站\汇森国际-东欧COD报价20260331生效(5).xlsx"

def cell_repr(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return repr(v)

def print_row_iter(row_tuple, row_num):
    parts = []
    for cell in row_tuple:
        v = cell.value
        if v is not None:
            letter = get_column_letter(cell.column)
            parts.append(f"({letter}: {cell_repr(v)})")
    line = f"Row {row_num}: " + "  ".join(parts) if parts else f"Row {row_num}: (empty / all None)"
    print(line)

def dump_sheet_range(ws, r1, r2):
    for r in range(r1, r2 + 1):
        rows = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))
        if not rows:
            print(f"Row {r}: (no row data)")
            continue
        print_row_iter(rows[0], r)

def dump_sheet_all(ws):
    max_r = ws.max_row or 1
    for r in range(1, max_r + 1):
        rows = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))
        if not rows:
            print(f"Row {r}: (no row data)")
            continue
        print_row_iter(rows[0], r)

def main():
    # 1. 欧洲COD 1-85
    print("=" * 80)
    print("SHEET: 欧洲COD — rows 1 to 85 (non-None cells only)")
    print("=" * 80)
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb["欧洲COD"]
    dump_sheet_range(ws, 1, 85)
    wb.close()

    # 2. IOSS欧盟税率
    print()
    print("=" * 80)
    print("SHEET: IOSS欧盟税率 — ALL rows")
    print("=" * 80)
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb["IOSS欧盟税率"]
    dump_sheet_all(ws)
    wb.close()

    # 3. 偏远邮编
    print()
    print("=" * 80)
    print("SHEET: 偏远邮编 — header, first 5 data rows, row count")
    print("=" * 80)
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb["偏远邮编"]
    max_r = ws.max_row or 0
    print(f"Total rows in sheet (max_row): {max_r}")
    print()
    print("--- Header (row 1) ---")
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if rows:
        print_row_iter(rows[0], 1)
    print()
    print("--- First 5 data rows (rows 2–6) ---")
    for r in range(2, min(7, max_r + 1)):
        row_list = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))
        if row_list:
            print_row_iter(row_list[0], r)
    wb.close()

    # 4. 币种表 (leading space)
    print()
    print("=" * 80)
    print("SHEET:  币种表 — ALL rows")
    print("=" * 80)
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb[" 币种表"]
    dump_sheet_all(ws)
    wb.close()
    print()
    print("Done.")

if __name__ == "__main__":
    main()
