"""
导出引擎 - 按客户+科目生成Excel对账单
"""
import os
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import current_app
from app import db
from app.models import Order, OrderFee, FeeCategory, Customer, ExchangeRate


def _safe_filename(name):
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def _get_exchange_rate(bill_date):
    query = ExchangeRate.query.filter_by(from_currency='EUR', to_currency='CNY')
    if bill_date:
        query = query.filter(ExchangeRate.date <= bill_date)
    rate = query.order_by(ExchangeRate.date.desc()).first()
    return rate.rate if rate else None


def generate_export(customer_id, bill_period_str, category_ids=None):
    customer = Customer.query.get(customer_id)
    if not customer:
        raise ValueError('客户不存在')

    bp = None
    if bill_period_str:
        try:
            parts = bill_period_str.split('-')
            bp = date(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            raise ValueError('日期格式错误')

    query = Order.query.filter_by(customer_id=customer_id)
    if bp:
        query = query.filter_by(bill_period=bp)
    orders = query.order_by(Order.waybill_no).all()

    if not orders:
        raise ValueError('该客户在指定账期无订单数据')

    categories = []
    if category_ids:
        categories = FeeCategory.query.filter(FeeCategory.id.in_(category_ids)).all()
    else:
        categories = FeeCategory.query.all()

    exchange_rate = _get_exchange_rate(bp)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = '汇总'

    header_font = Font(name='等线', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    _write_summary_sheet(ws_summary, customer, orders, categories,
                         header_font, header_fill, header_align, thin_border, bp, exchange_rate)

    for cat in categories:
        cat_orders = []
        for order in orders:
            fee = OrderFee.query.filter_by(
                order_id=order.id, category_id=cat.id
            ).first()
            if fee:
                cat_orders.append((order, fee))

        if not cat_orders:
            continue

        ws = wb.create_sheet(title=cat.name[:31])
        _write_category_sheet(ws, cat, cat_orders,
                              header_font, header_fill, header_align, thin_border)

    export_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'exports')
    os.makedirs(export_dir, exist_ok=True)

    period_str = bp.strftime('%Y%m%d') if bp else 'all'
    filename = f'{period_str}-{_safe_filename(customer.name)}-对账单.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return filepath


def _write_summary_sheet(ws, customer, orders, categories,
                          header_font, header_fill, header_align, border, bp, exchange_rate):
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    period_str = bp.strftime('%Y-%m-%d') if bp else ''
    title_cell.value = f'{customer.name} 对账单 {period_str}'
    title_cell.font = Font(name='等线', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws['A3'] = '科目'
    ws['B3'] = '订单数'
    ws['C3'] = '总金额(EUR)'
    if exchange_rate:
        ws['D3'] = f'总金额(CNY) 汇率:{exchange_rate:.4f}'
    else:
        ws['D3'] = '总金额(CNY)'

    for col in range(1, 5):
        cell = ws.cell(3, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row = 4
    grand_total_eur = 0
    grand_total_cny = 0
    for cat in categories:
        count = 0
        total = 0
        for order in orders:
            fee = OrderFee.query.filter_by(
                order_id=order.id, category_id=cat.id
            ).first()
            if fee:
                count += 1
                amt = fee.override_amount if fee.override_amount is not None else fee.calculated_amount
                if amt is not None:
                    total += amt

        if count == 0:
            continue

        ws.cell(row, 1, cat.name).border = border
        ws.cell(row, 2, count).border = border
        ws.cell(row, 3, round(total, 2)).border = border
        ws.cell(row, 3).number_format = '#,##0.00'

        cny_total = round(total * exchange_rate, 2) if exchange_rate else None
        if cny_total is not None:
            ws.cell(row, 4, cny_total).border = border
            ws.cell(row, 4).number_format = '#,##0.00'
            grand_total_cny += cny_total
        else:
            ws.cell(row, 4, '-').border = border

        grand_total_eur += total
        row += 1

    ws.cell(row, 1, '合计').font = Font(bold=True)
    ws.cell(row, 1).border = border
    ws.cell(row, 2).border = border
    ws.cell(row, 3, round(grand_total_eur, 2)).font = Font(bold=True)
    ws.cell(row, 3).border = border
    ws.cell(row, 3).number_format = '#,##0.00'
    if exchange_rate:
        ws.cell(row, 4, round(grand_total_cny, 2)).font = Font(bold=True)
        ws.cell(row, 4).number_format = '#,##0.00'
    else:
        ws.cell(row, 4, '-')
    ws.cell(row, 4).border = border

    for col_letter in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 22


def _write_category_sheet(ws, category, cat_orders,
                           header_font, header_fill, header_align, border):
    headers = ['序号', '运单号', '目的国', '重量(kg)', '计费重(kg)',
               '代理金额', '计算金额', '最终金额', '备注']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for idx, (order, fee) in enumerate(cat_orders, 1):
        final = fee.override_amount if fee.override_amount is not None else fee.calculated_amount

        ws.cell(idx + 1, 1, idx).border = border
        ws.cell(idx + 1, 2, order.waybill_no).border = border
        ws.cell(idx + 1, 3, order.region.name if order.region else '-').border = border
        ws.cell(idx + 1, 4, order.actual_weight).border = border
        ws.cell(idx + 1, 5, order.charge_weight_head or order.charge_weight_tail).border = border
        ws.cell(idx + 1, 6, fee.input_amount).border = border
        ws.cell(idx + 1, 7, fee.calculated_amount).border = border
        ws.cell(idx + 1, 8, final).border = border
        ws.cell(idx + 1, 9, fee.notes or '').border = border

        for c in [6, 7, 8]:
            ws.cell(idx + 1, c).number_format = '#,##0.00'

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions['B'].width = 22
