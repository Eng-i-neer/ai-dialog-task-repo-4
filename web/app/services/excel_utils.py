"""
Unified Excel loading utility.
Handles both .xlsx (openpyxl) and .xls (xlrd) formats with a consistent
worksheet interface.
"""
import os


class XlrdSheetWrapper:
    """Wraps an xlrd sheet to provide an openpyxl-like cell access interface."""

    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols
        self.title = xlrd_sheet.name

    def cell(self, row, col):
        r = row - 1
        c = col - 1
        if r < 0 or c < 0 or r >= self._sheet.nrows or c >= self._sheet.ncols:
            return _EmptyCell()
        return _XlrdCell(self._sheet, r, c)


class _XlrdCell:
    def __init__(self, sheet, row, col):
        self.value = sheet.cell_value(row, col)
        if self.value == '':
            self.value = None


class _EmptyCell:
    value = None


class XlrdWorkbookWrapper:
    """Wraps an xlrd workbook to provide an openpyxl-like interface."""

    def __init__(self, xlrd_wb):
        self._wb = xlrd_wb
        self.sheetnames = xlrd_wb.sheet_names()

    def __getitem__(self, name):
        return XlrdSheetWrapper(self._wb.sheet_by_name(name))

    def close(self):
        self._wb.release_resources()


def load_excel(filepath, data_only=True, read_only=False):
    """
    Load an Excel file, auto-detecting format.
    Returns an object with .sheetnames, [sheet_name] access, and .close().
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        return XlrdWorkbookWrapper(wb)
    else:
        from openpyxl import load_workbook
        return load_workbook(filepath, data_only=data_only, read_only=read_only)
