"""
报价文件解析引擎 — 从报价 Excel 自动提取计价规则
支持 汇森国际 东欧COD报价 格式
"""
import re
from datetime import date
from openpyxl import load_workbook
from app import db
from app.models import (PricingVersion, PricingRule, FeeCategory, Region,
                         RemotePostcode)

CARGO_TYPE_MAP = {
    '普货': 'GS',
    '特货': 'SC',
    '敏感货': 'IC',
}

COUNTRY_CODE_RE = re.compile(r'[（(]\s*([A-Z]{2})\s*[）)]')
COD_MIN_RE = re.compile(r'最低收费\s*([\d.]+)\s*EU', re.IGNORECASE)
COD_RATE_RE = re.compile(r'([\d.]+)\s*%')

NON_COUNTRY_PREFIXES = ('海外仓', '偏远', '普货', '特货', '敏感货', '备注',
                          '注意', '说明', '操作', '费用', '上架')


def _extract_country(cell_value):
    """从 '波兰(PL)' / '意大利(IT)' 提取国家名和代码"""
    if not cell_value:
        return None, None
    text = str(cell_value).strip()
    m = COUNTRY_CODE_RE.search(text)
    if not m:
        return None, None
    code = m.group(1)
    name_raw = COUNTRY_CODE_RE.sub('', text).strip()
    return name_raw, code


def _find_region(name, code):
    """按 code 或 name 查找 Region，支持模糊匹配"""
    if code:
        r = Region.query.filter_by(code=code).first()
        if r:
            return r
    if name:
        r = Region.query.filter_by(name=name).first()
        if r:
            return r
        for region in Region.query.all():
            if name in region.name or region.name in name:
                return region
    return None


def _find_or_create_region(name, code):
    """查找 Region，不存在则自动创建"""
    region = _find_region(name, code)
    if region:
        return region
    region = Region(name=name, code=code, currency='EUR')
    db.session.add(region)
    db.session.flush()
    return region


def _find_category(code):
    return FeeCategory.query.filter_by(code=code).first()


def _parse_return_rule(text):
    """解析退件规则文本 -> return_ratio"""
    if not text:
        return 1.0
    t = str(text).strip()
    if '70%' in t:
        return 0.7
    return 1.0


def _parse_cod_text(text):
    """解析 COD 手续费文本 -> (rate, min_amount)"""
    if not text:
        return None, None
    t = str(text)
    rate_m = COD_RATE_RE.search(t)
    rate = float(rate_m.group(1)) / 100 if rate_m else 0.03
    min_m = COD_MIN_RE.search(t)
    min_amt = float(min_m.group(1)) if min_m else 1.5
    return rate, min_amt


def _is_country_row(a_val):
    """Check if column A contains a country name (has country code in parens)."""
    if not a_val:
        return False
    text = str(a_val).strip()
    if not text:
        return False
    for pfx in NON_COUNTRY_PREFIXES:
        if text.startswith(pfx):
            return False
    return bool(COUNTRY_CODE_RE.search(text))


def _is_cargo_row(c_val):
    """Check if column C contains a cargo type."""
    if not c_val:
        return False
    return str(c_val).strip() in CARGO_TYPE_MAP


def _scan_country_blocks(ws):
    """
    Scan the COD sheet and return structured country blocks.
    Each block: {name, code, carriers: [{carrier, rows: [{cargo, d, e, f}]}], return_text, cod_text}
    Handles merged cells by carrying forward E/F values within a country.
    """
    blocks = []
    current_block = None
    current_carrier = None
    country_tail_e = None
    country_tail_f = None

    for row in ws.iter_rows(min_row=4, max_row=80, values_only=False):
        vals = [c.value for c in row[:11]]
        row_num = row[0].row
        a_val = vals[0]
        b_val = vals[1] if len(vals) > 1 else None
        c_val = vals[2] if len(vals) > 2 else None
        d_val = vals[3] if len(vals) > 3 else None
        e_val = vals[4] if len(vals) > 4 else None
        f_val = vals[5] if len(vals) > 5 else None
        g_val = vals[6] if len(vals) > 6 else None
        h_val = vals[7] if len(vals) > 7 else None

        if _is_country_row(a_val):
            name, code = _extract_country(a_val)
            if current_block:
                blocks.append(current_block)

            current_block = {
                'name': name, 'code': code,
                'carriers': [],
                'return_text': str(g_val) if g_val else None,
                'cod_text': str(h_val) if h_val else None,
            }
            carrier_name = str(b_val).strip() if b_val else ''
            current_carrier = {'carrier': carrier_name, 'rows': []}
            current_block['carriers'].append(current_carrier)

            if e_val and isinstance(e_val, (int, float)):
                country_tail_e = float(e_val)
            if f_val and isinstance(f_val, (int, float)):
                country_tail_f = float(f_val)

            if _is_cargo_row(c_val):
                eff_e = float(e_val) if e_val and isinstance(e_val, (int, float)) else country_tail_e
                eff_f = float(f_val) if f_val and isinstance(f_val, (int, float)) else country_tail_f
                current_carrier['rows'].append({
                    'cargo': str(c_val).strip(),
                    'd': float(d_val) if d_val and isinstance(d_val, (int, float)) else None,
                    'e': eff_e,
                    'f': eff_f,
                })
            continue

        if not current_block:
            continue

        if b_val and not _is_cargo_row(c_val):
            text = str(b_val).strip()
            if text and not any(text.startswith(p) for p in NON_COUNTRY_PREFIXES):
                if COUNTRY_CODE_RE.search(str(a_val) if a_val else ''):
                    pass
                elif _is_cargo_row(vals[2] if len(vals) > 2 else None):
                    pass
                else:
                    continue

        if b_val and _is_cargo_row(c_val) and not a_val:
            carrier_name = str(b_val).strip()
            existing = [cr for cr in current_block['carriers'] if cr['carrier'] == carrier_name]
            if not existing:
                current_carrier = {'carrier': carrier_name, 'rows': []}
                current_block['carriers'].append(current_carrier)
                if e_val and isinstance(e_val, (int, float)):
                    country_tail_e = float(e_val)
                if f_val and isinstance(f_val, (int, float)):
                    country_tail_f = float(f_val)

        if _is_cargo_row(c_val) and current_carrier is not None:
            if e_val and isinstance(e_val, (int, float)):
                country_tail_e = float(e_val)
            if f_val and isinstance(f_val, (int, float)):
                country_tail_f = float(f_val)

            current_carrier['rows'].append({
                'cargo': str(c_val).strip(),
                'd': float(d_val) if d_val and isinstance(d_val, (int, float)) else None,
                'e': country_tail_e,
                'f': country_tail_f,
            })

    if current_block:
        blocks.append(current_block)

    return blocks


def preview_pricing_file(filepath):
    """解析报价文件，返回预览结果（不写入数据库）。"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = {
        'countries': [],
        'rules_by_category': {},
        'vat_updates': [],
        'remote_postcodes_count': 0,
        'currencies': [],
        'warnings': [],
    }

    _preview_cod_sheet(wb, result)
    _preview_ioss_sheet(wb, result)
    _preview_remote_sheet(wb, result)
    _preview_currency_sheet(wb, result)

    wb.close()
    return result


def _get_sheet(wb, candidates):
    """Find a sheet by trying multiple name candidates."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    for sn in wb.sheetnames:
        for c in candidates:
            if c in sn:
                return wb[sn]
    return None


def _preview_cod_sheet(wb, result):
    ws = _get_sheet(wb, ['欧洲COD', 'COD'])
    if not ws:
        result['warnings'].append('未找到"欧洲COD"Sheet')
        return

    blocks = _scan_country_blocks(ws)
    rules_count = {}

    seen_countries = set()
    for blk in blocks:
        key = blk['code'] or blk['name']
        if key not in seen_countries:
            result['countries'].append({'name': blk['name'], 'code': blk['code']})
            seen_countries.add(key)

        for carrier in blk['carriers']:
            for row_data in carrier['rows']:
                if row_data['d'] is not None:
                    rules_count['HEAD_FREIGHT'] = rules_count.get('HEAD_FREIGHT', 0) + 1
                if row_data['e'] is not None:
                    rules_count['TAIL_FREIGHT'] = rules_count.get('TAIL_FREIGHT', 0) + 1

        if blk['return_text']:
            rules_count['RETURN_FEE'] = rules_count.get('RETURN_FEE', 0) + 1
        if blk['cod_text']:
            rules_count['COD_FEE'] = rules_count.get('COD_FEE', 0) + 1

    rules_count['SHELF_FEE'] = 1
    rules_count['F_SURCHARGE'] = 3
    rules_count['REMOTE_FEE'] = 1

    result['rules_by_category'] = rules_count


def _preview_ioss_sheet(wb, result):
    ws = _get_sheet(wb, ['IOSS欧盟税率', 'IOSS', 'VAT'])
    if not ws:
        result['warnings'].append('未找到"IOSS欧盟税率"Sheet')
        return

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        country_name = str(row[1]).strip()
        vat = row[2] if len(row) > 2 else None
        if vat and isinstance(vat, (int, float)):
            result['vat_updates'].append({'country': country_name, 'vat_rate': vat})


def _preview_remote_sheet(wb, result):
    ws = _get_sheet(wb, ['偏远邮编', '偏远'])
    if not ws:
        result['warnings'].append('未找到"偏远邮编"Sheet')
        return
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        for cell in row:
            if cell and str(cell).strip():
                pc = str(cell).strip()
                if pc.isdigit() or (len(pc) <= 10 and pc.replace('-', '').replace(' ', '').isdigit()):
                    count += 1
    result['remote_postcodes_count'] = count


def _preview_currency_sheet(wb, result):
    ws = _get_sheet(wb, ['币种表', ' 币种表'])
    if not ws:
        return
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0] and row[1]:
            result['currencies'].append({
                'country': str(row[0]).strip(),
                'currency': str(row[1]).strip()
            })


def commit_pricing_file(filepath, version_name, effective_date,
                         expire_date=None, source_filename=None):
    """
    解析报价文件并写入数据库。
    创建 PricingVersion + PricingRules + RemotePostcodes，更新 Region VAT。
    Returns summary dict.
    """
    existing = PricingVersion.query.filter_by(name=version_name).first()
    if existing:
        db.session.delete(existing)
        db.session.flush()

    version = PricingVersion(
        name=version_name,
        effective_date=effective_date,
        expire_date=expire_date,
        source_file=source_filename or '',
    )
    db.session.add(version)
    db.session.flush()

    wb = load_workbook(filepath, read_only=True, data_only=True)
    summary = {'rules_created': 0, 'vat_updated': 0, 'postcodes_created': 0}

    summary['rules_created'] += _commit_cod_sheet(wb, version)
    summary['vat_updated'] += _commit_ioss_sheet(wb)
    summary['postcodes_created'] += _commit_remote_sheet(wb, version)

    wb.close()
    db.session.commit()

    summary['version_id'] = version.id
    summary['version_name'] = version.name
    return summary


def _commit_cod_sheet(wb, version):
    ws = _get_sheet(wb, ['欧洲COD', 'COD'])
    if not ws:
        return 0

    cat_head = _find_category('HEAD_FREIGHT')
    cat_tail = _find_category('TAIL_FREIGHT')
    cat_return = _find_category('RETURN_FEE')
    cat_cod = _find_category('COD_FEE')
    cat_shelf = _find_category('SHELF_FEE')
    cat_f = _find_category('F_SURCHARGE')
    cat_remote = _find_category('REMOTE_FEE')

    blocks = _scan_country_blocks(ws)
    rules_created = 0

    for blk in blocks:
        region = _find_or_create_region(blk['name'], blk['code'])

        for carrier in blk['carriers']:
            carrier_name = carrier['carrier']
            for row_data in carrier['rows']:
                cargo_code = CARGO_TYPE_MAP.get(row_data['cargo'])
                if not cargo_code:
                    continue

                if row_data['d'] is not None and cat_head:
                    rule = PricingRule(
                        version_id=version.id,
                        category_id=cat_head.id,
                        region_id=region.id,
                        cargo_type=cargo_code,
                        rule_type='per_kg',
                    )
                    rule.set_params({
                        'rate_per_kg': row_data['d'],
                        'carrier': carrier_name,
                    })
                    db.session.add(rule)
                    rules_created += 1

                if row_data['e'] is not None and cat_tail:
                    rule = PricingRule(
                        version_id=version.id,
                        category_id=cat_tail.id,
                        region_id=region.id,
                        cargo_type=cargo_code,
                        rule_type='first_extra',
                    )
                    rule.set_params({
                        'first_weight': 2,
                        'first_price': row_data['e'],
                        'extra_per_kg': row_data['f'] or 0,
                        'carrier': carrier_name,
                    })
                    db.session.add(rule)
                    rules_created += 1

        if blk['return_text'] and cat_return:
            return_ratio = _parse_return_rule(blk['return_text'])
            first_carrier = blk['carriers'][0] if blk['carriers'] else None
            first_gs = None
            if first_carrier:
                for rd in first_carrier['rows']:
                    if rd['cargo'] == '普货':
                        first_gs = rd
                        break

            rule = PricingRule(
                version_id=version.id,
                category_id=cat_return.id,
                region_id=region.id,
                cargo_type=None,
                rule_type='first_extra',
            )
            rule.set_params({
                'first_weight': 2,
                'first_price': first_gs['e'] if first_gs and first_gs['e'] else 0,
                'extra_per_kg': first_gs['f'] if first_gs and first_gs['f'] else 0,
                'return_ratio': return_ratio,
            })
            db.session.add(rule)
            rules_created += 1

        if blk['cod_text'] and cat_cod:
            rate, min_amt = _parse_cod_text(blk['cod_text'])
            if rate is not None:
                rule = PricingRule(
                    version_id=version.id,
                    category_id=cat_cod.id,
                    region_id=region.id,
                    cargo_type=None,
                    rule_type='percentage',
                )
                rule.set_params({'rate': rate, 'min_amount': min_amt or 1.5})
                db.session.add(rule)
                rules_created += 1

    if cat_shelf:
        rule = PricingRule(
            version_id=version.id,
            category_id=cat_shelf.id,
            region_id=None,
            cargo_type=None,
            rule_type='fixed',
        )
        rule.set_params({'amount': 1.5})
        db.session.add(rule)
        rules_created += 1

    if cat_f:
        for cargo, amount, currency in [
            ('SC', 1.5, 'EUR'),
            ('IC', 2.0, 'EUR'),
            ('F手表', 30.0, 'CNY'),
        ]:
            rule = PricingRule(
                version_id=version.id,
                category_id=cat_f.id,
                region_id=None,
                cargo_type=cargo,
                rule_type='fixed',
            )
            rule.set_params({'amount': amount, 'currency': currency})
            db.session.add(rule)
            rules_created += 1

    if cat_remote:
        rule = PricingRule(
            version_id=version.id,
            category_id=cat_remote.id,
            region_id=None,
            cargo_type=None,
            rule_type='tiered',
        )
        rule.set_params({
            'zones': {
                '西西里岛': {'type': 'per_kg', 'amount': 0.8},
                '撒丁岛': {'type': 'per_kg', 'amount': 0.8},
                '卡拉布里亚': {'type': 'per_kg', 'amount': 0.8},
                '其他岛屿': {'type': 'per_piece', 'amount': 21.0},
            }
        })
        db.session.add(rule)
        rules_created += 1

    return rules_created


def _commit_ioss_sheet(wb):
    ws = _get_sheet(wb, ['IOSS欧盟税率', 'IOSS', 'VAT'])
    if not ws:
        return 0

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        country_name = str(row[1]).strip()
        vat = row[2] if len(row) > 2 else None
        if vat is None or not isinstance(vat, (int, float)):
            continue

        region = _find_region(country_name, None)
        if region and region.vat_rate != vat:
            region.vat_rate = float(vat)
            updated += 1

    return updated


def _commit_remote_sheet(wb, version):
    ws = _get_sheet(wb, ['偏远邮编', '偏远'])
    if not ws:
        return 0

    header = None
    created = 0
    batch = []

    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue

        if row[0] and str(row[0]).strip():
            pc = str(row[0]).strip()
            country_col = row[1] if len(row) > 1 else None
            country = str(country_col).strip() if country_col else '克罗地亚'
            if pc.isdigit() or (len(pc) <= 10 and pc.replace('-', '').replace(' ', '').isdigit()):
                batch.append(RemotePostcode(
                    version_id=version.id,
                    postcode=pc,
                    country=country,
                    zone=country,
                    surcharge_type='per_kg',
                    surcharge_amount=0.8,
                ))

        for col_pc, col_zone, default_country in [(3, 4, '意大利'), (5, 6, '意大利'), (7, 8, '意大利')]:
            if len(row) <= col_pc:
                continue
            pc_val = row[col_pc]
            if not pc_val:
                continue
            pc = str(pc_val).strip()
            if not pc or not (pc.isdigit() or pc.replace('-', '').replace(' ', '').isdigit()):
                continue

            zone_val = row[col_zone] if len(row) > col_zone and row[col_zone] else ''
            zone = str(zone_val).strip() if zone_val else ''

            if '西西里' in zone:
                stype, samt = 'per_kg', 0.8
            elif '撒丁' in zone:
                stype, samt = 'per_kg', 0.8
            elif '岛屿' in zone or '威尼斯' in zone:
                stype, samt = 'per_piece', 21.0
            else:
                stype, samt = 'per_kg', 0.8

            batch.append(RemotePostcode(
                version_id=version.id,
                postcode=pc,
                country=default_country,
                zone=zone or default_country,
                surcharge_type=stype,
                surcharge_amount=samt,
            ))

        if len(batch) >= 500:
            db.session.add_all(batch)
            created += len(batch)
            batch = []

    if batch:
        db.session.add_all(batch)
        created += len(batch)

    return created
